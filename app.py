# app.py
import os
import logging
import uuid
import secrets
from datetime import datetime, date, timedelta
from flask import Flask, jsonify, request, render_template, send_from_directory, make_response, session, redirect, url_for
from dotenv import load_dotenv
from supabase import create_client, Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from werkzeug.security import generate_password_hash, check_password_hash

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', 'rs-co-default-luxury-secret-key')

# Vercel serverless compatibility - adjust session cookie settings
if os.getenv('VERCEL'):
    # For Vercel, use looser cookie settings since it's behind HTTPS
    app.config['SESSION_COOKIE_SECURE'] = False
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
else:
    # Local development
    app.config['SESSION_COOKIE_SECURE'] = True
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# CSRF Protection
def generate_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

def validate_csrf_token(token):
    return 'csrf_token' in session and secrets.compare_digest(session['csrf_token'], token)

app.jinja_env.globals['csrf_token'] = lambda: generate_csrf_token()

# Simple Rate Limiter
rate_limit_storage = {}

def check_rate_limit(identifier, max_requests=5, window_seconds=60):
    current_time = datetime.now()
    if identifier not in rate_limit_storage:
        rate_limit_storage[identifier] = []
    
    # Remove old requests outside the time window
    rate_limit_storage[identifier] = [
        req_time for req_time in rate_limit_storage[identifier]
        if (current_time - req_time).total_seconds() < window_seconds
    ]
    
    # Check if limit exceeded
    if len(rate_limit_storage[identifier]) >= max_requests:
        return False
    
    # Add current request
    rate_limit_storage[identifier].append(current_time)
    return True

# Supabase Initialization with Fallback
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = None
SUPABASE_KEY_NAME = None
for key_name in ['SUPABASE_KEY', 'SUPABASE_ANON_KEY', 'SUPABASE_SERVICE_ROLE_KEY', 'SUPABASE_SERVICE_ROLE', 'SUPABASE_SERVICE_KEY', 'SUPABASE_API_KEY']:
    value = os.getenv(key_name)
    if value:
        SUPABASE_KEY = value
        SUPABASE_KEY_NAME = key_name
        break
SUPABASE_CLIENT = None
SUPABASE_CONNECTED = False

# ─── FILE-BASED PERSISTENCE ─────────────────────────────────────────────────
import json

DB_FILE = os.path.join(os.path.dirname(__file__), 'data', 'db.json')

# Keys that are user-created and must be persisted across restarts
PERSISTED_KEYS = ['manual_billing', 'expenses', 'rmc_orders', 'rmc_order_items',
                   'rmc_invoices', 'rmc_payments', 'rmc_clients', 'submissions', 'newsletter',
                   'rmc_users', 'client_billing_summaries', 'supplier_payment_summaries']
BILLING_PDF_DIR = os.path.join(os.path.dirname(__file__), 'data', 'billing_pdfs')

def save_db():
    """Persist dynamic user data to disk."""
    try:
        os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
        payload = {k: MOCK_DATABASE.get(k, []) for k in PERSISTED_KEYS}
        with open(DB_FILE, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"save_db error: {e}")

def load_db():
    """Load persisted user data from disk into MOCK_DATABASE."""
    if not os.path.exists(DB_FILE):
        return
    try:
        with open(DB_FILE, 'r', encoding='utf-8') as f:
            payload = json.load(f)
        for k in PERSISTED_KEYS:
            if k in payload:
                MOCK_DATABASE[k] = payload[k]
        logger.info(f"Loaded persisted data: {', '.join(f'{k}={len(MOCK_DATABASE.get(k,[]))}' for k in PERSISTED_KEYS if isinstance(MOCK_DATABASE.get(k), list))}")
    except Exception as e:
        logger.error(f"load_db error: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# Mock In-Memory Database for local preview/fallback
MOCK_DATABASE = {
    "founder": {
        "name": "S Maghesh",
        "role": "Founder & Chief Executive Officer",
        "bio": "With over 25 years of pioneering leadership in infrastructure development, S Maghesh has steered RS & CO from a local paving contractor to an elite infrastructure enterprise. His engineering acumen paired with an uncompromising commitment to premium execution has redefined highway construction standards across the region.",
        "photo_url": "/static/images/founder_ceo.jpg?v=4",
        "vision": "To build foundational pathways that endure for generations, fusing cutting-edge materials engineering with timeless architectural luxury and safety."
    },
    "rmc_users": [
        {
            "id": "user-1",
            "username": "admin",
            "email": "admin@rsandco.com",
            "password_hash": generate_password_hash("admin123"),
            "full_name": "System Administrator",
            "role": "admin",
            "phone": "+91 98765 43211",
            "is_active": True
        },
        {
            "id": "user-2",
            "username": "manager",
            "email": "manager@rsandco.com",
            "password_hash": generate_password_hash("manager123"),
            "full_name": "Operations Manager",
            "role": "manager",
            "phone": "+91 98765 43212",
            "is_active": True
        },
        {
            "id": "user-3",
            "username": "client",
            "email": "client@test.com",
            "password_hash": generate_password_hash("client123"),
            "full_name": "Test Client",
            "role": "client",
            "phone": "+91 98765 43213",
            "is_active": True
        }
    ],
    "manual_billing": [
        {
            "id": "bill-1",
            "company_name": "ABC Construction Ltd",
            "phone": "9876543210",
            "address": "123 Main St, Chennai",
            "date": "2025-12-11",
            "time": "10:30 AM",
            "vehicle_number": "TN-01-AB-1234",
            "driver_name": "R Kumar",
            "grade": "M25",
            "cubic_meters": 5.5,
            "rate_per_cubic": 4500,
            "total_amount": 24750,
            "advance_amount": 10000,
            "balance_amount": 14750,
            "billing_status": "unbilled",
            "billing_summary_id": None,
            "loading_time": "09:45 AM",
            "unloading_time": "10:30 AM",
            "created_at": "2025-12-11T10:30:00"
        },
        {
            "id": "bill-2",
            "company_name": "XYZ Builders",
            "phone": "9876543211",
            "address": "456 Oak Ave, Chennai",
            "date": "2025-12-10",
            "time": "02:15 PM",
            "vehicle_number": "TN-02-CD-5678",
            "driver_name": "S Raj",
            "grade": "M30",
            "cubic_meters": 8.0,
            "rate_per_cubic": 4800,
            "total_amount": 38400,
            "advance_amount": 15000,
            "balance_amount": 23400,
            "billing_status": "unbilled",
            "billing_summary_id": None,
            "loading_time": "01:30 PM",
            "unloading_time": "02:15 PM",
            "created_at": "2025-12-10T14:15:00"
        }
    ],
    "client_billing_summaries": [],
    "supplier_payment_summaries": [],
    "expenses": [],
    "plants": [
        {
            "id": "plant-1",
            "name": "Golden Falcon Hotmix Plant",
            "type": "Hotmix",
            "capacity": "160 TPH",
            "status": "Active",
            "live_metrics": {
                "temperature": "165°C",
                "hourly_output": "142 tons",
                "fuel_efficiency": "94%",
                "active_recipe": "Superpave PG 76-22"
            },
            "last_updated": "Just now"
        },
        {
            "id": "plant-2",
            "name": "R. Sundaram & Co. RMC Plant",
            "type": "RMC",
            "capacity": "90 m³/hr",
            "status": "Active",
            "live_metrics": {
                "active_mix": "M25 (High Durability)",
                "current_slump": "120mm",
                "silo_levels": {"cement": "84%", "flyash": "72%"},
                "water_cement_ratio": "0.42"
            },
            "address": "Quarry Rd, Tiruneermalai, Chennai, Tamil Nadu 600132",
            "map_url": "https://maps.app.goo.gl/E94H9iCz8hXbowjo9",
            "last_updated": "Just now"
        },
        {
            "id": "plant-3",
            "name": "Imperial Hotmix Plant 2",
            "type": "Hotmix",
            "capacity": "120 TPH",
            "status": "Maintenance",
            "live_metrics": {
                "temperature": "45°C",
                "hourly_output": "0 tons",
                "maintenance_reason": "Scheduled burner nozzle calibration",
                "estimated_resume": "2 hours"
            },
            "last_updated": "Just now"
        }
    ],
    "projects": [
        {
            "id": "project-1",
            "name": "The Grand Horizon Expressway",
            "description": "An elite 8-lane concrete highway connecting major industrial corridors, integrating automated tollways and sustainable green zones.",
            "progress": 85,
            "image_url": "https://images.unsplash.com/photo-1541888946425-d81bb19240f5?auto=format&fit=crop&q=80&w=800",
            "status": "Ongoing"
        },
        {
            "id": "project-2",
            "name": "Royal Vista Bridge & Flyover",
            "description": "A majestic pre-stressed concrete flyover utilizing premium grade M45 concrete and architectural LED lighting accents.",
            "progress": 100,
            "image_url": "https://images.unsplash.com/photo-1513828729170-d62b358730eb?auto=format&fit=crop&q=80&w=800",
            "status": "Completed"
        },
        {
            "id": "project-3",
            "name": "Aurelia Boulevard Corridor",
            "description": "Urban revitalization paving project featuring advanced noise-reducing stone mastic asphalt and luxury pedestrian pathways.",
            "progress": 30,
            "image_url": "https://images.unsplash.com/photo-1504307651254-35680f356dfd?auto=format&fit=crop&q=80&w=800",
            "status": "Ongoing"
        }
    ],
    "submissions": [],
    "newsletter": [],
    # RMC Module Mock Data
    "rmc_concrete_grades": [
        {"id": "grade-1", "grade_code": "M10", "grade_name": "M10 Grade", "description": "Standard concrete for non-structural applications", "price_per_cubic_meter": 4500.00, "is_active": True},
        {"id": "grade-2", "grade_code": "M15", "grade_name": "M15 Grade", "description": "Light duty concrete for pathways and floors", "price_per_cubic_meter": 4800.00, "is_active": True},
        {"id": "grade-3", "grade_code": "M20", "grade_name": "M20 Grade", "description": "Standard grade for residential construction", "price_per_cubic_meter": 5200.00, "is_active": True},
        {"id": "grade-4", "grade_code": "M25", "grade_name": "M25 Grade", "description": "High durability concrete for commercial buildings", "price_per_cubic_meter": 5600.00, "is_active": True},
        {"id": "grade-5", "grade_code": "M30", "grade_name": "M30 Grade", "description": "Premium grade for heavy structural applications", "price_per_cubic_meter": 6100.00, "is_active": True},
        {"id": "grade-6", "grade_code": "M35", "grade_name": "M35 Grade", "description": "Ultra-high strength for specialized infrastructure", "price_per_cubic_meter": 6800.00, "is_active": True},
        {"id": "grade-7", "grade_code": "M40", "grade_name": "M40 Grade", "description": "Elite grade for critical infrastructure projects", "price_per_cubic_meter": 7500.00, "is_active": True}
    ],
    "rmc_clients": [
        {
            "id": "client-1",
            "company_name": "Apex Construction Ltd",
            "contact_person": "Rajesh Kumar",
            "email": "rajesh@apexconstruction.com",
            "phone": "+91 98765 43210",
            "billing_address": "123 Industrial Area, Chennai, Tamil Nadu 600001",
            "shipping_address": "123 Industrial Area, Chennai, Tamil Nadu 600001",
            "gst_number": "33AABCU9603R1ZN",
            "credit_limit": 5000000.00,
            "current_balance": 0.00,
            "is_active": True
        }
    ],
    "rmc_orders": [
        {
            "id": "order-1",
            "order_number": "RMC-20260610-0001",
            "user_id": "user-3",
            "customer_name": "Test Customer",
            "phone": "+91 98765 43213",
            "order_date": "2026-06-10T10:00:00",
            "delivery_date": "2026-06-15",
            "delivery_address": "123 Test Street, Chennai",
            "status": "Approved",
            "total_amount": 15000.00,
            "notes": "Test order",
            "items": [
                {
                    "grade_id": "grade-1",
                    "quantity": 5,
                    "unit_price": 3000.00,
                    "subtotal": 15000.00
                }
            ],
            "approved_by": "admin",
            "approved_at": "2026-06-10T10:30:00"
        }
    ],
    "rmc_order_items": [
        {
            "id": "item-1",
            "order_id": "order-1",
            "grade_id": "grade-1",
            "quantity_cubic_meters": 5,
            "unit_price": 3000.00,
            "subtotal": 15000.00
        }
    ],
    "rmc_invoices": [
        {
            "id": "invoice-1",
            "invoice_number": "INV-20260610-0001",
            "order_id": "order-1",
            "invoice_date": "2026-06-10",
            "due_date": "2026-07-10",
            "subtotal": 15000.00,
            "tax_amount": 0,
            "total_amount": 15000.00,
            "status": "Pending",
            "notes": "Auto-generated invoice for order RMC-20260610-0001"
        }
    ],
    "rmc_payments": []
}

if SUPABASE_URL and SUPABASE_KEY:
    try:
        SUPABASE_CLIENT = create_client(SUPABASE_URL, SUPABASE_KEY)
        SUPABASE_CONNECTED = True
        logger.info(f"Connected to Supabase using key env var: {SUPABASE_KEY_NAME}")
    except Exception as e:
        SUPABASE_CONNECTED = False
        logger.error(f"Error initializing Supabase client: {e}. Falling back to mock database.")
else:
    logger.warning("Supabase credentials not found. App running in LOCAL MOCK DATABASE mode.")

# Keep the local data available for explicit local/mock mode.  In particular,
# do not lose locally created records merely because cloud credentials exist
# but the cloud service is temporarily unavailable.
load_db()

@app.route('/api/supabase-status', methods=['GET'])
def supabase_status():
    return jsonify({
        "connected": bool(SUPABASE_CONNECTED),
        "mode": "supabase" if SUPABASE_CONNECTED else "mock",
        "supabase_url_present": bool(SUPABASE_URL),
        "supabase_key_name": SUPABASE_KEY_NAME,
        "supabase_url": SUPABASE_URL or ""
    })

# ----------------- Frontend Routes -----------------
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/login')
def login():
    return render_template('login.html')

@app.route('/register')
def register():
    return render_template('register.html')

@app.route('/robots.txt')
def robots_txt():
    return send_from_directory('static', 'robots.txt')

@app.route('/static/videos/<path:filename>')
def serve_video(filename):
    """Serve video files with byte-range support for mobile/Safari streaming."""
    import os
    from flask import Response, request as flask_request
    video_path = os.path.join(app.static_folder, 'videos', filename)
    if not os.path.exists(video_path):
        return '', 404

    file_size = os.path.getsize(video_path)
    range_header = flask_request.headers.get('Range', None)

    if range_header:
        # Parse Range header: bytes=start-end
        byte_range = range_header.replace('bytes=', '').split('-')
        start = int(byte_range[0])
        end = int(byte_range[1]) if byte_range[1] else file_size - 1
        length = end - start + 1

        def generate_chunk():
            with open(video_path, 'rb') as f:
                f.seek(start)
                remaining = length
                chunk_size = 65536  # 64KB chunks
                while remaining > 0:
                    data = f.read(min(chunk_size, remaining))
                    if not data:
                        break
                    remaining -= len(data)
                    yield data

        resp = Response(
            generate_chunk(),
            status=206,
            mimetype='video/mp4',
            direct_passthrough=True
        )
        resp.headers['Content-Range'] = f'bytes {start}-{end}/{file_size}'
        resp.headers['Accept-Ranges'] = 'bytes'
        resp.headers['Content-Length'] = length
        resp.headers['Cache-Control'] = 'public, max-age=86400'
        return resp
    else:
        # Full file response
        resp = send_from_directory(
            os.path.join(app.static_folder, 'videos'),
            filename,
            mimetype='video/mp4'
        )
        resp.headers['Accept-Ranges'] = 'bytes'
        resp.headers['Content-Length'] = file_size
        resp.headers['Cache-Control'] = 'public, max-age=86400'
        return resp


@app.route('/sitemap.xml')
def sitemap_xml():
    sitemap = '''<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
    <url>
        <loc>https://rsandco.com/</loc>
        <lastmod>2026-06-10</lastmod>
        <changefreq>weekly</changefreq>
        <priority>1.0</priority>
    </url>
    <url>
        <loc>https://rsandco.com/#about</loc>
        <lastmod>2026-06-10</lastmod>
        <changefreq>monthly</changefreq>
        <priority>0.8</priority>
    </url>
    <url>
        <loc>https://rsandco.com/#plants-section</loc>
        <lastmod>2026-06-10</lastmod>
        <changefreq>weekly</changefreq>
        <priority>0.8</priority>
    </url>
    <url>
        <loc>https://rsandco.com/#projects</loc>
        <lastmod>2026-06-10</lastmod>
        <changefreq>weekly</changefreq>
        <priority>0.8</priority>
    </url>
    <url>
        <loc>https://rsandco.com/#contact</loc>
        <lastmod>2026-06-10</lastmod>
        <changefreq>monthly</changefreq>
        <priority>0.6</priority>
    </url>
</urlset>'''
    response = make_response(sitemap)
    response.headers['Content-Type'] = 'application/xml'
    return response

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/billing-expense')
def billing_expense():
    # Role-based access control - only admin and manager can access
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # Check user role from session
    user_role = session.get('role')
    
    if user_role not in ['admin', 'manager']:
        return jsonify({"error": "Access denied. Admin or Manager role required."}), 403
    
    return render_template('billing_expense.html')

@app.route('/supplier-payment-summary')
def supplier_payment_summary_page():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session.get('role') not in ['admin', 'manager']:
        return jsonify({"error": "Access denied. Admin or Manager role required."}), 403
    return render_template('supplier_payment_summary.html')

@app.route('/invoice-letterpad')
def invoice_letterpad():
    return render_template('invoice_letterpad.html')

# ----------------- Authentication Routes -----------------
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    # Rate limiting based on IP address
    client_ip = request.remote_addr
    if not check_rate_limit(f"login_{client_ip}", max_requests=5, window_seconds=60):
        return jsonify({"error": "Too many login attempts. Please try again later."}), 429
    
    data = request.json
    username_or_email = data.get('username') or data.get('email')
    password = data.get('password')
    
    if not username_or_email or not password:
        return jsonify({"error": "Missing username/email or password"}), 400
    
    if SUPABASE_CLIENT:
        try:
            # Query user from Supabase
            response = SUPABASE_CLIENT.table("rmc_users").select("*").or_(f"username.eq.{username_or_email},email.eq.{username_or_email}").execute()
            
            if not response.data:
                return jsonify({"error": "Invalid credentials"}), 401
            
            user = response.data[0]
            
            if not user['is_active']:
                return jsonify({"error": "Account is inactive"}), 401
            
            if not check_password_hash(user['password_hash'], password):
                return jsonify({"error": "Invalid credentials"}), 401
            
            # Set session
            session['user_id'] = str(user['id'])
            session['username'] = user['username']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            
            return jsonify({
                "success": True,
                "user": {
                    "id": str(user['id']),
                    "username": user['username'],
                    "email": user['email'],
                    "full_name": user['full_name'],
                    "role": user['role']
                }
            })
        except Exception as e:
            logger.error(f"Supabase login error: {e}")
            return jsonify({"error": str(e)}), 500
    
    # Mock login
    users = MOCK_DATABASE.get("rmc_users", [])
    user = None
    for u in users:
        if u['username'] == username_or_email or u['email'] == username_or_email:
            user = u
            break
    
    if not user:
        return jsonify({"error": "Invalid credentials"}), 401
    
    if not user['is_active']:
        return jsonify({"error": "Account is inactive"}), 401
    
    if not check_password_hash(user['password_hash'], password):
        return jsonify({"error": "Invalid credentials"}), 401
    
    # Set session
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['role'] = user['role']
    session['full_name'] = user['full_name']
    
    return jsonify({
        "success": True,
        "user": {
            "id": user['id'],
            "username": user['username'],
            "email": user['email'],
            "full_name": user['full_name'],
            "role": user['role']
        },
        "mode": "mock"
    })

@app.route('/api/auth/register', methods=['POST'])
def api_register():
    # Rate limiting based on IP address
    client_ip = request.remote_addr
    if not check_rate_limit(f"register_{client_ip}", max_requests=3, window_seconds=3600):
        return jsonify({"error": "Too many registration attempts. Please try again later."}), 429
    
    data = request.json
    username = data.get('username')
    email = data.get('email')
    password = data.get('password')
    full_name = data.get('full_name')
    phone = data.get('phone')
    
    if not all([username, email, password, full_name]):
        return jsonify({"error": "Missing required fields"}), 400
    
    # Check if username or email already exists
    if SUPABASE_CLIENT:
        try:
            existing = SUPABASE_CLIENT.table("rmc_users").select("*").or_(f"username.eq.{username},email.eq.{email}").execute()
            if existing.data:
                return jsonify({"error": "Username or email already exists"}), 400
            
            # Create user
            password_hash = generate_password_hash(password)
            user_response = SUPABASE_CLIENT.table("rmc_users").insert({
                "username": username,
                "email": email,
                "password_hash": password_hash,
                "full_name": full_name,
                "role": "client",
                "phone": phone
            }).execute()
            
            return jsonify({"success": True, "user_id": str(user_response.data[0]['id'])})
        except Exception as e:
            logger.error(f"Supabase registration error: {e}")
            return jsonify({"error": str(e)}), 500
    
    # Mock registration
    users = MOCK_DATABASE.get("rmc_users", [])
    for u in users:
        if u['username'] == username or u['email'] == email:
            return jsonify({"error": "Username or email already exists"}), 400
    
    new_user = {
        "id": f"user-{uuid.uuid4().hex[:8]}",
        "username": username,
        "email": email,
        "password_hash": generate_password_hash(password),
        "full_name": full_name,
        "role": "client",
        "phone": phone or "",
        "is_active": True,
        "created_at": datetime.utcnow().isoformat()
    }
    MOCK_DATABASE["rmc_users"].append(new_user)
    save_db()   # ← persist so login survives server restart
    
    return jsonify({"success": True, "user_id": new_user['id'], "mode": "mock"})

@app.route('/api/auth/me', methods=['GET'])
def api_me():
    if 'user_id' not in session:
        return jsonify({"error": "Not authenticated"}), 401

    # Look up full user record to get email, phone, created_at
    uid = session['user_id']
    full_user = next((u for u in MOCK_DATABASE.get('rmc_users', []) if str(u.get('id')) == str(uid)), None)

    return jsonify({
        "user": {
            "id":         session['user_id'],
            "username":   session['username'],
            "role":       session['role'],
            "full_name":  session['full_name'],
            "email":      full_user.get('email', '')      if full_user else '',
            "phone":      full_user.get('phone', '')      if full_user else '',
            "created_at": full_user.get('created_at', '') if full_user else '',
        }
    })


@app.route('/rmc-portal')
def rmc_portal():
    # Check if user is logged in
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if user is a client
    if session.get('role') != 'client':
        return redirect(url_for('login'))
    
    return render_template('rmc_portal.html')

@app.route('/rmc-admin')
def rmc_admin():
    # Check if user is logged in
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if user is admin or manager
    if session.get('role') not in ['admin', 'manager']:
        return redirect(url_for('login'))
    
    return render_template('rmc_admin.html')

# ----------------- API Endpoints -----------------

@app.route('/api/founder', methods=['GET'])
def get_founder():
    if SUPABASE_CLIENT:
        try:
            response = SUPABASE_CLIENT.table("founder_ceo").select("*").limit(1).execute()
            if response.data:
                data = response.data[0]
                data["photo_url"] = "/static/images/founder_ceo.jpg?v=4"
                return jsonify(data)
        except Exception as e:
            logger.error(f"Supabase founder query error: {e}")
    
    # Return mock data as fallback
    return jsonify(MOCK_DATABASE["founder"])


@app.route('/api/plants', methods=['GET'])
def get_plants():
    if SUPABASE_CLIENT:
        try:
            response = SUPABASE_CLIENT.table("plants").select("*").execute()
            if response.data:
                return jsonify(response.data)
        except Exception as e:
            logger.error(f"Supabase plants query error: {e}")
            
    return jsonify(MOCK_DATABASE["plants"])


@app.route('/api/projects', methods=['GET'])
def get_projects():
    if SUPABASE_CLIENT:
        try:
            response = SUPABASE_CLIENT.table("projects").select("*").execute()
            if response.data:
                return jsonify(response.data)
        except Exception as e:
            logger.error(f"Supabase projects query error: {e}")
            
    return jsonify(MOCK_DATABASE["projects"])


@app.route('/api/contact', methods=['POST'])
def submit_contact():
    data = request.json or {}
    name = data.get('name')
    email = data.get('email')
    message = data.get('message')
    
    if not name or not email or not message:
        return jsonify({"error": "Missing required fields"}), 400
        
    if SUPABASE_CLIENT:
        try:
            SUPABASE_CLIENT.table("contact_submissions").insert({
                "name": name,
                "email": email,
                "message": message
            }).execute()
            return jsonify({"success": True, "message": "Message received via Supabase."})
        except Exception as e:
            logger.error(f"Supabase contact submission error: {e}")
            
    # Mock fallback submission
    MOCK_DATABASE["submissions"].append(data)
    logger.info(f"Local Mock Submission received: {data}")
    return jsonify({"success": True, "message": "Message submitted locally (Mock Mode)."})


@app.route('/api/subscribe', methods=['POST'])
def submit_subscribe():
    data = request.json or {}
    email = data.get('email')
    
    if not email:
        return jsonify({"error": "Email is required"}), 400
        
    if SUPABASE_CLIENT:
        try:
            SUPABASE_CLIENT.table("newsletter_subscribers").insert({
                "email": email
            }).execute()
            return jsonify({"success": True, "message": "Subscribed via Supabase."})
        except Exception as e:
            logger.error(f"Supabase newsletter subscription error: {e}")
            
    MOCK_DATABASE["newsletter"].append(email)
    logger.info(f"Local Mock Newsletter Subscription: {email}")
    return jsonify({"success": True, "message": "Subscribed locally (Mock Mode)."})


# ----------------- Protected Admin API Routes -----------------
# For mock mode or authentication bypass, we read standard payload fields. 
# Real Supabase admin uses JS SDK direct inserts with RLS, but these endpoints support alternate clients.

@app.route('/api/admin/update-plant', methods=['POST'])
def update_plant():
    data = request.json or {}
    plant_id = data.get('id')
    status = data.get('status')
    live_metrics = data.get('live_metrics')
    
    if not plant_id or not status:
        return jsonify({"error": "Missing plant ID or status"}), 400
        
    if SUPABASE_CLIENT:
        try:
            payload = {"status": status}
            if live_metrics is not None:
                payload["live_metrics"] = live_metrics
            payload["last_updated"] = "now()"
            
            SUPABASE_CLIENT.table("plants").update(payload).eq("id", plant_id).execute()
            return jsonify({"success": True})
        except Exception as e:
            logger.error(f"Supabase admin plant update error: {e}")
            return jsonify({"error": str(e)}), 500
            
    # Mock update
    for plant in MOCK_DATABASE["plants"]:
        if plant["id"] == plant_id:
            plant["status"] = status
            if live_metrics is not None:
                plant["live_metrics"] = live_metrics
            plant["last_updated"] = "Just now"
            return jsonify({"success": True, "mode": "mock"})
            
    return jsonify({"error": "Plant not found"}), 404


@app.route('/api/admin/update-project', methods=['POST'])
def update_project():
    data = request.json or {}
    project_id = data.get('id')
    progress = data.get('progress')
    
    if not project_id or progress is None:
        return jsonify({"error": "Missing project ID or progress value"}), 400
        
    if SUPABASE_CLIENT:
        try:
            SUPABASE_CLIENT.table("projects").update({"progress": int(progress)}).eq("id", project_id).execute()
            return jsonify({"success": True})
        except Exception as e:
            logger.error(f"Supabase admin project update error: {e}")
            return jsonify({"error": str(e)}), 500
            
    # Mock update
    for project in MOCK_DATABASE["projects"]:
        if project["id"] == project_id:
            project["progress"] = int(progress)
            return jsonify({"success": True, "mode": "mock"})
            
    return jsonify({"error": "Project not found"}), 404


# ----------------- RMC Module API Endpoints -----------------

@app.route('/api/rmc/grades', methods=['GET'])
def get_rmc_grades():
    if SUPABASE_CLIENT:
        try:
            response = SUPABASE_CLIENT.table("rmc_concrete_grades").select("*").eq("is_active", True).execute()
            if response.data:
                return jsonify(response.data)
        except Exception as e:
            logger.error(f"Supabase grades query error: {e}")
    return jsonify(MOCK_DATABASE["rmc_concrete_grades"])

@app.route('/api/rmc/clients', methods=['GET', 'POST'])
def rmc_clients():
    if request.method == 'GET':
        if SUPABASE_CLIENT:
            try:
                response = SUPABASE_CLIENT.table("rmc_clients").select("*").execute()
                if response.data:
                    return jsonify(response.data)
            except Exception as e:
                logger.error(f"Supabase clients query error: {e}")
        return jsonify(MOCK_DATABASE["rmc_clients"])
    
    elif request.method == 'POST':
        data = request.json
        if not data.get('company_name') or not data.get('email') or not data.get('phone'):
            return jsonify({"error": "Missing required fields"}), 400
        
        if SUPABASE_CLIENT:
            try:
                response = SUPABASE_CLIENT.table("rmc_clients").insert({
                    "company_name": data['company_name'],
                    "contact_person": data.get('contact_person', ''),
                    "email": data['email'],
                    "phone": data['phone'],
                    "billing_address": data.get('billing_address', ''),
                    "shipping_address": data.get('shipping_address', ''),
                    "gst_number": data.get('gst_number', ''),
                    "credit_limit": data.get('credit_limit', 0)
                }).execute()
                return jsonify({"success": True, "client": response.data[0]})
            except Exception as e:
                logger.error(f"Supabase client insert error: {e}")
        
        # Mock insert
        new_client = {
            "id": f"client-{uuid.uuid4().hex[:8]}",
            **data,
            "current_balance": 0.00,
            "is_active": True
        }
        MOCK_DATABASE["rmc_clients"].append(new_client)
        return jsonify({"success": True, "client": new_client, "mode": "mock"})

@app.route('/api/rmc/orders', methods=['GET', 'POST'])
def rmc_orders():
    if request.method == 'GET':
        client_id = request.args.get('client_id')
        
        # Check authentication and role
        user_role = session.get('role')
        user_id = session.get('user_id')
        
        if SUPABASE_CLIENT:
            try:
                query = SUPABASE_CLIENT.table("rmc_orders").select("*")
                
                # Filter by role
                if user_role == 'client' and user_id:
                    query = query.eq("user_id", user_id)
                elif client_id:
                    query = query.eq("client_id", client_id)
                
                response = query.execute()
                if response.data:
                    return jsonify(response.data)
            except Exception as e:
                logger.error(f"Supabase orders query error: {e}")
        
        # Mock
        orders = MOCK_DATABASE["rmc_orders"]
        
        # Filter by role
        if user_role == 'client' and user_id:
            orders = [o for o in orders if o.get('user_id') == user_id]
        elif client_id:
            orders = [o for o in orders if o.get('client_id') == client_id]
        
        return jsonify(orders)
    
    elif request.method == 'POST':
        data = request.json
        if not data.get('customer_name') or not data.get('items'):
            return jsonify({"error": "Missing customer_name or items"}), 400
        
        # Check authentication
        user_id = session.get('user_id')
        user_role = session.get('role')
        
        if not user_id:
            return jsonify({"error": "Not authenticated"}), 401
        
        # Calculate total
        total_amount = 0
        for item in data['items']:
            total_amount += item.get('subtotal', 0)
        
        order_number = f"RMC-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        
        if SUPABASE_CLIENT:
            try:
                # Create order
                order_response = SUPABASE_CLIENT.table("rmc_orders").insert({
                    "order_number": order_number,
                    "user_id": user_id,
                    "customer_name": data['customer_name'],
                    "phone": data.get('phone'),
                    "delivery_date": data.get('delivery_date'),
                    "delivery_address": data.get('delivery_address'),
                    "status": "Pending",
                    "total_amount": total_amount,
                    "notes": data.get('notes', '')
                }).execute()
                order_id = order_response.data[0]['id']
                
                # Create order items
                for item in data['items']:
                    SUPABASE_CLIENT.table("rmc_order_items").insert({
                        "order_id": order_id,
                        "grade_id": item['grade_id'],
                        "quantity_cubic_meters": item['quantity'],
                        "unit_price": item['unit_price'],
                        "subtotal": item['subtotal']
                    }).execute()
                
                return jsonify({"success": True, "order_id": order_id, "order_number": order_number})
            except Exception as e:
                logger.error(f"Supabase order insert error: {e}")
                return jsonify({"error": str(e)}), 500
        
        # Mock insert
        new_order = {
            "id": f"order-{uuid.uuid4().hex[:8]}",
            "order_number": order_number,
            "user_id": user_id,
            "customer_name": data['customer_name'],
            "phone": data.get('phone'),
            "order_date": datetime.now().isoformat(),
            "delivery_date": data.get('delivery_date'),
            "delivery_address": data.get('delivery_address'),
            "status": "Pending",
            "total_amount": total_amount,
            "notes": data.get('notes', ''),
            "items": data['items']
        }
        MOCK_DATABASE["rmc_orders"].append(new_order)
        return jsonify({"success": True, "order_id": new_order['id'], "order_number": order_number, "mode": "mock"})

@app.route('/api/rmc/orders/<order_id>/approve', methods=['POST'])
def approve_order(order_id):
    data = request.json
    approved_by = data.get('approved_by', 'Admin')
    
    # Find the order first
    order = None
    if SUPABASE_CLIENT:
        try:
            response = SUPABASE_CLIENT.table("rmc_orders").select("*").eq("id", order_id).execute()
            if response.data:
                order = response.data[0]
        except Exception as e:
            logger.error(f"Supabase order query error: {e}")
    else:
        for o in MOCK_DATABASE["rmc_orders"]:
            if o['id'] == order_id:
                order = o
                break
    
    if not order:
        return jsonify({"error": "Order not found"}), 404
    
    # Update order status
    if SUPABASE_CLIENT:
        try:
            SUPABASE_CLIENT.table("rmc_orders").update({
                "status": "Approved",
                "approved_by": approved_by,
                "approved_at": datetime.now().isoformat()
            }).eq("id", order_id).execute()
        except Exception as e:
            logger.error(f"Supabase order approve error: {e}")
            return jsonify({"error": str(e)}), 500
    else:
        for o in MOCK_DATABASE["rmc_orders"]:
            if o['id'] == order_id:
                o['status'] = "Approved"
                o['approved_by'] = approved_by
                o['approved_at'] = datetime.now().isoformat()
                order = o
                break
    
    # Generate invoice automatically
    invoice_number = f"INV-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
    
    if SUPABASE_CLIENT:
        try:
            # Create invoice
            invoice_response = SUPABASE_CLIENT.table("rmc_invoices").insert({
                "invoice_number": invoice_number,
                "order_id": order_id,
                "invoice_date": datetime.now().date().isoformat(),
                "due_date": (datetime.now() + timedelta(days=30)).date().isoformat(),
                "subtotal": order.get('total_amount', 0),
                "tax_amount": 0,
                "total_amount": order.get('total_amount', 0),
                "status": "Pending",
                "notes": f"Auto-generated invoice for order {order.get('order_number')}"
            }).execute()
            invoice_id = invoice_response.data[0]['id']
        except Exception as e:
            logger.error(f"Supabase invoice creation error: {e}")
    else:
        # Mock invoice creation
        new_invoice = {
            "id": f"invoice-{uuid.uuid4().hex[:8]}",
            "invoice_number": invoice_number,
            "order_id": order_id,
            "invoice_date": datetime.now().date().isoformat(),
            "due_date": (datetime.now() + timedelta(days=30)).date().isoformat(),
            "subtotal": order.get('total_amount', 0),
            "tax_amount": 0,
            "total_amount": order.get('total_amount', 0),
            "status": "Pending",
            "notes": f"Auto-generated invoice for order {order.get('order_number')}"
        }
        MOCK_DATABASE.setdefault("rmc_invoices", []).append(new_invoice)
        invoice_id = new_invoice['id']
    
    return jsonify({
        "success": True,
        "invoice_id": invoice_id,
        "invoice_number": invoice_number,
        "mode": "mock" if not SUPABASE_CLIENT else "supabase"
    })

@app.route('/api/rmc/orders/<order_id>', methods=['PUT'])
def update_order(order_id):
    data = request.json
    status = data.get('status')
    
    if not status:
        return jsonify({"error": "Missing status"}), 400
    
    if SUPABASE_CLIENT:
        try:
            SUPABASE_CLIENT.table("rmc_orders").update({
                "status": status,
                "updated_at": datetime.now().isoformat()
            }).eq("id", order_id).execute()
            return jsonify({"success": True})
        except Exception as e:
            logger.error(f"Supabase order update error: {e}")
            return jsonify({"error": str(e)}), 500
    
    # Mock update
    for order in MOCK_DATABASE["rmc_orders"]:
        if order['id'] == order_id:
            order['status'] = status
            order['updated_at'] = datetime.now().isoformat()
            return jsonify({"success": True, "mode": "mock"})
    
    return jsonify({"error": "Order not found"}), 404

@app.route('/api/rmc/invoices', methods=['GET', 'POST'])
def rmc_invoices():
    if request.method == 'GET':
        order_id = request.args.get('order_id')
        if SUPABASE_CLIENT:
            try:
                query = SUPABASE_CLIENT.table("rmc_invoices").select("*")
                if order_id:
                    query = query.eq("order_id", order_id)
                response = query.execute()
                if response.data:
                    return jsonify(response.data)
            except Exception as e:
                logger.error(f"Supabase invoices query error: {e}")
        
        invoices = MOCK_DATABASE["rmc_invoices"]
        if order_id:
            invoices = [i for i in invoices if i.get('order_id') == order_id]
        return jsonify(invoices)
    
    elif request.method == 'POST':
        data = request.json
        if not data.get('order_id') or not data.get('client_id'):
            return jsonify({"error": "Missing order_id or client_id"}), 400
        
        invoice_number = f"INV-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        due_date = (datetime.now() + timedelta(days=30)).date()
        
        if SUPABASE_CLIENT:
            try:
                response = SUPABASE_CLIENT.table("rmc_invoices").insert({
                    "invoice_number": invoice_number,
                    "order_id": data['order_id'],
                    "client_id": data['client_id'],
                    "due_date": due_date.isoformat(),
                    "subtotal": data['subtotal'],
                    "tax_amount": data.get('tax_amount', 0),
                    "total_amount": data['total_amount'],
                    "status": "Sent",
                    "payment_terms": data.get('payment_terms', 'Net 30 days')
                }).execute()
                return jsonify({"success": True, "invoice": response.data[0]})
            except Exception as e:
                logger.error(f"Supabase invoice insert error: {e}")
                return jsonify({"error": str(e)}), 500
        
        # Mock insert
        new_invoice = {
            "id": f"invoice-{uuid.uuid4().hex[:8]}",
            "invoice_number": invoice_number,
            "order_id": data['order_id'],
            "client_id": data['client_id'],
            "invoice_date": datetime.now().date().isoformat(),
            "due_date": due_date.isoformat(),
            "subtotal": data['subtotal'],
            "tax_amount": data.get('tax_amount', 0),
            "total_amount": data['total_amount'],
            "status": "Sent",
            "payment_terms": data.get('payment_terms', 'Net 30 days')
        }
        MOCK_DATABASE["rmc_invoices"].append(new_invoice)
        return jsonify({"success": True, "invoice": new_invoice, "mode": "mock"})

@app.route('/api/rmc/payments', methods=['GET', 'POST'])
def rmc_payments():
    if request.method == 'GET':
        client_id = request.args.get('client_id')
        if SUPABASE_CLIENT:
            try:
                query = SUPABASE_CLIENT.table("rmc_payment_records").select("*")
                if client_id:
                    query = query.eq("client_id", client_id)
                response = query.execute()
                if response.data:
                    return jsonify(response.data)
            except Exception as e:
                logger.error(f"Supabase payments query error: {e}")
        
        payments = MOCK_DATABASE["rmc_payments"]
        if client_id:
            payments = [p for p in payments if p.get('client_id') == client_id]
        return jsonify(payments)
    
    elif request.method == 'POST':
        data = request.json
        if not data.get('client_id') or not data.get('amount'):
            return jsonify({"error": "Missing client_id or amount"}), 400
        
        payment_number = f"PAY-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        
        if SUPABASE_CLIENT:
            try:
                response = SUPABASE_CLIENT.table("rmc_payment_records").insert({
                    "payment_number": payment_number,
                    "invoice_id": data.get('invoice_id'),
                    "client_id": data['client_id'],
                    "amount": data['amount'],
                    "payment_method": data.get('payment_method', 'Bank Transfer'),
                    "reference_number": data.get('reference_number', ''),
                    "notes": data.get('notes', '')
                }).execute()
                return jsonify({"success": True, "payment": response.data[0]})
            except Exception as e:
                logger.error(f"Supabase payment insert error: {e}")
                return jsonify({"error": str(e)}), 500
        
        # Mock insert
        new_payment = {
            "id": f"payment-{uuid.uuid4().hex[:8]}",
            "payment_number": payment_number,
            "invoice_id": data.get('invoice_id'),
            "client_id": data['client_id'],
            "payment_date": datetime.now().date().isoformat(),
            "amount": data['amount'],
            "payment_method": data.get('payment_method', 'Bank Transfer'),
            "reference_number": data.get('reference_number', ''),
            "notes": data.get('notes', '')
        }
        MOCK_DATABASE["rmc_payments"].append(new_payment)
        return jsonify({"success": True, "payment": new_payment, "mode": "mock"})

@app.route('/api/rmc/export/orders', methods=['GET'])
def export_orders_excel():
    client_id = request.args.get('client_id')
    
    # Get data
    if SUPABASE_CLIENT:
        try:
            query = SUPABASE_CLIENT.table("rmc_orders").select("*")
            if client_id:
                query = query.eq("client_id", client_id)
            response = query.execute()
            orders = response.data if response.data else []
        except Exception as e:
            logger.error(f"Supabase export error: {e}")
            orders = MOCK_DATABASE["rmc_orders"]
    else:
        orders = MOCK_DATABASE["rmc_orders"]
        if client_id:
            orders = [o for o in orders if o.get('client_id') == client_id]
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Headers
    headers = ["Order Number", "Client ID", "Order Date", "Delivery Date", "Status", "Total Amount", "Notes"]
    ws.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for order in orders:
        ws.append([
            order.get('order_number', ''),
            order.get('client_id', ''),
            order.get('order_date', ''),
            order.get('delivery_date', ''),
            order.get('status', ''),
            order.get('total_amount', 0),
            order.get('notes', '')
        ])
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=rmc_orders_export.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/api/rmc/invoices/<invoice_id>/pdf', methods=['GET'])
def generate_invoice_pdf(invoice_id):
    from io import BytesIO
    import os
    from reportlab.lib.units import inch
    from reportlab.platypus import Image, Table, TableStyle, Flowable
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, Spacer, SimpleDocTemplate
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.platypus import PageTemplate, Frame
    from reportlab.lib.pagesizes import A4
    
    try:
        # Find the invoice
        invoice = None
        order = None
        
        if SUPABASE_CLIENT:
            try:
                inv_response = SUPABASE_CLIENT.table("rmc_invoices").select("*").eq("id", invoice_id).execute()
                if inv_response.data:
                    invoice = inv_response.data[0]
                    order_response = SUPABASE_CLIENT.table("rmc_orders").select("*").eq("id", invoice['order_id']).execute()
                    if order_response.data:
                        order = order_response.data[0]
            except Exception as e:
                logger.error(f"Supabase invoice query error: {e}")
        else:
            for inv in MOCK_DATABASE.get("rmc_invoices", []):
                if inv['id'] == invoice_id:
                    invoice = inv
                    for ord in MOCK_DATABASE.get("rmc_orders", []):
                        if ord['id'] == invoice['order_id']:
                            order = ord
                            break
                    break
        
        if not invoice or not order:
            logger.error(f"Invoice or order not found for invoice_id: {invoice_id}")
            return jsonify({"error": "Invoice or order not found"}), 404
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20, bottomMargin=20, leftMargin=20, rightMargin=20)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles for overlay text
        invoice_title_style = ParagraphStyle(
            'InvoiceTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.black,
            spaceAfter=0,
            alignment=TA_RIGHT
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            spaceAfter=0,
            alignment=TA_LEFT
        )
        
        bold_style = ParagraphStyle(
            'Bold',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            fontName='Helvetica-Bold',
            spaceAfter=0,
            alignment=TA_LEFT
        )
        
        # Add invoice template as background image
        sheet_path = os.path.join('static', 'images', 'invoice_template.jpg')
        
        def on_first_page(canvas, doc):
            if os.path.exists(sheet_path):
                try:
                    canvas.drawImage(sheet_path, 0, 0, width=A4[0], height=A4[1], preserveAspectRatio=True)
                except Exception as e:
                    logger.error(f"Error drawing invoice template: {e}")
        
        # Create a custom flowable for absolute positioning
        class AbsolutePosition(Flowable):
            def __init__(self, x, y, text, style):
                Flowable.__init__(self)
                self.x = x
                self.y = y
                self.text = text
                self.style = style
            
            def wrap(self, availWidth, availHeight):
                return (0, 0)
            
            def draw(self):
                self.canv.saveState()
                self.canv.setFont(self.style.fontName, self.style.fontSize)
                self.canv.setFillColor(self.style.textColor)
                self.canv.drawString(self.x, self.y, self.text)
                self.canv.restoreState()
        
        # Add text overlays at specific positions (coordinates in points)
        # These coordinates need to match the template layout
        text_overlays = [
            # Invoice No - top right
            (450, 780, invoice.get('invoice_number', ''), bold_style),
            # Invoice Date - top right below invoice no
            (450, 760, invoice.get('invoice_date', ''), normal_style),
            # Due Date - top right below date
            (450, 740, invoice.get('due_date', ''), normal_style),
            # Order No - top right below due date
            (450, 720, order.get('order_number', ''), normal_style),
            # Company Name - left side
            (80, 680, order.get('customer_name', ''), bold_style),
            # Address - below company
            (80, 660, order.get('delivery_address', ''), normal_style),
            # Phone - below address
            (80, 640, order.get('phone', ''), normal_style),
            # Total Amount - bottom right
            (450, 200, f"₹{invoice.get('total_amount', 0):,.2f}", bold_style),
        ]
        
        for x, y, text, style in text_overlays:
            elements.append(AbsolutePosition(x, y, text, style))
        
        # Build PDF with background image
        doc.build(elements, onFirstPage=on_first_page)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=invoice_{invoice["invoice_number"]}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        return jsonify({"error": f"Error generating PDF: {str(e)}"}), 500

@app.route('/api/rmc/export/invoices', methods=['GET'])
def export_invoices_excel():
    if SUPABASE_CLIENT:
        try:
            response = SUPABASE_CLIENT.table("rmc_invoices").select("*").execute()
            invoices = response.data if response.data else []
        except Exception as e:
            logger.error(f"Supabase export error: {e}")
            invoices = MOCK_DATABASE["rmc_invoices"]
    else:
        invoices = MOCK_DATABASE["rmc_invoices"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    headers = ["Invoice Number", "Order ID", "Client ID", "Invoice Date", "Due Date", "Subtotal", "Tax", "Total", "Status"]
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for invoice in invoices:
        ws.append([
            invoice.get('invoice_number', ''),
            invoice.get('order_id', ''),
            invoice.get('client_id', ''),
            invoice.get('invoice_date', ''),
            invoice.get('due_date', ''),
            invoice.get('subtotal', 0),
            invoice.get('tax_amount', 0),
            invoice.get('total_amount', 0),
            invoice.get('status', '')
        ])
    
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=rmc_invoices_export.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

# Manual Billing Module
@app.route('/api/manual-billing', methods=['GET', 'POST'])
def manual_billing():
    if request.method == 'GET':
        if SUPABASE_CLIENT:
            try:
                response = SUPABASE_CLIENT.table("manual_billing").select("*").limit(10000).execute()
                billing_records = response.data if response.data else []
            except Exception as e:
                logger.error(f"Supabase billing fetch error: {e}")
                billing_records = MOCK_DATABASE.get("manual_billing", [])
        else:
            billing_records = MOCK_DATABASE.get("manual_billing", [])
        return jsonify({"success": True, "data": billing_records})
    
    elif request.method == 'POST':
        data = request.json
        required_fields = ['company_name', 'phone', 'address', 'date', 'time', 'vehicle_number', 'driver_name', 'unloading_time', 'grade', 'cubic_meters', 'loading_time', 'rate_per_cubic']
        
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({"error": f"Missing required field: {field}"}), 400
        
        billing_record = {
            "id": f"bill-{uuid.uuid4().hex[:8]}",
            "company_name": data['company_name'],
            "phone": data['phone'],
            "address": data['address'],
            "date": data['date'],
            "time": data['time'],
            "vehicle_number": data['vehicle_number'],
            "driver_name": data['driver_name'],
            "unloading_time": data['unloading_time'],
            "grade": data['grade'],
            "cubic_meters": data['cubic_meters'],
            "loading_time": data['loading_time'],
            "rate_per_cubic": data['rate_per_cubic'],
            "total_amount": data.get('total_amount', 0),
            "advance_amount": data.get('advance_amount', 0),
            "balance_amount": data.get('balance_amount', 0),
            "billing_status": "unbilled",
            "billing_summary_id": None,
            "created_at": datetime.now().isoformat(),
            "created_by": session.get('username', 'system')
        }

        # Use resilient insert (retries without settlement columns if DB not migrated)
        billing_record = _insert_manual_billing_record(billing_record)
        return jsonify({"success": True, "data": billing_record})

@app.route('/api/manual-billing/<billing_id>', methods=['PUT', 'DELETE'])
def manual_billing_detail(billing_id):
    if request.method == 'PUT':
        data = request.json
        
        if SUPABASE_CLIENT:
            try:
                SUPABASE_CLIENT.table("manual_billing").update(data).eq("id", billing_id).execute()
            except Exception as e:
                logger.error(f"Supabase billing update error: {e}")
        else:
            for record in MOCK_DATABASE.get("manual_billing", []):
                if record['id'] == billing_id:
                    record.update(data)
                    break
            save_db()  # Persist to disk
        
        return jsonify({"success": True})
    
    elif request.method == 'DELETE':
        if SUPABASE_CLIENT:
            try:
                SUPABASE_CLIENT.table("manual_billing").delete().eq("id", billing_id).execute()
            except Exception as e:
                logger.error(f"Supabase billing delete error: {e}")
        else:
            MOCK_DATABASE["manual_billing"] = [r for r in MOCK_DATABASE.get("manual_billing", []) if r['id'] != billing_id]
            save_db()  # Persist to disk
        
        return jsonify({"success": True})

@app.route('/api/manual-billing/<billing_id>/pdf', methods=['GET'])
def manual_billing_pdf(billing_id):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.units import mm
    import os

    try:
        billing_record = None

        if SUPABASE_CLIENT:
            try:
                response = SUPABASE_CLIENT.table("manual_billing").select("*").eq("id", billing_id).execute()
                if response.data:
                    billing_record = response.data[0]
            except Exception as e:
                logger.error(f"Supabase billing fetch error: {e}")

        # Fallback to local store (insert may have landed here if Supabase failed)
        if not billing_record:
            for record in MOCK_DATABASE.get("manual_billing", []):
                if record.get('id') == billing_id:
                    billing_record = record
                    break

        if not billing_record:
            return jsonify({"error": "Billing record not found"}), 404

        buffer = BytesIO()
        W, H = A4  # 595.28 x 841.89 points
        c = pdf_canvas.Canvas(buffer, pagesize=A4)

        # ── COLORS ──────────────────────────────────────────────
        GOLD        = colors.HexColor('#D4AF37')
        DARK_GOLD   = colors.HexColor('#B8962E')
        BLACK       = colors.HexColor('#1a1a1a')
        WHITE       = colors.white
        LIGHT_GRAY  = colors.HexColor('#f5f5f5')
        MID_GRAY    = colors.HexColor('#e0e0e0')
        TEXT_GRAY   = colors.HexColor('#555555')

        # ── HELPER ───────────────────────────────────────────────
        def draw_text(x, y, text, font='Helvetica', size=9, color=BLACK):
            c.setFont(font, size)
            c.setFillColor(color)
            c.drawString(x, y, str(text))

        def draw_right_text(x, y, text, font='Helvetica', size=9, color=BLACK):
            c.setFont(font, size)
            c.setFillColor(color)
            c.drawRightString(x, y, str(text))

        def draw_rect(x, y, w, h, fill_color=None, stroke_color=None, stroke_width=0.5):
            c.setLineWidth(stroke_width)
            if fill_color:
                c.setFillColor(fill_color)
            if stroke_color:
                c.setStrokeColor(stroke_color)
            if fill_color and stroke_color:
                c.rect(x, y, w, h, fill=1, stroke=1)
            elif fill_color:
                c.rect(x, y, w, h, fill=1, stroke=0)
            elif stroke_color:
                c.rect(x, y, w, h, fill=0, stroke=1)

        # ════════════════════════════════════════════════════════
        # SECTION 1 — TOP GOLD BANNER (logo box + company info)
        # ════════════════════════════════════════════════════════
        banner_h = 100
        banner_y = H - banner_h

        # Full gold banner background
        draw_rect(0, banner_y, W, banner_h, fill_color=GOLD)

        # Black logo box (left side)
        logo_box_w = 90
        logo_box_h = 76
        logo_box_x = 18
        logo_box_y = banner_y + (banner_h - logo_box_h) / 2
        draw_rect(logo_box_x, logo_box_y, logo_box_w, logo_box_h, fill_color=BLACK)

        # Draw logo image inside black box
        logo_path = os.path.join('static', 'images', 'logo.png')
        if os.path.exists(logo_path):
            try:
                c.drawImage(logo_path, logo_box_x + 5, logo_box_y + 5,
                            width=logo_box_w - 10, height=logo_box_h - 10,
                            preserveAspectRatio=True, mask='auto')
            except Exception as e:
                logger.error(f"Logo draw error: {e}")

        # ── White section to the right of gold banner (company + proprietor)
        white_section_x = logo_box_x + logo_box_w + 16
        white_section_y = banner_y
        white_section_w = W - white_section_x - 10

        # Company name + tagline (left of white section)
        company_x = white_section_x
        company_mid_y = banner_y + banner_h / 2

        draw_text(company_x, company_mid_y + 18, "R SUNDARAM & CO",
                  font='Helvetica-Bold', size=16, color=BLACK)
        draw_text(company_x, company_mid_y + 4, "STRONG ROADS  BUILD TO LAST",
                  font='Helvetica', size=8, color=BLACK)
        draw_text(company_x, company_mid_y - 10, "READYMIX CONCRETE",
                  font='Helvetica', size=8, color=BLACK)

        # Proprietor info (right side of banner)
        prop_x = W - 170
        prop_y = banner_y + 72
        draw_text(prop_x, prop_y,      "Proprietor",        font='Helvetica',      size=8,  color=BLACK)
        draw_text(prop_x, prop_y - 14, "S MAGHESH",         font='Helvetica-Bold', size=12, color=BLACK)
        draw_text(prop_x, prop_y - 27, "\u25a0 9940270304",   font='Helvetica',      size=8,  color=BLACK)
        draw_text(prop_x, prop_y - 38, "\u25a0 rsundaram&co@gmail.com", font='Helvetica', size=8, color=BLACK)
        draw_text(prop_x, prop_y - 49, "\u25a0 Quarry Rd, Tiruneermalai", font='Helvetica', size=8, color=BLACK)
        draw_text(prop_x, prop_y - 60, "Chennai, Tamil Nadu 600132", font='Helvetica', size=8, color=BLACK)

        # ════════════════════════════════════════════════════════
        # SECTION 2 — THIN DARK GOLD SEPARATOR LINE
        # ════════════════════════════════════════════════════════
        sep_y = banner_y - 3
        c.setStrokeColor(DARK_GOLD)
        c.setLineWidth(2)
        c.line(0, sep_y, W, sep_y)

        # ════════════════════════════════════════════════════════
        # SECTION 3 — DELIVERY CHALLAN TITLE
        # ════════════════════════════════════════════════════════
        title_y = sep_y - 38
        c.setFont('Helvetica-Bold', 22)
        c.setFillColor(GOLD)
        title_text = "DELIVERY CHALLAN"
        title_w = c.stringWidth(title_text, 'Helvetica-Bold', 22)
        c.drawString((W - title_w) / 2, title_y, title_text)

        # Underline
        c.setStrokeColor(GOLD)
        c.setLineWidth(1.5)
        underline_x = (W - title_w) / 2
        c.line(underline_x, title_y - 4, underline_x + title_w, title_y - 4)

        # ════════════════════════════════════════════════════════
        # SECTION 4 — BILL TO / CHALLAN DETAILS (two columns)
        # ════════════════════════════════════════════════════════
        info_top_y = title_y - 30
        col_left_x  = 30
        col_right_x = W / 2 + 10

        # "Bill To" label
        draw_text(col_left_x, info_top_y, "Bill To", font='Helvetica-Bold', size=9, color=TEXT_GRAY)
        # Bill To values
        r = billing_record
        draw_text(col_left_x, info_top_y - 14, "Company:", font='Helvetica-Bold', size=9, color=BLACK)
        draw_text(col_left_x + 55, info_top_y - 14, r.get('company_name', ''), font='Helvetica', size=9, color=BLACK)
        draw_text(col_left_x, info_top_y - 26, "Phone:",   font='Helvetica-Bold', size=9, color=BLACK)
        draw_text(col_left_x + 55, info_top_y - 26, str(r.get('phone', '')), font='Helvetica', size=9, color=BLACK)
        draw_text(col_left_x, info_top_y - 38, "Address:", font='Helvetica-Bold', size=9, color=BLACK)
        draw_text(col_left_x + 55, info_top_y - 38, str(r.get('address', '')), font='Helvetica', size=9, color=BLACK)

        # "Challan Details" label
        draw_text(col_right_x, info_top_y, "Challan Details", font='Helvetica-Bold', size=9, color=TEXT_GRAY)
        draw_text(col_right_x, info_top_y - 14, "Challan No:", font='Helvetica-Bold', size=9, color=BLACK)
        draw_text(col_right_x + 65, info_top_y - 14, str(r.get('id', '')), font='Helvetica', size=9, color=BLACK)
        draw_text(col_right_x, info_top_y - 26, "Date:", font='Helvetica-Bold', size=9, color=BLACK)
        draw_text(col_right_x + 65, info_top_y - 26, str(r.get('date', '')), font='Helvetica', size=9, color=BLACK)
        draw_text(col_right_x, info_top_y - 38, "Time:", font='Helvetica-Bold', size=9, color=BLACK)
        draw_text(col_right_x + 65, info_top_y - 38, str(r.get('time', '')), font='Helvetica', size=9, color=BLACK)

        # Horizontal rule below info section
        rule_y = info_top_y - 55
        c.setStrokeColor(MID_GRAY)
        c.setLineWidth(0.5)
        c.line(30, rule_y, W - 30, rule_y)

        # ════════════════════════════════════════════════════════
        # SECTION 5 — DETAILS TABLE
        # ════════════════════════════════════════════════════════
        table_top_y  = rule_y - 8
        table_left   = 30
        table_right  = W - 30
        table_width  = table_right - table_left
        col1_w       = table_width * 0.40   # Field column
        col2_w       = table_width * 0.60   # Details column
        row_h        = 26
        header_h     = 26

        # Table header row (gold background)
        draw_rect(table_left, table_top_y - header_h, table_width, header_h, fill_color=GOLD)
        draw_text(table_left + 10, table_top_y - header_h + 8, "Field",   font='Helvetica-Bold', size=9, color=BLACK)
        draw_text(table_left + col1_w + 10, table_top_y - header_h + 8, "Details", font='Helvetica-Bold', size=9, color=BLACK)

        # Table rows
        rows = [
            ("Vehicle Number",  str(r.get('vehicle_number', ''))),
            ("Driver Name",     str(r.get('driver_name', ''))),
            ("Grade",           str(r.get('grade', ''))),
            ("Quantity (m\u00b3)", str(r.get('cubic_meters', ''))),
            ("Rate per m\u00b3",   f"Rs. {float(r.get('rate_per_cubic', 0)):.2f}"),
            ("Loading Time",    str(r.get('loading_time', ''))),
            ("Unloading Time",  str(r.get('unloading_time', ''))),
        ]

        row_y = table_top_y - header_h
        for i, (field, detail) in enumerate(rows):
            row_bg = LIGHT_GRAY if i % 2 == 0 else WHITE
            draw_rect(table_left, row_y - row_h, table_width, row_h, fill_color=row_bg, stroke_color=MID_GRAY, stroke_width=0.4)
            # Vertical divider
            c.setStrokeColor(MID_GRAY)
            c.setLineWidth(0.4)
            c.line(table_left + col1_w, row_y - row_h, table_left + col1_w, row_y)
            draw_text(table_left + 10, row_y - row_h + 8, field,  font='Helvetica',      size=9, color=BLACK)
            draw_text(table_left + col1_w + 10, row_y - row_h + 8, detail, font='Helvetica', size=9, color=BLACK)
            row_y -= row_h

        # ════════════════════════════════════════════════════════
        # SECTION 6 — AMOUNTS SUMMARY (right-aligned)
        # ════════════════════════════════════════════════════════
        amt_top_y    = row_y - 20
        amt_right_x  = W - 30
        amt_label_x  = W - 200
        amt_val_x    = W - 30
        line_h       = 22

        total   = float(r.get('total_amount', 0))
        advance = float(r.get('advance_amount', 0))
        balance = float(r.get('balance_amount', 0))

        # Total Amount
        draw_text(amt_label_x, amt_top_y, "Total Amount:", font='Helvetica', size=10, color=BLACK)
        draw_right_text(amt_val_x, amt_top_y, f"Rs. {total:,.2f}", font='Helvetica-Bold', size=10, color=BLACK)

        # Advance Given
        draw_text(amt_label_x, amt_top_y - line_h, "Advance Given:", font='Helvetica', size=10, color=BLACK)
        draw_right_text(amt_val_x, amt_top_y - line_h, f"Rs. {advance:,.2f}", font='Helvetica', size=10, color=BLACK)

        # Gold divider line
        c.setStrokeColor(GOLD)
        c.setLineWidth(1.2)
        c.line(amt_label_x, amt_top_y - line_h - 8, amt_val_x, amt_top_y - line_h - 8)

        # Balance Amount (highlighted gold)
        draw_text(amt_label_x, amt_top_y - line_h * 2 - 10, "Balance Amount:", font='Helvetica-Bold', size=11, color=GOLD)
        draw_right_text(amt_val_x, amt_top_y - line_h * 2 - 10, f"Rs. {balance:,.2f}", font='Helvetica-Bold', size=11, color=GOLD)

        # ════════════════════════════════════════════════════════
        # SECTION 7 — FOOTER
        # ════════════════════════════════════════════════════════
        footer_h = 36
        draw_rect(0, 0, W, footer_h, fill_color=BLACK)
        draw_text(30, footer_h - 14, "R SUNDARAM & CO  \u2022  READYMIX CONCRETE",
                  font='Helvetica-Bold', size=8, color=GOLD)
        draw_text(30, footer_h - 26, "Strong Roads  \u2022  Build To Last",
                  font='Helvetica', size=7, color=WHITE)

        # Outer border
        c.setStrokeColor(GOLD)
        c.setLineWidth(1.5)
        c.rect(10, 10, W - 20, H - 20, fill=0, stroke=1)

        c.showPage()
        c.save()
        buffer.seek(0)

        response = make_response(buffer.getvalue())
        response.headers['Content-Disposition'] = f'inline; filename=delivery_challan_{billing_id}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response

    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": f"Error generating PDF: {str(e)}"}), 500

@app.route('/api/manual-billing/export', methods=['GET'])
def export_manual_billing_excel():
    if SUPABASE_CLIENT:
        try:
            response = SUPABASE_CLIENT.table("manual_billing").select("*").limit(10000).execute()
            billing_records = response.data if response.data else []
        except Exception as e:
            logger.error(f"Supabase billing export error: {e}")
            billing_records = MOCK_DATABASE.get("manual_billing", [])
    else:
        billing_records = MOCK_DATABASE.get("manual_billing", [])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Manual Billing"
    
    headers = ["ID", "Company Name", "Phone", "Address", "Date", "Time", "Vehicle Number", "Driver Name", "Grade", "Cubic Meters", "Rate per m³", "Total Amount", "Advance Given", "Balance Amount", "Loading Time", "Unloading Time", "Created At"]
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for record in billing_records:
        ws.append([
            record.get('id', ''),
            record.get('company_name', ''),
            record.get('phone', ''),
            record.get('address', ''),
            record.get('date', ''),
            record.get('time', ''),
            record.get('vehicle_number', ''),
            record.get('driver_name', ''),
            record.get('grade', ''),
            record.get('cubic_meters', 0),
            record.get('rate_per_cubic', 0),
            record.get('total_amount', 0),
            record.get('advance_amount', 0),
            record.get('balance_amount', 0),
            record.get('loading_time', ''),
            record.get('unloading_time', ''),
            record.get('created_at', '')
        ])
    
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=manual_billing_export.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

# Expense Management Module
@app.route('/api/expenses', methods=['GET', 'POST'])
def expenses():
    if request.method == 'GET':
        if SUPABASE_CLIENT:
            try:
                response = SUPABASE_CLIENT.table("expenses").select("*").execute()
                expense_records = response.data if response.data else []
            except Exception as e:
                logger.error(f"Supabase expenses fetch error: {e}")
                expense_records = MOCK_DATABASE.get("expenses", [])
        else:
            expense_records = MOCK_DATABASE.get("expenses", [])
        return jsonify({"success": True, "data": expense_records})
    
    elif request.method == 'POST':
        data = request.get_json(silent=True)

        if not isinstance(data, dict):
            return jsonify({"error": "A valid JSON expense record is required"}), 400

        # Only category and date are always required
        if not data.get('category') or not data.get('date'):
            return jsonify({"error": "Category and Date are required"}), 400

        # Build the record — capture ALL fields the frontend may send
        expense_record = {
            "id": f"expense-{uuid.uuid4().hex[:8]}",
            "category":       data.get('category', ''),
            "date":           data.get('date', ''),
            "description":    data.get('description', ''),
            "company_name":   data.get('company_name', data.get('description', '')),
            "gst_no":         data.get('gst_no', ''),
            # Raw-material specific fields
            "material_type":  data.get('material_type', ''),
            "quality":        data.get('quality', ''),
            "quantity":       data.get('quantity', ''),
            "rate_per_ton":   data.get('rate_per_ton', ''),
            # Financial fields
            "amount":         data.get('amount', 0),
            "total_amount":   data.get('total_amount', data.get('amount', 0)),
            "advance_amount": data.get('advance_amount', 0),
            "balance_amount": data.get('balance_amount', data.get('amount', 0)),
            "supplier_payment_status": data.get('supplier_payment_status', 'unpaid'),
            "supplier_statement_id": data.get('supplier_statement_id'),
            "created_at":     datetime.now().isoformat(),
            "created_by":     session.get('username', 'system')
        }

        if SUPABASE_CLIENT:
            try:
                response = SUPABASE_CLIENT.table("expenses").insert(expense_record).execute()
                # An insert must return the row we just wrote.  Treat an empty
                # response as a failure so the UI never displays a false success.
                if not response.data:
                    raise RuntimeError("The database did not confirm the expense insert")
            except Exception as e:
                logger.error(f"Supabase expense insert error: {e}")
                return jsonify({
                    "error": "Expense was not saved to the database. Please try again or contact an administrator."
                }), 502
        else:
            MOCK_DATABASE.setdefault("expenses", []).append(expense_record)
            save_db()  # Persist to disk

        return jsonify({"success": True, "data": expense_record})

@app.route('/api/expenses/<expense_id>', methods=['PUT', 'DELETE'])
def expense_detail(expense_id):
    if request.method == 'PUT':
        data = request.json
        
        if SUPABASE_CLIENT:
            try:
                SUPABASE_CLIENT.table("expenses").update(data).eq("id", expense_id).execute()
            except Exception as e:
                logger.error(f"Supabase expense update error: {e}")
        else:
            for record in MOCK_DATABASE.get("expenses", []):
                if record['id'] == expense_id:
                    record.update(data)
                    break
            save_db()  # Persist to disk
        
        return jsonify({"success": True})
    
    elif request.method == 'DELETE':
        if SUPABASE_CLIENT:
            try:
                SUPABASE_CLIENT.table("expenses").delete().eq("id", expense_id).execute()
            except Exception as e:
                logger.error(f"Supabase expense delete error: {e}")
        else:
            MOCK_DATABASE["expenses"] = [r for r in MOCK_DATABASE.get("expenses", []) if r['id'] != expense_id]
            save_db()  # Persist to disk
        return jsonify({"success": True})


@app.route('/api/supplier-payment-summary', methods=['GET'])
def supplier_payment_summary():
    if 'username' not in session:
        return jsonify({"error": "Authentication required"}), 401
    if session.get('role') not in ['admin', 'manager']:
        return jsonify({"error": "Access denied"}), 403

    supplier = request.args.get('supplier', '').strip()
    start_date = request.args.get('start', '').strip()
    end_date = request.args.get('end', '').strip()
    include_paid = request.args.get('include_paid', '').lower() in ('1', 'true', 'yes')

    if not supplier:
        return jsonify({"error": "Supplier name is required"}), 400

    try:
        summary = _compute_supplier_payment_summary(supplier, start_date, end_date, include_paid=include_paid)
        return jsonify(summary)
    except Exception as e:
        logger.error(f"Supplier payment summary error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/supplier-payment-summary/history', methods=['GET'])
def supplier_payment_summary_history():
    if 'username' not in session:
        return jsonify({"error": "Authentication required"}), 401
    if session.get('role') not in ['admin', 'manager']:
        return jsonify({"error": "Access denied"}), 403

    supplier = request.args.get('supplier', '').strip()
    try:
        rows = _fetch_supplier_payment_summaries()
        if supplier:
            wanted = _normalize_supplier_name(supplier)
            rows = [r for r in rows if wanted in _normalize_supplier_name(r.get('supplier_name'))]
        rows.sort(key=lambda r: str(r.get('statement_date') or r.get('created_at') or ''), reverse=True)
        return jsonify({"success": True, "data": rows})
    except Exception as e:
        logger.error(f"Supplier payment history error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/supplier-payment-summary/confirm', methods=['POST'])
def supplier_payment_summary_confirm():
    if 'username' not in session:
        return jsonify({"error": "Authentication required"}), 401
    if session.get('role') not in ['admin', 'manager']:
        return jsonify({"error": "Access denied"}), 403

    data = request.json or {}
    if not data.get('confirm'):
        return jsonify({"error": "Confirmation required. Set confirm=true to settle."}), 400

    supplier = (data.get('supplier') or '').strip()
    start_date = (data.get('start') or '').strip()
    end_date = (data.get('end') or '').strip()
    requested_ids = data.get('transaction_ids')

    if not supplier:
        return jsonify({"error": "Supplier name is required"}), 400

    try:
        summary = _compute_supplier_payment_summary(supplier, start_date, end_date, include_paid=False)
        if not summary.get('matched'):
            return jsonify({"error": summary.get('message', 'Supplier not found')}), 404

        txn_ids = list(summary.get('transaction_ids') or [])
        if requested_ids is not None:
            requested_set = set(requested_ids)
            txn_ids = [i for i in txn_ids if i in requested_set]
            summary['transactions'] = [t for t in summary['transactions'] if t['id'] in txn_ids]
            summary['transaction_ids'] = txn_ids
            summary['total_amount'] = round(sum(t['price'] for t in summary['transactions']), 2)
            summary['advance_amount'] = round(sum(float(next((r.get('advance_amount') for r in _fetch_all_expenses() if r.get('id') == t['id']), 0)) for t in summary['transactions']), 2)
            if summary['advance_amount'] > summary['total_amount']:
                summary['advance_deducted'] = round(summary['total_amount'], 2)
                summary['balance_amount'] = 0.0
                summary['advance_carried_forward'] = round(summary['advance_amount'] - summary['total_amount'], 2)
                summary['balance_label'] = 'Advance Carried Forward'
                summary['display_balance'] = summary['advance_carried_forward']
            else:
                summary['advance_deducted'] = round(summary['advance_amount'], 2)
                summary['balance_amount'] = round(summary['total_amount'] - summary['advance_amount'], 2)
                summary['advance_carried_forward'] = 0.0
                summary['balance_label'] = 'Balance Amount'
                summary['display_balance'] = summary['balance_amount']

        statement_date = datetime.now().strftime('%Y-%m-%d')
        statement_number = f"SPS-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
        summary_id = f"sps-{uuid.uuid4().hex[:10]}"

        pdf_bytes, pdf_filename = _build_supplier_payment_pdf(summary, statement_number, statement_date)
        pdf_path = None
        try:
            os.makedirs(BILLING_PDF_DIR, exist_ok=True)
            abs_path = os.path.join(BILLING_PDF_DIR, pdf_filename)
            with open(abs_path, 'wb') as f:
                f.write(pdf_bytes)
            pdf_path = abs_path
        except Exception as e:
            logger.warning(f"Could not save supplier statement PDF to disk: {e}")
            pdf_path = f"memory:{pdf_filename}"

        summary_record = {
            "id": summary_id,
            "statement_number": statement_number,
            "supplier_name": summary['supplier_name'],
            "statement_date": statement_date,
            "period_start": summary['period']['start'],
            "period_end": summary['period']['end'],
            "total_amount": summary['total_amount'],
            "advance_deducted": summary['advance_amount'],
            "balance_amount": summary['balance_amount'],
            "advance_carried_forward": summary['advance_carried_forward'],
            "transaction_ids": json.dumps(txn_ids),
            "pdf_path": pdf_path,
            "pdf_filename": pdf_filename,
            "notes": summary['balance_label'],
            "created_at": datetime.now().isoformat(),
            "created_by": session.get('username', 'system')
        }
        _insert_supplier_payment_summary(summary_record)

        for tid in txn_ids:
            _update_expense_record(tid, {
                "supplier_payment_status": "paid",
                "supplier_statement_id": summary_id
            })

        return jsonify({
            "success": True,
            "message": "Supplier statement confirmed and generated successfully",
            "statement": summary_record,
            "summary": {
                "supplier_name": summary['supplier_name'],
                "total_amount": summary['total_amount'],
                "advance_deducted": summary['advance_amount'],
                "balance_amount": summary['balance_amount'],
                "advance_carried_forward": summary['advance_carried_forward'],
                "balance_label": summary['balance_label'],
                "records_settled": len(txn_ids)
            },
            "pdf_url": f"/api/supplier-payment-summary/{summary_id}/pdf"
        })
    except Exception as e:
        logger.error(f"Supplier payment confirm error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/supplier-payment-summary/<summary_id>/pdf', methods=['GET'])
def supplier_payment_summary_pdf(summary_id):
    if 'username' not in session:
        return jsonify({"error": "Authentication required"}), 401
    if session.get('role') not in ['admin', 'manager']:
        return jsonify({"error": "Access denied"}), 403

    try:
        cache = getattr(app, '_billing_pdf_cache', {})
        if summary_id in cache:
            pdf_bytes = cache[summary_id]
            fname = f"supplier_statement_{summary_id}.pdf"
            response = make_response(pdf_bytes)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'inline; filename={fname}'
            return response

        rows = _fetch_supplier_payment_summaries()
        rec = next((r for r in rows if r.get('id') == summary_id), None)
        if not rec:
            return jsonify({"error": "Supplier statement not found"}), 404

        pdf_path = rec.get('pdf_path') or ''
        fname = rec.get('pdf_filename') or f"{summary_id}.pdf"
        if pdf_path and not pdf_path.startswith('memory:') and os.path.exists(pdf_path):
            with open(pdf_path, 'rb') as f:
                pdf_bytes = f.read()
            response = make_response(pdf_bytes)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'inline; filename={fname}'
            return response

        return jsonify({"error": "PDF file not available"}), 404
    except Exception as e:
        logger.error(f"Supplier payment summary PDF error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

# ── Per-expense PDF Invoice ──────────────────────────────────────────────────
@app.route('/api/expenses/<expense_id>/pdf', methods=['GET'])
def expense_pdf(expense_id):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    import os

    try:
        rec = None
        if SUPABASE_CLIENT:
            try:
                resp = SUPABASE_CLIENT.table("expenses").select("*").eq("id", expense_id).execute()
                if resp.data:
                    rec = resp.data[0]
            except Exception as e:
                logger.error(f"Supabase expense fetch error: {e}")
        else:
            for r in MOCK_DATABASE.get("expenses", []):
                if r['id'] == expense_id:
                    rec = r
                    break

        if not rec:
            return jsonify({"error": "Expense record not found"}), 404

        buffer = BytesIO()
        W, H = A4
        c = pdf_canvas.Canvas(buffer, pagesize=A4)

        GOLD      = colors.HexColor('#D4AF37')
        DARK_GOLD = colors.HexColor('#B8962E')
        BLACK     = colors.HexColor('#1a1a1a')
        WHITE     = colors.white
        LIGHT_GRAY = colors.HexColor('#f5f5f5')
        MID_GRAY  = colors.HexColor('#e0e0e0')
        GREEN     = colors.HexColor('#22c55e')

        def dt(x, y, text, font='Helvetica', size=9, color=BLACK):
            c.setFont(font, size); c.setFillColor(color); c.drawString(x, y, str(text))

        def drt(x, y, text, font='Helvetica', size=9, color=BLACK):
            c.setFont(font, size); c.setFillColor(color); c.drawRightString(x, y, str(text))

        def drect(x, y, w, h, fill=None, stroke=None, sw=0.5):
            c.setLineWidth(sw)
            if fill: c.setFillColor(fill)
            if stroke: c.setStrokeColor(stroke)
            if fill and stroke: c.rect(x, y, w, h, fill=1, stroke=1)
            elif fill: c.rect(x, y, w, h, fill=1, stroke=0)
            elif stroke: c.rect(x, y, w, h, fill=0, stroke=1)

        # ── HEADER BANNER ─────────────────────────────────────────
        banner_h = 100; banner_y = H - banner_h
        drect(0, banner_y, W, banner_h, fill=GOLD)

        logo_box_w, logo_box_h = 90, 76
        logo_box_x = 18; logo_box_y = banner_y + (banner_h - logo_box_h) / 2
        drect(logo_box_x, logo_box_y, logo_box_w, logo_box_h, fill=BLACK)

        logo_path = os.path.join('static', 'images', 'logo.png')
        if os.path.exists(logo_path):
            try:
                c.drawImage(logo_path, logo_box_x+5, logo_box_y+5,
                            width=logo_box_w-10, height=logo_box_h-10,
                            preserveAspectRatio=True, mask='auto')
            except Exception as e:
                logger.error(f"Logo draw error: {e}")

        cx = logo_box_x + logo_box_w + 16
        cy = banner_y + banner_h / 2
        dt(cx, cy+18, "R SUNDARAM & CO",         font='Helvetica-Bold', size=16, color=BLACK)
        dt(cx, cy+4,  "STRONG ROADS  BUILD TO LAST", font='Helvetica', size=8,  color=BLACK)
        dt(cx, cy-10, "READYMIX CONCRETE",        font='Helvetica', size=8,  color=BLACK)

        px = W - 170; py = banner_y + 72
        dt(px, py,    "Proprietor",              font='Helvetica',      size=8,  color=BLACK)
        dt(px, py-14, "S MAGHESH",               font='Helvetica-Bold', size=12, color=BLACK)
        dt(px, py-27, "\u25a0 9940270304",          font='Helvetica',      size=8,  color=BLACK)
        dt(px, py-38, "\u25a0 rsundaram&co@gmail.com", font='Helvetica',  size=8,  color=BLACK)
        dt(px, py-49, "\u25a0 Quarry Rd, Tiruneermalai", font='Helvetica', size=8, color=BLACK)
        dt(px, py-60, "Chennai, Tamil Nadu 600132", font='Helvetica',    size=8,  color=BLACK)

        # Separator
        c.setStrokeColor(DARK_GOLD); c.setLineWidth(2); c.line(0, banner_y-3, W, banner_y-3)

        # ── TITLE ─────────────────────────────────────────────────
        title_y = banner_y - 44
        c.setFont('Helvetica-Bold', 20); c.setFillColor(GOLD)
        title = "EXPENSE INVOICE"
        tw = c.stringWidth(title, 'Helvetica-Bold', 20)
        c.drawString((W-tw)/2, title_y, title)
        c.setStrokeColor(GOLD); c.setLineWidth(1.5)
        c.line((W-tw)/2, title_y-4, (W-tw)/2+tw, title_y-4)

        # ── META ROW ─────────────────────────────────────────────
        meta_y = title_y - 28
        dt(30,     meta_y, "Expense ID:", font='Helvetica-Bold', size=9, color=BLACK)
        dt(100,    meta_y, str(rec.get('id','')), size=9)
        dt(W/2+10, meta_y, "Date:", font='Helvetica-Bold', size=9, color=BLACK)
        dt(W/2+50, meta_y, str(rec.get('date','')), size=9)

        # ── CATEGORY BADGE ───────────────────────────────────────
        cat_y = meta_y - 26
        cat = rec.get('category','')
        dt(30, cat_y, "Category:", font='Helvetica-Bold', size=9, color=BLACK)
        badge_x = 95; badge_w = c.stringWidth(cat,'Helvetica-Bold',9)+16
        drect(badge_x, cat_y-3, badge_w, 16, fill=GOLD)
        dt(badge_x+8, cat_y+1, cat, font='Helvetica-Bold', size=9, color=BLACK)

        # description
        if rec.get('description'):
            dt(30, cat_y-18, "Notes:", font='Helvetica-Bold', size=9, color=BLACK)
            dt(70, cat_y-18, str(rec.get('description','')), size=9)

        # ── DETAILS TABLE ────────────────────────────────────────
        is_raw = cat == 'Raw Material'
        tbl_top = cat_y - (50 if rec.get('description') else 36)
        tbl_left = 30; tbl_w = W - 60
        col1 = tbl_w * 0.45; row_h = 28; hdr_h = 28

        drect(tbl_left, tbl_top-hdr_h, tbl_w, hdr_h, fill=GOLD)
        dt(tbl_left+10, tbl_top-hdr_h+9, "Field",   font='Helvetica-Bold', size=9, color=BLACK)
        dt(tbl_left+col1+10, tbl_top-hdr_h+9, "Value", font='Helvetica-Bold', size=9, color=BLACK)

        rows_data = []
        if is_raw:
            rows_data = [
                ("Material Type",    str(rec.get('material_type','\u2014'))),
                ("Quality",          str(rec.get('quality','\u2014')) if rec.get('quality') else '\u2014'),
                ("Quantity",         f"{rec.get('quantity','\u2014')} Tons"),
                ("Rate per Ton",     f"\u20b9 {float(rec.get('rate_per_ton',0)):,.2f}"),
            ]
        rows_data.append(("Total Amount", f"\u20b9 {float(rec.get('total_amount', rec.get('amount',0))):,.2f}"))
        if is_raw:
            rows_data.append(("Advance Given",  f"\u20b9 {float(rec.get('advance_amount',0)):,.2f}"))
            rows_data.append(("Balance Due",    f"\u20b9 {float(rec.get('balance_amount',0)):,.2f}"))

        row_y = tbl_top - hdr_h
        for i, (field, val) in enumerate(rows_data):
            bg = LIGHT_GRAY if i%2==0 else WHITE
            drect(tbl_left, row_y-row_h, tbl_w, row_h, fill=bg, stroke=MID_GRAY, sw=0.4)
            c.setStrokeColor(MID_GRAY); c.setLineWidth(0.4)
            c.line(tbl_left+col1, row_y-row_h, tbl_left+col1, row_y)
            is_last = (i == len(rows_data)-1)
            val_color = GREEN if (is_raw and is_last) else BLACK
            dt(tbl_left+10, row_y-row_h+9, field, font='Helvetica', size=9, color=BLACK)
            dt(tbl_left+col1+10, row_y-row_h+9, val, font='Helvetica-Bold', size=9, color=val_color)
            row_y -= row_h

        # ── FOOTER ───────────────────────────────────────────────
        drect(0, 0, W, 36, fill=BLACK)
        dt(30, 22, "R SUNDARAM & CO  \u2022  READYMIX CONCRETE", font='Helvetica-Bold', size=8, color=GOLD)
        dt(30, 10, "Strong Roads  \u2022  Build To Last", font='Helvetica', size=7, color=WHITE)

        c.setStrokeColor(GOLD); c.setLineWidth(1.5)
        c.rect(10, 10, W-20, H-20, fill=0, stroke=1)

        c.showPage(); c.save(); buffer.seek(0)

        resp = make_response(buffer.getvalue())
        resp.headers['Content-Disposition'] = f'inline; filename=expense_{expense_id}.pdf'
        resp.headers['Content-Type'] = 'application/pdf'
        return resp

    except Exception as e:
        logger.error(f"Expense PDF error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ── Per-expense Excel ────────────────────────────────────────────────────────
@app.route('/api/expenses/<expense_id>/excel', methods=['GET'])
def expense_excel(expense_id):
    try:
        rec = None
        if SUPABASE_CLIENT:
            try:
                resp = SUPABASE_CLIENT.table("expenses").select("*").eq("id", expense_id).execute()
                if resp.data:
                    rec = resp.data[0]
            except Exception as e:
                logger.error(f"Supabase expense fetch error: {e}")
        else:
            for r in MOCK_DATABASE.get("expenses", []):
                if r['id'] == expense_id:
                    rec = r
                    break

        if not rec:
            return jsonify({"error": "Expense record not found"}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Invoice"

        gold_fill   = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
        black_fill  = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
        header_font = Font(bold=True, color="1A1A1A", size=11)
        white_font  = Font(bold=True, color="FFFFFF", size=11)
        gold_font   = Font(bold=True, color="D4AF37", size=11)
        center      = Alignment(horizontal="center", vertical="center")
        left        = Alignment(horizontal="left", vertical="center")

        # Title rows
        ws.merge_cells('A1:H1')
        ws['A1'] = "R SUNDARAM & CO — EXPENSE INVOICE"
        ws['A1'].font = Font(bold=True, color="D4AF37", size=14)
        ws['A1'].fill = black_fill
        ws['A1'].alignment = center

        ws.merge_cells('A2:H2')
        ws['A2'] = "Strong Roads  •  Build To Last"
        ws['A2'].font = Font(color="888888", size=9)
        ws['A2'].fill = black_fill
        ws['A2'].alignment = center

        ws.append([])  # blank row

        # Column headers
        headers = ["Expense ID","Date","Category","Material Type","Quality","Qty (Tons)","Rate/Ton (₹)","Total (₹)","Advance (₹)","Balance (₹)","Description","Created At"]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.fill = gold_fill
            cell.font = header_font
            cell.alignment = center

        # Data row
        ws.append([
            rec.get('id',''),
            rec.get('date',''),
            rec.get('category',''),
            rec.get('material_type','—'),
            rec.get('quality','—') if rec.get('quality') else '—',
            rec.get('quantity','—'),
            rec.get('rate_per_ton','—'),
            float(rec.get('total_amount', rec.get('amount', 0))),
            float(rec.get('advance_amount', 0)),
            float(rec.get('balance_amount', 0)),
            rec.get('description',''),
            rec.get('created_at',''),
        ])
        for col_idx in range(1, 13):
            ws.cell(row=5, column=col_idx).alignment = left

        # Column widths
        for col, width in zip(['A','B','C','D','E','F','G','H','I','J','K','L'], [20,14,16,16,12,12,14,14,14,14,28,22]):
            ws.column_dimensions[col].width = width
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[4].height = 20

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf); buf.seek(0)

        resp = make_response(buf.getvalue())
        resp.headers['Content-Disposition'] = f'attachment; filename=expense_{expense_id}.xlsx'
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return resp

    except Exception as e:
        logger.error(f"Expense Excel error: {e}")
        return jsonify({"error": str(e)}), 500



@app.route('/api/expenses/export', methods=['GET'])
def export_expenses_excel():
    if SUPABASE_CLIENT:
        try:
            response = SUPABASE_CLIENT.table("expenses").select("*").execute()
            expense_records = response.data if response.data else []
        except Exception as e:
            logger.error(f"Supabase expenses export error: {e}")
            expense_records = MOCK_DATABASE.get("expenses", [])
    else:
        expense_records = MOCK_DATABASE.get("expenses", [])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    headers = ["ID", "Category", "Amount", "Company Name", "GST NO", "Date", "Created At", "Created By"]
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for record in expense_records:
        ws.append([
            record.get('id', ''),
            record.get('category', ''),
            record.get('amount', 0),
            record.get('description', ''),
            record.get('gst_no', ''),
            record.get('date', ''),
            record.get('created_at', ''),
            record.get('created_by', '')
        ])
    
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=expenses_export.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/api/rmc/export/payments', methods=['GET'])
def export_payments_excel():
    client_id = request.args.get('client_id')
    
    if SUPABASE_CLIENT:
        try:
            query = SUPABASE_CLIENT.table("rmc_payment_records").select("*")
            if client_id:
                query = query.eq("client_id", client_id)
            response = query.execute()
            payments = response.data if response.data else []
        except Exception as e:
            logger.error(f"Supabase export error: {e}")
            payments = MOCK_DATABASE["rmc_payments"]
    else:
        payments = MOCK_DATABASE["rmc_payments"]
        if client_id:
            payments = [p for p in payments if p.get('client_id') == client_id]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    headers = ["Payment Number", "Invoice ID", "Client ID", "Payment Date", "Amount", "Method", "Reference", "Notes"]
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for payment in payments:
        ws.append([
            payment.get('payment_number', ''),
            payment.get('invoice_id', ''),
            payment.get('client_id', ''),
            payment.get('payment_date', ''),
            payment.get('amount', 0),
            payment.get('payment_method', ''),
            payment.get('reference_number', ''),
            payment.get('notes', '')
        ])
    
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=rmc_payments_export.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/api/rmc/payment/<payment_id>/receipt', methods=['GET'])
def generate_payment_receipt(payment_id):
    from io import BytesIO
    
    # Get payment data
    payment = None
    if SUPABASE_CLIENT:
        try:
            response = SUPABASE_CLIENT.table("rmc_payment_records").select("*").eq("id", payment_id).execute()
            if response.data:
                payment = response.data[0]
        except Exception as e:
            logger.error(f"Supabase payment query error: {e}")
    
    if not payment:
        payment = next((p for p in MOCK_DATABASE["rmc_payments"] if p['id'] == payment_id), None)
    
    if not payment:
        return jsonify({"error": "Payment not found"}), 404
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30
    )
    elements.append(Paragraph("PAYMENT RECEIPT", title_style))
    elements.append(Spacer(1, 20))
    
    receipt_data = [
        ["Payment Number:", payment.get('payment_number', '')],
        ["Payment Date:", payment.get('payment_date', '')],
        ["Amount:", f"₹{payment.get('amount', 0):,.2f}"],
        ["Payment Method:", payment.get('payment_method', '')],
        ["Reference:", payment.get('reference_number', 'N/A')]
    ]
    
    receipt_table = Table(receipt_data, colWidths=[150, 300])
    receipt_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(receipt_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=receipt_{payment.get("payment_number", "")}.pdf'
    response.headers['Content-Type'] = 'application/pdf'
    return response

# ── Combined Report PDF ──────────────────────────────────────────────────────
@app.route('/api/reports/combined/pdf', methods=['GET'])
def combined_report_pdf():
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from datetime import datetime as dt2
    import os

    report_type = request.args.get('type', 'combined')
    start_date  = request.args.get('start', '')
    end_date    = request.args.get('end', '')
    company_name= request.args.get('company', '')

    try:
        # ── Fetch data ────────────────────────────────────────────
        if SUPABASE_CLIENT:
            try:
                bills = SUPABASE_CLIENT.table("manual_billing").select("*").limit(10000).execute().data or []
                exps  = SUPABASE_CLIENT.table("expenses").select("*").limit(10000).execute().data or []
            except Exception as e:
                logger.error(f"Supabase fetch error: {e}")
                bills = MOCK_DATABASE.get("manual_billing", [])
                exps  = MOCK_DATABASE.get("expenses", [])
        else:
            bills = MOCK_DATABASE.get("manual_billing", [])
            exps  = MOCK_DATABASE.get("expenses", [])

        # ── Apply filters ─────────────────────────────────────
        if start_date:
            bills = [r for r in bills if str(r.get('date', '')) >= start_date]
            exps  = [r for r in exps  if str(r.get('date', '')) >= start_date]
        if end_date:
            bills = [r for r in bills if str(r.get('date', '')) <= end_date]
            exps  = [r for r in exps  if str(r.get('date', '')) <= end_date]

        if company_name:
            bills = [r for r in bills if company_name.lower() in str(r.get('company_name', '')).lower()]
            if report_type == 'combined':
                report_type = 'billing' # expenses don't have company_name, hide them.

        bills.sort(key=lambda x: str(x.get('date', '')))
        exps.sort(key=lambda x: str(x.get('date', '')))

        if report_type == 'billing':
            exps = []
        elif report_type == 'expenses':
            bills = []

        # ── Totals ────────────────────────────────────────────────
        billing_total  = sum(float(r.get('total_amount', 0)) for r in bills)
        advance_total  = sum(float(r.get('advance_amount', 0)) for r in bills)
        expense_total  = sum(float(r.get('total_amount', r.get('amount', 0))) for r in exps)
        raw_mat_total  = sum(float(r.get('total_amount', 0)) for r in exps if r.get('category') == 'Raw Material')
        grand_total    = billing_total + expense_total
        
        date_range = f"{start_date} to {end_date}" if start_date and end_date else "All Records"
        if company_name:
            date_range += f"  |  Company: {company_name}"

        # ── Canvas helpers ────────────────────────────────────────
        buffer = BytesIO()
        W, H = A4
        c = pdf_canvas.Canvas(buffer, pagesize=A4)

        GOLD       = colors.HexColor('#D4AF37')
        DARK_GOLD  = colors.HexColor('#B8962E')
        BLACK      = colors.HexColor('#1a1a1a')
        WHITE      = colors.white
        LIGHT_GRAY = colors.HexColor('#f5f5f5')
        MID_GRAY   = colors.HexColor('#e0e0e0')
        GREEN      = colors.HexColor('#22c55e')
        RED        = colors.HexColor('#f87171')

        logo_path = os.path.join('static', 'images', 'logo.png')

        def draw_header(page_title="COMBINED REPORT"):
            banner_h = 90; banner_y = H - banner_h
            c.setFillColor(GOLD); c.rect(0, banner_y, W, banner_h, fill=1, stroke=0)
            # Logo box
            lbw, lbh = 80, 66; lbx = 16; lby = banner_y + (banner_h - lbh) / 2
            c.setFillColor(BLACK); c.rect(lbx, lby, lbw, lbh, fill=1, stroke=0)
            if os.path.exists(logo_path):
                try: c.drawImage(logo_path, lbx+4, lby+4, width=lbw-8, height=lbh-8, preserveAspectRatio=True, mask='auto')
                except: pass
            # Company name
            cx2 = lbx + lbw + 14; cy2 = banner_y + banner_h / 2
            c.setFont('Helvetica-Bold', 14); c.setFillColor(BLACK); c.drawString(cx2, cy2+14, "R SUNDARAM & CO")
            c.setFont('Helvetica', 8); c.drawString(cx2, cy2+2, "STRONG ROADS  BUILD TO LAST")
            c.setFont('Helvetica', 8); c.drawString(cx2, cy2-10, "READYMIX CONCRETE")
            # Proprietor
            px2 = W - 160; py2 = banner_y + 66
            c.setFont('Helvetica', 8); c.drawString(px2, py2, "Proprietor")
            c.setFont('Helvetica-Bold', 11); c.drawString(px2, py2-13, "S MAGHESH")
            c.setFont('Helvetica', 7.5)
            c.drawString(px2, py2-25, "\u25a0 9940270304")
            c.drawString(px2, py2-35, "\u25a0 rsundaram&co@gmail.com")
            c.drawString(px2, py2-45, "\u25a0 Quarry Rd, Tiruneermalai, Chennai")
            # Separator
            c.setStrokeColor(DARK_GOLD); c.setLineWidth(2); c.line(0, banner_y-3, W, banner_y-3)
            # Page title
            title_y = banner_y - 38
            c.setFont('Helvetica-Bold', 18); c.setFillColor(GOLD)
            tw = c.stringWidth(page_title, 'Helvetica-Bold', 18)
            c.drawString((W - tw) / 2, title_y, page_title)
            c.setStrokeColor(GOLD); c.setLineWidth(1.5)
            c.line((W-tw)/2, title_y-4, (W-tw)/2+tw, title_y-4)
            return title_y - 28  # return y cursor after title

        def draw_footer(page_num):
            c.setFillColor(BLACK); c.rect(0, 0, W, 30, fill=1, stroke=0)
            c.setFont('Helvetica-Bold', 7.5); c.setFillColor(GOLD)
            c.drawString(30, 18, "R SUNDARAM & CO  \u2022  READYMIX CONCRETE  \u2022  Strong Roads  \u2022  Build To Last")
            c.setFont('Helvetica', 7); c.setFillColor(colors.HexColor('#888888'))
            c.drawRightString(W-30, 18, f"Page {page_num}")
            c.setStrokeColor(GOLD); c.setLineWidth(1.5)
            c.rect(10, 10, W-20, H-20, fill=0, stroke=1)

        def stat_box(x, y, w, h, label, value, val_color=GOLD):
            c.setFillColor(colors.HexColor('#222222')); c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.roundRect(x, y, w, h, 6, fill=1, stroke=1)
            c.setFont('Helvetica', 8); c.setFillColor(colors.HexColor('#888888'))
            c.drawString(x+12, y+h-18, label)
            c.setFont('Helvetica-Bold', 13); c.setFillColor(val_color)
            c.drawString(x+12, y+10, str(value))

        # ════════════════════════════════════════════════════════
        # PAGE 1 — SUMMARY
        # ════════════════════════════════════════════════════════
        y = draw_header("COMBINED REPORT")
        page_num = 1

        # Meta strip
        c.setFont('Helvetica', 8.5); c.setFillColor(colors.HexColor('#888888'))
        c.drawString(30, y, f"Period: {date_range}   |   Generated: {dt2.now().strftime('%d %b %Y  %H:%M')}")
        y -= 28

        # Stat cards row 1
        bw = (W - 60) / 4 - 8; bh = 60
        stat_box(30,          y-bh, bw, bh, "TOTAL BILLS",           str(len(bills)),            GOLD)
        stat_box(30+bw+8,     y-bh, bw, bh, "BILLING REVENUE",       f"\u20b9 {billing_total:,.2f}", GOLD)
        stat_box(30+2*(bw+8), y-bh, bw, bh, "TOTAL EXPENSES",        str(len(exps)),              GOLD)
        stat_box(30+3*(bw+8), y-bh, bw, bh, "EXPENSE AMOUNT",        f"\u20b9 {expense_total:,.2f}", GOLD)
        y -= bh + 14

        bw2 = (W - 60) / 3 - 8
        stat_box(30,           y-bh, bw2, bh, "RAW MATERIAL SPEND",  f"\u20b9 {raw_mat_total:,.2f}", GOLD)
        stat_box(30+bw2+8,     y-bh, bw2, bh, "OTHER EXPENSES",      f"\u20b9 {expense_total-raw_mat_total:,.2f}", GOLD)
        stat_box(30+2*(bw2+8), y-bh, bw2, bh, "GRAND TOTAL",         f"\u20b9 {grand_total:,.2f}", GREEN)
        y -= bh + 22

        # Expense category breakdown
        cat_totals = {}
        for r in exps:
            cat = r.get('category', 'Other')
            cat_totals[cat] = cat_totals.get(cat, 0) + float(r.get('total_amount', r.get('amount', 0)))

        if cat_totals:
            c.setFont('Helvetica-Bold', 10); c.setFillColor(GOLD)
            c.drawString(30, y, "Expense Breakdown by Category")
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.line(30, y-4, 250, y-4)
            y -= 18
            for cat, amt in sorted(cat_totals.items(), key=lambda x: -x[1]):
                pct = (amt / expense_total * 100) if expense_total else 0
                c.setFont('Helvetica', 9); c.setFillColor(BLACK)
                c.drawString(40, y, f"\u2022  {cat}")
                c.setFont('Helvetica-Bold', 9)
                c.drawString(180, y, f"\u20b9 {amt:,.2f}")
                c.setFont('Helvetica', 9); c.setFillColor(colors.HexColor('#888888'))
                c.drawString(280, y, f"({pct:.1f}%)")
                c.setFillColor(BLACK)
                y -= 16
            y -= 8

        # Grade-wise m³ breakdown and Advance info
        grade_totals = {}
        for r in bills:
            g = str(r.get('grade', 'Unknown')).strip()
            if not g: g = 'Unknown'
            m3 = float(r.get('cubic_meters', 0))
            grade_totals[g] = grade_totals.get(g, 0) + m3

        if grade_totals or advance_total > 0:
            c.setFont('Helvetica-Bold', 10); c.setFillColor(GOLD)
            c.drawString(30, y, "Billing Summary")
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.line(30, y-4, 250, y-4)
            y -= 18
            
            c.setFont('Helvetica', 9); c.setFillColor(BLACK)
            c.drawString(40, y, "\u2022  Total Advance Given:")
            c.setFont('Helvetica-Bold', 9)
            c.drawString(180, y, f"\u20b9 {advance_total:,.2f}")
            y -= 16
            
            c.setFont('Helvetica', 9); c.setFillColor(BLACK)
            c.drawString(40, y, "\u2022  Total Volume by Grade:")
            y -= 16
            
            for g, m3 in sorted(grade_totals.items()):
                c.setFont('Helvetica', 9); c.setFillColor(colors.HexColor('#444444'))
                c.drawString(60, y, f"- {g}")
                c.setFont('Helvetica-Bold', 9); c.setFillColor(BLACK)
                c.drawString(180, y, f"{m3:,.2f} m\u00b3")
                y -= 14
            y -= 8

        draw_footer(page_num)

        # ── Shared table dimensions (used by both billing & expense pages) ──
        tbl_x = 25; row_h = 20; hdr_h = 22

        # ════════════════════════════════════════════════════════
        # PAGE 2 — BILLING RECORDS TABLE
        # ════════════════════════════════════════════════════════
        if bills:
            c.showPage(); page_num += 1
            y = draw_header("BILLING RECORDS")

            # Table header
            cols_b = [('Date', 60), ('Company', 115), ('Vehicle', 75), ('Grade', 40),
                      ('m\u00b3', 35), ('Total \u20b9', 70), ('Balance \u20b9', 70)]


            # Draw header
            c.setFillColor(GOLD); c.rect(tbl_x, y-hdr_h, W-50, hdr_h, fill=1, stroke=0)
            cx3 = tbl_x + 5
            for label, cw in cols_b:
                c.setFont('Helvetica-Bold', 8); c.setFillColor(BLACK)
                c.drawString(cx3, y-hdr_h+7, label)
                cx3 += cw
            y -= hdr_h

            for i, r in enumerate(bills):
                if y < 60:  # new page
                    draw_footer(page_num); c.showPage(); page_num += 1
                    y = draw_header("BILLING RECORDS (cont.)")

                bg = LIGHT_GRAY if i % 2 == 0 else WHITE
                c.setFillColor(bg); c.rect(tbl_x, y-row_h, W-50, row_h, fill=1, stroke=0)
                c.setStrokeColor(MID_GRAY); c.setLineWidth(0.3)
                c.rect(tbl_x, y-row_h, W-50, row_h, fill=0, stroke=1)

                vals = [
                    str(r.get('date', '')),
                    str(r.get('company_name', ''))[:18],
                    str(r.get('vehicle_number', '')),
                    str(r.get('grade', '')),
                    str(r.get('cubic_meters', '')),
                    f"\u20b9{float(r.get('total_amount',0)):,.0f}",
                    f"\u20b9{float(r.get('balance_amount',0)):,.0f}",
                ]
                cx3 = tbl_x + 5
                for j, (val, (_, cw)) in enumerate(zip(vals, cols_b)):
                    c.setFont('Helvetica', 7.5); c.setFillColor(BLACK)
                    c.drawString(cx3, y-row_h+6, val)
                    cx3 += cw
                y -= row_h

            draw_footer(page_num)

        # ════════════════════════════════════════════════════════
        # PAGE 3 — EXPENSE RECORDS TABLE
        # ════════════════════════════════════════════════════════
        if exps:
            c.showPage(); page_num += 1
            y = draw_header("EXPENSE RECORDS")

            cols_e = [('Date', 60), ('Category', 90), ('Material', 80),
                      ('Qty(T)', 45), ('Rate/T', 55), ('Total \u20b9', 70), ('Balance \u20b9', 65)]
            tbl_x = 25

            c.setFillColor(GOLD); c.rect(tbl_x, y-hdr_h, W-50, hdr_h, fill=1, stroke=0)
            cx3 = tbl_x + 5
            for label, cw in cols_e:
                c.setFont('Helvetica-Bold', 8); c.setFillColor(BLACK)
                c.drawString(cx3, y-hdr_h+7, label)
                cx3 += cw
            y -= hdr_h

            for i, r in enumerate(exps):
                if y < 60:
                    draw_footer(page_num); c.showPage(); page_num += 1
                    y = draw_header("EXPENSE RECORDS (cont.)")

                bg = LIGHT_GRAY if i % 2 == 0 else WHITE
                c.setFillColor(bg); c.rect(tbl_x, y-row_h, W-50, row_h, fill=1, stroke=0)
                c.setStrokeColor(MID_GRAY); c.setLineWidth(0.3)
                c.rect(tbl_x, y-row_h, W-50, row_h, fill=0, stroke=1)

                total   = float(r.get('total_amount', r.get('amount', 0)))
                balance = float(r.get('balance_amount', total))
                vals = [
                    str(r.get('date', '')),
                    str(r.get('category', ''))[:14],
                    str(r.get('material_type', '\u2014'))[:12],
                    str(r.get('quantity', '\u2014')),
                    f"\u20b9{float(r.get('rate_per_ton',0)):,.0f}" if r.get('rate_per_ton') else '\u2014',
                    f"\u20b9{total:,.0f}",
                    f"\u20b9{balance:,.0f}",
                ]
                cx3 = tbl_x + 5
                for val, (_, cw) in zip(vals, cols_e):
                    c.setFont('Helvetica', 7.5); c.setFillColor(BLACK)
                    c.drawString(cx3, y-row_h+6, val)
                    cx3 += cw
                y -= row_h

            draw_footer(page_num)

        c.save(); buffer.seek(0)

        resp = make_response(buffer.getvalue())
        resp.headers['Content-Disposition'] = f'inline; filename=combined_report_{dt2.now().strftime("%Y%m%d")}.pdf'
        resp.headers['Content-Type'] = 'application/pdf'
        return resp

    except Exception as e:
        logger.error(f"Combined report PDF error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500



# ── Expense Category-Filtered PDF Report ─────────────────────────────────────
@app.route('/api/expenses/category-report/pdf', methods=['GET'])
def expense_category_report_pdf():
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from datetime import datetime as dt2
    import os

    category   = request.args.get('category', '').strip()
    mat_type   = request.args.get('material',  '').strip()
    start_date = request.args.get('start',     '').strip()
    end_date   = request.args.get('end',       '').strip()

    try:
        # ── Fetch all expense records ──────────────────────────────────────
        if SUPABASE_CLIENT:
            try:
                resp = SUPABASE_CLIENT.table("expenses").select("*").execute()
                all_recs = resp.data if resp.data else []
            except Exception as e:
                logger.error(f"Supabase expense fetch error: {e}")
                all_recs = MOCK_DATABASE.get("expenses", [])
        else:
            all_recs = MOCK_DATABASE.get("expenses", [])

        # ── Apply filters ─────────────────────────────────────────────────
        if category:
            all_recs = [r for r in all_recs if r.get('category', '') == category]
        if mat_type:
            all_recs = [r for r in all_recs if r.get('material_type', '') == mat_type]
        if start_date:
            all_recs = [r for r in all_recs if str(r.get('date', '')) >= start_date]
        if end_date:
            all_recs = [r for r in all_recs if str(r.get('date', '')) <= end_date]

        # Sort by date ascending
        all_recs.sort(key=lambda x: str(x.get('date', '')))

        # Group by category for multi-category rendering
        from collections import OrderedDict
        groups = OrderedDict()
        for r in all_recs:
            cat = r.get('category', 'Other')
            groups.setdefault(cat, []).append(r)

        # ── Design tokens ─────────────────────────────────────────────────
        buffer = BytesIO()
        W, H   = A4
        c      = pdf_canvas.Canvas(buffer, pagesize=A4)

        GOLD       = colors.HexColor('#D4AF37')
        DARK_GOLD  = colors.HexColor('#B8962E')
        BLACK      = colors.HexColor('#1a1a1a')
        WHITE      = colors.white
        LIGHT_GRAY = colors.HexColor('#f5f5f5')
        MID_GRAY   = colors.HexColor('#e0e0e0')
        GREEN      = colors.HexColor('#22c55e')
        RED        = colors.HexColor('#f87171')

        logo_path = os.path.join('static', 'images', 'logo.png')

        # ── Helpers ───────────────────────────────────────────────────────
        def draw_header(title="EXPENSE CATEGORY REPORT"):
            banner_h = 90
            banner_y = H - banner_h
            c.setFillColor(GOLD)
            c.rect(0, banner_y, W, banner_h, fill=1, stroke=0)
            # Logo box
            lbw, lbh = 80, 66
            lbx = 16
            lby = banner_y + (banner_h - lbh) / 2
            c.setFillColor(BLACK)
            c.rect(lbx, lby, lbw, lbh, fill=1, stroke=0)
            if os.path.exists(logo_path):
                try:
                    c.drawImage(logo_path, lbx+4, lby+4, width=lbw-8, height=lbh-8,
                                preserveAspectRatio=True, mask='auto')
                except Exception:
                    pass
            # Company info
            cx2 = lbx + lbw + 14
            cy2 = banner_y + banner_h / 2
            c.setFont('Helvetica-Bold', 14); c.setFillColor(BLACK)
            c.drawString(cx2, cy2 + 14, "R SUNDARAM & CO")
            c.setFont('Helvetica', 8)
            c.drawString(cx2, cy2 + 2,  "STRONG ROADS  BUILD TO LAST")
            c.drawString(cx2, cy2 - 10, "READYMIX CONCRETE")
            # Proprietor
            px2 = W - 160
            py2 = banner_y + 66
            c.setFont('Helvetica', 8); c.drawString(px2, py2,      "Proprietor")
            c.setFont('Helvetica-Bold', 11); c.drawString(px2, py2 - 13, "S MAGHESH")
            c.setFont('Helvetica', 7.5)
            c.drawString(px2, py2 - 25, "\u25a0 9940270304")
            c.drawString(px2, py2 - 35, "\u25a0 rsundaram&co@gmail.com")
            c.drawString(px2, py2 - 45, "\u25a0 Quarry Rd, Tiruneermalai, Chennai")
            # Separator
            c.setStrokeColor(DARK_GOLD); c.setLineWidth(2)
            c.line(0, banner_y - 3, W, banner_y - 3)
            # Title
            title_y = banner_y - 38
            c.setFont('Helvetica-Bold', 16); c.setFillColor(GOLD)
            tw = c.stringWidth(title, 'Helvetica-Bold', 16)
            c.drawString((W - tw) / 2, title_y, title)
            c.setStrokeColor(GOLD); c.setLineWidth(1.5)
            c.line((W - tw) / 2, title_y - 4, (W - tw) / 2 + tw, title_y - 4)
            return title_y - 28

        def draw_footer(page_num):
            c.setFillColor(BLACK); c.rect(0, 0, W, 30, fill=1, stroke=0)
            c.setFont('Helvetica-Bold', 7.5); c.setFillColor(GOLD)
            c.drawString(30, 18, "R SUNDARAM & CO  \u2022  READYMIX CONCRETE  \u2022  Strong Roads  \u2022  Build To Last")
            c.setFont('Helvetica', 7); c.setFillColor(colors.HexColor('#888888'))
            c.drawRightString(W - 30, 18, f"Page {page_num}")
            c.setStrokeColor(GOLD); c.setLineWidth(1.5)
            c.rect(10, 10, W - 20, H - 20, fill=0, stroke=1)

        def stat_box(x, y, w, h, label, value, val_color=GOLD):
            c.setFillColor(colors.HexColor('#222222'))
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.roundRect(x, y, w, h, 5, fill=1, stroke=1)
            c.setFont('Helvetica', 7.5); c.setFillColor(colors.HexColor('#888888'))
            c.drawString(x + 10, y + h - 15, label)
            c.setFont('Helvetica-Bold', 12); c.setFillColor(val_color)
            c.drawString(x + 10, y + 9, str(value))

        # ── Table column spec ─────────────────────────────────────────────
        tbl_x  = 25
        row_h  = 18
        hdr_h  = 22
        cols   = [
            ('Date',        62),
            ('Description', 110),
            ('Mat. Type',   72),
            ('Qty (T)',     42),
            ('Rate/T',      52),
            ('Total \u20b9',        65),
            ('Advance \u20b9',      65),
            ('Balance \u20b9',      60),
        ]

        def draw_table_header(y):
            c.setFillColor(GOLD)
            c.rect(tbl_x, y - hdr_h, W - 50, hdr_h, fill=1, stroke=0)
            cx3 = tbl_x + 5
            for label, cw in cols:
                c.setFont('Helvetica-Bold', 7.5); c.setFillColor(BLACK)
                c.drawString(cx3, y - hdr_h + 7, label)
                cx3 += cw
            return y - hdr_h

        def draw_row(y, r, idx):
            bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
            c.setFillColor(bg)
            c.rect(tbl_x, y - row_h, W - 50, row_h, fill=1, stroke=0)
            c.setStrokeColor(MID_GRAY); c.setLineWidth(0.3)
            c.rect(tbl_x, y - row_h, W - 50, row_h, fill=0, stroke=1)

            total   = float(r.get('total_amount', r.get('amount', 0)))
            advance = float(r.get('advance_amount', 0))
            balance = float(r.get('balance_amount', total))
            qty_val = r.get('quantity', '')
            rate_val = r.get('rate_per_ton', '')

            vals = [
                str(r.get('date', '')),
                str(r.get('description', '\u2014'))[:22],
                str(r.get('material_type', '\u2014'))[:12],
                str(qty_val) if qty_val else '\u2014',
                f"\u20b9{float(rate_val):,.0f}" if rate_val else '\u2014',
                f"\u20b9{total:,.2f}",
                f"\u20b9{advance:,.2f}",
                f"\u20b9{balance:,.2f}",
            ]
            cx3 = tbl_x + 5
            for j, (val, (_, cw)) in enumerate(zip(vals, cols)):
                c.setFont('Helvetica', 7); c.setFillColor(BLACK)
                c.drawString(cx3, y - row_h + 5, val)
                cx3 += cw
            return y - row_h

        def draw_totals_row(y, total_amt, total_adv, total_bal):
            c.setFillColor(colors.HexColor('#222222'))
            c.rect(tbl_x, y - hdr_h, W - 50, hdr_h, fill=1, stroke=0)
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.rect(tbl_x, y - hdr_h, W - 50, hdr_h, fill=0, stroke=1)
            # "TOTALS" label
            c.setFont('Helvetica-Bold', 8); c.setFillColor(GOLD)
            c.drawString(tbl_x + 5, y - hdr_h + 7, "TOTALS")
            # Calculate x positions for total columns
            cx3 = tbl_x + 5
            for idx2, (_, cw) in enumerate(cols):
                if idx2 == 5:   # Total ₹
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(GOLD)
                    c.drawString(cx3, y - hdr_h + 7, f"\u20b9{total_amt:,.2f}")
                elif idx2 == 6:  # Advance ₹
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(GREEN)
                    c.drawString(cx3, y - hdr_h + 7, f"\u20b9{total_adv:,.2f}")
                elif idx2 == 7:  # Balance ₹
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(RED)
                    c.drawString(cx3, y - hdr_h + 7, f"\u20b9{total_bal:,.2f}")
                cx3 += cw
            return y - hdr_h

        # ════════════════════════════════════════════════════════════════════
        # SUMMARY PAGE
        # ════════════════════════════════════════════════════════════════════
        page_num = 1
        y = draw_header("EXPENSE CATEGORY REPORT")

        # Meta strip
        cat_label   = category  if category  else "All Categories"
        mat_label   = mat_type  if mat_type  else "All Types"
        period_lbl  = f"{start_date} to {end_date}" if start_date and end_date else "All Time"
        filter_lbl  = f"Category: {cat_label}"
        if mat_type:
            filter_lbl += f"  |  Material: {mat_label}"
        filter_lbl += f"  |  Period: {period_lbl}"

        c.setFont('Helvetica', 8.5); c.setFillColor(colors.HexColor('#888888'))
        c.drawString(30, y,       filter_lbl)
        c.drawString(30, y - 14, f"Generated: {dt2.now().strftime('%d %b %Y  %H:%M')}")
        y -= 36

        # Overall totals
        grand_total_amt = sum(float(r.get('total_amount', r.get('amount', 0))) for r in all_recs)
        grand_total_adv = sum(float(r.get('advance_amount', 0)) for r in all_recs)
        grand_total_bal = sum(float(r.get('balance_amount', 0)) for r in all_recs)
        total_count     = len(all_recs)

        bw  = (W - 60) / 4 - 8
        bh  = 56
        stat_box(30,            y - bh, bw, bh, "TOTAL RECORDS",   str(total_count),                      GOLD)
        stat_box(30 + bw + 8,   y - bh, bw, bh, "TOTAL AMOUNT",    f"\u20b9 {grand_total_amt:,.2f}",      GOLD)
        stat_box(30 + 2*(bw+8), y - bh, bw, bh, "TOTAL ADVANCE",   f"\u20b9 {grand_total_adv:,.2f}",      GREEN)
        stat_box(30 + 3*(bw+8), y - bh, bw, bh, "TOTAL BALANCE",   f"\u20b9 {grand_total_bal:,.2f}",      RED)
        y -= bh + 20

        # Category breakdown table
        if len(groups) > 1:
            c.setFont('Helvetica-Bold', 10); c.setFillColor(GOLD)
            c.drawString(30, y, "Breakdown by Category")
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.line(30, y - 4, 230, y - 4)
            y -= 20

            for gcat, grecs in groups.items():
                gamt = sum(float(r.get('total_amount', r.get('amount', 0))) for r in grecs)
                gadv = sum(float(r.get('advance_amount', 0)) for r in grecs)
                c.setFont('Helvetica', 8.5); c.setFillColor(BLACK)
                c.drawString(40, y, f"\u2022  {gcat}  ({len(grecs)} record(s))")
                c.setFont('Helvetica-Bold', 8.5)
                c.drawString(220, y, f"\u20b9 {gamt:,.2f}")
                c.setFont('Helvetica', 8.5); c.setFillColor(GREEN)
                c.drawString(320, y, f"Adv: \u20b9 {gadv:,.2f}")
                c.setFillColor(BLACK)
                y -= 16
                if y < 60:
                    draw_footer(page_num); c.showPage(); page_num += 1
                    y = draw_header("EXPENSE CATEGORY REPORT (cont.)")

        draw_footer(page_num)

        # ════════════════════════════════════════════════════════════════════
        # DETAIL PAGES — one section per category
        # ════════════════════════════════════════════════════════════════════
        for gcat, grecs in groups.items():
            c.showPage(); page_num += 1
            section_title = f"EXPENSE RECORDS — {gcat.upper()}"
            if mat_type:
                section_title += f" / {mat_type.upper()}"
            y = draw_header(section_title)

            # Category totals
            cat_total_amt = sum(float(r.get('total_amount', r.get('amount', 0))) for r in grecs)
            cat_total_adv = sum(float(r.get('advance_amount', 0)) for r in grecs)
            cat_total_bal = sum(float(r.get('balance_amount', 0)) for r in grecs)

            bw2 = (W - 60) / 3 - 8
            stat_box(30,              y - bh, bw2, bh, "TOTAL AMOUNT",   f"\u20b9 {cat_total_amt:,.2f}",  GOLD)
            stat_box(30 + bw2 + 8,   y - bh, bw2, bh, "TOTAL ADVANCE",  f"\u20b9 {cat_total_adv:,.2f}",  GREEN)
            stat_box(30 + 2*(bw2+8), y - bh, bw2, bh, "TOTAL BALANCE",  f"\u20b9 {cat_total_bal:,.2f}",  RED)
            y -= bh + 16

            # Date range info
            c.setFont('Helvetica', 8); c.setFillColor(colors.HexColor('#888888'))
            c.drawString(30, y, f"Period: {period_lbl}   |   {len(grecs)} record(s)   |   Sorted by date")
            y -= 20

            # Table
            y = draw_table_header(y)

            for idx3, r in enumerate(grecs):
                if y < 80:
                    draw_footer(page_num); c.showPage(); page_num += 1
                    y = draw_header(section_title + " (cont.)")
                    y = draw_table_header(y)
                y = draw_row(y, r, idx3)

            # Totals footer row
            if y < 80:
                draw_footer(page_num); c.showPage(); page_num += 1
                y = draw_header(section_title + " (cont.)")
            y = draw_totals_row(y, cat_total_amt, cat_total_adv, cat_total_bal)

            draw_footer(page_num)

        c.save()
        buffer.seek(0)

        safe_cat = (category or 'all').replace(' ', '_')
        fname = f"expense_report_{safe_cat}_{dt2.now().strftime('%Y%m%d')}.pdf"
        resp = make_response(buffer.getvalue())
        resp.headers['Content-Disposition'] = f'inline; filename={fname}'
        resp.headers['Content-Type'] = 'application/pdf'
        return resp

    except Exception as e:
        logger.error(f"Expense category report PDF error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ── Expense Description-Search PDF Report ────────────────────────────────────
@app.route('/api/expenses/description-report/pdf', methods=['GET'])
def expense_description_report_pdf():
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from datetime import datetime as dt2
    import os, re

    keyword    = request.args.get('keyword',  '').strip()
    category   = request.args.get('category', '').strip()
    start_date = request.args.get('start',    '').strip()
    end_date   = request.args.get('end',      '').strip()

    if not keyword:
        return jsonify({"error": "keyword is required"}), 400

    try:
        # ── Fetch ─────────────────────────────────────────────────────────
        if SUPABASE_CLIENT:
            try:
                resp = SUPABASE_CLIENT.table("expenses").select("*").execute()
                all_recs = resp.data if resp.data else []
            except Exception as e:
                logger.error(f"Supabase expense fetch error: {e}")
                all_recs = MOCK_DATABASE.get("expenses", [])
        else:
            all_recs = MOCK_DATABASE.get("expenses", [])

        # ── Filter ────────────────────────────────────────────────────────
        kw_lower = keyword.lower()
        search_terms = [kw_lower]
        if kw_lower == 'labour': search_terms.append('labor')
        elif kw_lower == 'labor': search_terms.append('labour')

        all_recs = [
            r for r in all_recs
            if any(
                term in (r.get('description') or '').lower() or
                term in (r.get('material_type') or '').lower() or
                term in (r.get('category') or '').lower()
                for term in search_terms
            )
        ]
        if category:
            all_recs = [r for r in all_recs if r.get('category', '') == category]
        if start_date:
            all_recs = [r for r in all_recs if str(r.get('date', '')) >= start_date]
        if end_date:
            all_recs = [r for r in all_recs if str(r.get('date', '')) <= end_date]

        # Sort by date ascending
        all_recs.sort(key=lambda x: str(x.get('date', '')))

        # ── Design tokens ─────────────────────────────────────────────────
        buffer = BytesIO()
        W, H   = A4
        c      = pdf_canvas.Canvas(buffer, pagesize=A4)

        GOLD       = colors.HexColor('#D4AF37')
        DARK_GOLD  = colors.HexColor('#B8962E')
        BLACK      = colors.HexColor('#1a1a1a')
        WHITE      = colors.white
        LIGHT_GRAY = colors.HexColor('#f5f5f5')
        MID_GRAY   = colors.HexColor('#e0e0e0')
        GREEN      = colors.HexColor('#22c55e')
        RED        = colors.HexColor('#f87171')
        AMBER      = colors.HexColor('#fbbf24')

        logo_path = os.path.join('static', 'images', 'logo.png')

        def draw_header(title="EXPENSE DESCRIPTION REPORT"):
            banner_h = 90
            banner_y = H - banner_h
            c.setFillColor(GOLD)
            c.rect(0, banner_y, W, banner_h, fill=1, stroke=0)
            lbw, lbh = 80, 66
            lbx = 16
            lby = banner_y + (banner_h - lbh) / 2
            c.setFillColor(BLACK)
            c.rect(lbx, lby, lbw, lbh, fill=1, stroke=0)
            if os.path.exists(logo_path):
                try:
                    c.drawImage(logo_path, lbx+4, lby+4, width=lbw-8, height=lbh-8,
                                preserveAspectRatio=True, mask='auto')
                except Exception:
                    pass
            cx2 = lbx + lbw + 14
            cy2 = banner_y + banner_h / 2
            c.setFont('Helvetica-Bold', 14); c.setFillColor(BLACK)
            c.drawString(cx2, cy2 + 14, "R SUNDARAM & CO")
            c.setFont('Helvetica', 8)
            c.drawString(cx2, cy2 + 2,  "STRONG ROADS  BUILD TO LAST")
            c.drawString(cx2, cy2 - 10, "READYMIX CONCRETE")
            px2 = W - 160
            py2 = banner_y + 66
            c.setFont('Helvetica', 8); c.drawString(px2, py2, "Proprietor")
            c.setFont('Helvetica-Bold', 11); c.drawString(px2, py2 - 13, "S MAGHESH")
            c.setFont('Helvetica', 7.5)
            c.drawString(px2, py2 - 25, "\u25a0 9940270304")
            c.drawString(px2, py2 - 35, "\u25a0 rsundaram&co@gmail.com")
            c.drawString(px2, py2 - 45, "\u25a0 Quarry Rd, Tiruneermalai, Chennai")
            c.setStrokeColor(DARK_GOLD); c.setLineWidth(2)
            c.line(0, banner_y - 3, W, banner_y - 3)
            title_y = banner_y - 38
            c.setFont('Helvetica-Bold', 15); c.setFillColor(GOLD)
            tw = c.stringWidth(title, 'Helvetica-Bold', 15)
            c.drawString((W - tw) / 2, title_y, title)
            c.setStrokeColor(GOLD); c.setLineWidth(1.5)
            c.line((W - tw) / 2, title_y - 4, (W - tw) / 2 + tw, title_y - 4)
            return title_y - 28

        def draw_footer(page_num):
            c.setFillColor(BLACK); c.rect(0, 0, W, 30, fill=1, stroke=0)
            c.setFont('Helvetica-Bold', 7.5); c.setFillColor(GOLD)
            c.drawString(30, 18, "R SUNDARAM & CO  \u2022  READYMIX CONCRETE  \u2022  Strong Roads  \u2022  Build To Last")
            c.setFont('Helvetica', 7); c.setFillColor(colors.HexColor('#888888'))
            c.drawRightString(W - 30, 18, f"Page {page_num}")
            c.setStrokeColor(GOLD); c.setLineWidth(1.5)
            c.rect(10, 10, W - 20, H - 20, fill=0, stroke=1)

        def stat_box(x, y, w, h, label, value, val_color=GOLD):
            c.setFillColor(colors.HexColor('#222222'))
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.roundRect(x, y, w, h, 5, fill=1, stroke=1)
            c.setFont('Helvetica', 7.5); c.setFillColor(colors.HexColor('#888888'))
            c.drawString(x + 10, y + h - 15, label)
            c.setFont('Helvetica-Bold', 12); c.setFillColor(val_color)
            c.drawString(x + 10, y + 9, str(value))

        # ── Column spec ───────────────────────────────────────────────────
        tbl_x = 25
        row_h = 20
        hdr_h = 22
        cols  = [
            ('Date',        58),
            ('Category',    75),
            ('Description', 120),
            ('Material',    65),
            ('Qty (T)',     40),
            ('Rate/T',      50),
            ('Total \u20b9',    62),
            ('Advance \u20b9',  62),
            ('Balance \u20b9',  58),
        ]

        def draw_table_header(y):
            c.setFillColor(GOLD)
            c.rect(tbl_x, y - hdr_h, W - 50, hdr_h, fill=1, stroke=0)
            cx3 = tbl_x + 5
            for label, cw in cols:
                c.setFont('Helvetica-Bold', 7.5); c.setFillColor(BLACK)
                c.drawString(cx3, y - hdr_h + 7, label)
                cx3 += cw
            return y - hdr_h

        def draw_row(y, r, idx):
            bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
            c.setFillColor(bg)
            c.rect(tbl_x, y - row_h, W - 50, row_h, fill=1, stroke=0)
            c.setStrokeColor(MID_GRAY); c.setLineWidth(0.3)
            c.rect(tbl_x, y - row_h, W - 50, row_h, fill=0, stroke=1)

            total   = float(r.get('total_amount',   r.get('amount', 0)))
            advance = float(r.get('advance_amount', 0))
            balance = float(r.get('balance_amount', total))
            qty_val  = r.get('quantity', '')
            rate_val = r.get('rate_per_ton', '')
            desc_raw = str(r.get('description', '\u2014'))[:26]

            vals = [
                str(r.get('date', '')),
                str(r.get('category', '\u2014'))[:12],
                desc_raw,
                str(r.get('material_type', '\u2014'))[:10],
                str(qty_val) if qty_val else '\u2014',
                f"\u20b9{float(rate_val):,.0f}" if rate_val else '\u2014',
                f"\u20b9{total:,.2f}",
                f"\u20b9{advance:,.2f}",
                f"\u20b9{balance:,.2f}",
            ]
            cx3 = tbl_x + 5
            for j, (val, (_, cw)) in enumerate(zip(vals, cols)):
                # Colour amount columns
                if j == 6:
                    c.setFont('Helvetica-Bold', 7); c.setFillColor(BLACK)
                elif j == 7:
                    c.setFont('Helvetica', 7); c.setFillColor(GREEN)
                elif j == 8:
                    c.setFont('Helvetica', 7); c.setFillColor(RED)
                else:
                    c.setFont('Helvetica', 7); c.setFillColor(BLACK)
                c.drawString(cx3, y - row_h + 6, val)
                cx3 += cw
            return y - row_h

        def draw_totals_row(y, total_amt, total_adv, total_bal):
            c.setFillColor(colors.HexColor('#1a1a1a'))
            c.rect(tbl_x, y - hdr_h, W - 50, hdr_h, fill=1, stroke=0)
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.rect(tbl_x, y - hdr_h, W - 50, hdr_h, fill=0, stroke=1)
            c.setFont('Helvetica-Bold', 8); c.setFillColor(GOLD)
            c.drawString(tbl_x + 5, y - hdr_h + 7, "TOTALS")
            cx3 = tbl_x + 5
            for idx2, (_, cw) in enumerate(cols):
                if idx2 == 6:
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(GOLD)
                    c.drawString(cx3, y - hdr_h + 7, f"\u20b9{total_amt:,.2f}")
                elif idx2 == 7:
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(GREEN)
                    c.drawString(cx3, y - hdr_h + 7, f"\u20b9{total_adv:,.2f}")
                elif idx2 == 8:
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(RED)
                    c.drawString(cx3, y - hdr_h + 7, f"\u20b9{total_bal:,.2f}")
                cx3 += cw
            return y - hdr_h

        # ════════════════════════════════════════════════════════════════════
        # PAGE 1 — SUMMARY
        # ════════════════════════════════════════════════════════════════════
        page_num = 1
        y = draw_header("EXPENSE DESCRIPTION REPORT")

        # Totals
        grand_amt = sum(float(r.get('total_amount', r.get('amount', 0))) for r in all_recs)
        grand_adv = sum(float(r.get('advance_amount', 0)) for r in all_recs)
        grand_bal = sum(float(r.get('balance_amount', 0)) for r in all_recs)
        count     = len(all_recs)

        cat_lbl    = category   if category   else "All Categories"
        period_lbl = f"{start_date} to {end_date}" if start_date and end_date else "All Time"

        # Meta
        c.setFont('Helvetica', 8.5); c.setFillColor(colors.HexColor('#888888'))
        c.drawString(30, y,       f"Keyword: \"{keyword}\"   |   Category: {cat_lbl}   |   Period: {period_lbl}")
        c.drawString(30, y - 14, f"Generated: {dt2.now().strftime('%d %b %Y  %H:%M')}   |   {count} record(s) matched")
        y -= 38

        # Keyword badge
        kw_box_w = c.stringWidth(f'  \"{keyword}\"  ', 'Helvetica-Bold', 11) + 20
        c.setFillColor(colors.HexColor('#D4AF3333'))
        c.setStrokeColor(GOLD); c.setLineWidth(1)
        c.roundRect(30, y - 24, kw_box_w, 22, 4, fill=1, stroke=1)
        c.setFont('Helvetica-Bold', 11); c.setFillColor(GOLD)
        c.drawString(40, y - 16, f'"{keyword}"')
        c.setFont('Helvetica', 9); c.setFillColor(colors.HexColor('#888888'))
        c.drawString(30 + kw_box_w + 10, y - 16, "← matched in description / material type")
        y -= 38

        # Stat boxes
        bw = (W - 60) / 4 - 8
        bh = 56
        stat_box(30,            y - bh, bw, bh, "MATCHED RECORDS", str(count),                          GOLD)
        stat_box(30 + bw + 8,   y - bh, bw, bh, "TOTAL AMOUNT",    f"\u20b9 {grand_amt:,.2f}",           GOLD)
        stat_box(30 + 2*(bw+8), y - bh, bw, bh, "TOTAL ADVANCE",   f"\u20b9 {grand_adv:,.2f}",           GREEN)
        stat_box(30 + 3*(bw+8), y - bh, bw, bh, "TOTAL BALANCE",   f"\u20b9 {grand_bal:,.2f}",           RED)
        y -= bh + 20

        # Category breakdown (if multiple categories present)
        from collections import Counter
        cat_counts = Counter(r.get('category', 'Other') for r in all_recs)
        if len(cat_counts) > 1:
            c.setFont('Helvetica-Bold', 10); c.setFillColor(GOLD)
            c.drawString(30, y, "Matched Records by Category")
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.line(30, y - 4, 270, y - 4)
            y -= 18
            for gcat, cnt in cat_counts.most_common():
                gamt = sum(float(r.get('total_amount', r.get('amount', 0))) for r in all_recs if r.get('category') == gcat)
                gadv = sum(float(r.get('advance_amount', 0)) for r in all_recs if r.get('category') == gcat)
                c.setFont('Helvetica', 8.5); c.setFillColor(BLACK)
                c.drawString(40, y, f"\u2022  {gcat}  ({cnt} record(s))")
                c.setFont('Helvetica-Bold', 8.5)
                c.drawString(200, y, f"\u20b9 {gamt:,.2f}")
                c.setFont('Helvetica', 8.5); c.setFillColor(GREEN)
                c.drawString(310, y, f"Adv: \u20b9 {gadv:,.2f}")
                c.setFillColor(BLACK)
                y -= 15
                if y < 60:
                    draw_footer(page_num); c.showPage(); page_num += 1
                    y = draw_header("EXPENSE DESCRIPTION REPORT (cont.)")

        draw_footer(page_num)

        # ════════════════════════════════════════════════════════════════════
        # PAGE 2+ — DETAIL TABLE
        # ════════════════════════════════════════════════════════════════════
        if all_recs:
            c.showPage(); page_num += 1
            y = draw_header(f'MATCHED RECORDS — "{keyword}"')

            # Stat summary strip
            bw2 = (W - 60) / 3 - 8
            stat_box(30,              y - bh, bw2, bh, "TOTAL AMOUNT",  f"\u20b9 {grand_amt:,.2f}", GOLD)
            stat_box(30 + bw2 + 8,   y - bh, bw2, bh, "TOTAL ADVANCE", f"\u20b9 {grand_adv:,.2f}", GREEN)
            stat_box(30 + 2*(bw2+8), y - bh, bw2, bh, "TOTAL BALANCE", f"\u20b9 {grand_bal:,.2f}", RED)
            y -= bh + 14

            c.setFont('Helvetica', 8); c.setFillColor(colors.HexColor('#888888'))
            c.drawString(30, y, f"Keyword: \"{keyword}\"   |   {count} record(s)   |   Period: {period_lbl}   |   Sorted by date")
            y -= 20

            y = draw_table_header(y)

            for idx3, r in enumerate(all_recs):
                if y < 80:
                    draw_footer(page_num); c.showPage(); page_num += 1
                    y = draw_header(f'MATCHED RECORDS — "{keyword}" (cont.)')
                    y = draw_table_header(y)
                y = draw_row(y, r, idx3)

            if y < 80:
                draw_footer(page_num); c.showPage(); page_num += 1
                y = draw_header(f'MATCHED RECORDS — "{keyword}" (cont.)')
            y = draw_totals_row(y, grand_amt, grand_adv, grand_bal)

            draw_footer(page_num)

        c.save()
        buffer.seek(0)

        safe_kw = re.sub(r'[^a-zA-Z0-9_-]', '_', keyword)[:30]
        fname = f"expense_desc_{safe_kw}_{dt2.now().strftime('%Y%m%d')}.pdf"
        resp = make_response(buffer.getvalue())
        resp.headers['Content-Disposition'] = f'inline; filename={fname}'
        resp.headers['Content-Type'] = 'application/pdf'
        return resp

    except Exception as e:
        logger.error(f"Expense description report PDF error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500
# ── Company Billing PDF Report ─────────────────────────────────────────────
@app.route('/api/reports/company-billing/pdf', methods=['GET'])
def company_billing_report_pdf():
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from datetime import datetime as dt2
    import os, re

    company_kw = request.args.get('company', '').strip()
    start_date = request.args.get('start', '').strip()
    end_date   = request.args.get('end', '').strip()

    if not company_kw:
        return jsonify({"error": "company name keyword is required"}), 400

    try:
        # Fetch billing records
        if SUPABASE_CLIENT:
            try:
                resp = SUPABASE_CLIENT.table("manual_billing").select("*").limit(10000).execute()
                all_bills = resp.data if resp.data else []
            except Exception as e:
                logger.error(f"Supabase billing fetch error: {e}")
                all_bills = MOCK_DATABASE.get("manual_billing", [])
        else:
            all_bills = MOCK_DATABASE.get("manual_billing", [])

        # Resolve a spelling variation to one company, then filter by that exact
        # normalized name.  Per-word fuzzy matching could merge different customers.
        import difflib
        def normalize_company_name(value):
            return re.sub(r'[^a-z0-9]', '', str(value or '').lower())

        wanted_name = normalize_company_name(company_kw)
        company_names = {str(b.get('company_name', '')).strip() for b in all_bills if b.get('company_name')}
        best_name = ''
        best_score = 0.0
        for name in company_names:
            normalized_name = normalize_company_name(name)
            score = difflib.SequenceMatcher(None, wanted_name, normalized_name).ratio()
            if normalized_name == wanted_name:
                best_name, best_score = name, 1.0
                break
            if score > best_score:
                best_name, best_score = name, score

        all_bills = [b for b in all_bills if best_score >= 0.70 and
                     normalize_company_name(b.get('company_name')) == normalize_company_name(best_name)]
        
        if start_date:
            all_bills = [b for b in all_bills if str(b.get('date', '')) >= start_date]
        if end_date:
            all_bills = [b for b in all_bills if str(b.get('date', '')) <= end_date]

        all_bills.sort(key=lambda x: str(x.get('date', '')))

        # Design Tokens
        buffer = BytesIO()
        W, H   = A4
        c      = pdf_canvas.Canvas(buffer, pagesize=A4)

        GOLD       = colors.HexColor('#D4AF37')
        DARK_GOLD  = colors.HexColor('#B8962E')
        BLACK      = colors.HexColor('#1a1a1a')
        WHITE      = colors.white
        LIGHT_GRAY = colors.HexColor('#f5f5f5')
        MID_GRAY   = colors.HexColor('#e0e0e0')
        GREEN      = colors.HexColor('#22c55e')
        RED        = colors.HexColor('#f87171')
        AMBER      = colors.HexColor('#fbbf24')

        logo_path = os.path.join('static', 'images', 'logo.png')

        def draw_header(title="COMPANY BILLING REPORT"):
            banner_h = 90
            banner_y = H - banner_h
            c.setFillColor(GOLD)
            c.rect(0, banner_y, W, banner_h, fill=1, stroke=0)
            lbw, lbh = 80, 66
            lbx = 16
            lby = banner_y + (banner_h - lbh) / 2
            c.setFillColor(BLACK)
            c.rect(lbx, lby, lbw, lbh, fill=1, stroke=0)
            if os.path.exists(logo_path):
                try:
                    c.drawImage(logo_path, lbx+4, lby+4, width=lbw-8, height=lbh-8,
                                preserveAspectRatio=True, mask='auto')
                except Exception:
                    pass
            cx2 = lbx + lbw + 14
            cy2 = banner_y + banner_h / 2
            c.setFont('Helvetica-Bold', 14); c.setFillColor(BLACK)
            c.drawString(cx2, cy2 + 14, "R SUNDARAM & CO")
            c.setFont('Helvetica', 8)
            c.drawString(cx2, cy2 + 2,  "STRONG ROADS  BUILD TO LAST")
            c.drawString(cx2, cy2 - 10, "READYMIX CONCRETE")
            px2 = W - 160
            py2 = banner_y + 66
            c.setFont('Helvetica', 8); c.drawString(px2, py2, "Proprietor")
            c.setFont('Helvetica-Bold', 11); c.drawString(px2, py2 - 13, "S MAGHESH")
            c.setFont('Helvetica', 7.5)
            c.drawString(px2, py2 - 25, "\u25a0 9940270304")
            c.drawString(px2, py2 - 35, "\u25a0 rsundaram&co@gmail.com")
            c.drawString(px2, py2 - 45, "\u25a0 Quarry Rd, Tiruneermalai, Chennai")
            c.setStrokeColor(DARK_GOLD); c.setLineWidth(2)
            c.line(0, banner_y - 3, W, banner_y - 3)
            title_y = banner_y - 38
            c.setFont('Helvetica-Bold', 15); c.setFillColor(GOLD)
            tw = c.stringWidth(title, 'Helvetica-Bold', 15)
            c.drawString((W - tw) / 2, title_y, title)
            c.setStrokeColor(GOLD); c.setLineWidth(1.5)
            c.line((W - tw) / 2, title_y - 4, (W - tw) / 2 + tw, title_y - 4)
            return title_y - 28

        def draw_footer(page_num):
            c.setFillColor(BLACK); c.rect(0, 0, W, 30, fill=1, stroke=0)
            c.setFont('Helvetica-Bold', 7.5); c.setFillColor(GOLD)
            c.drawString(30, 18, "R SUNDARAM & CO  \u2022  READYMIX CONCRETE  \u2022  Strong Roads  \u2022  Build To Last")
            c.setFont('Helvetica', 7); c.setFillColor(colors.HexColor('#888888'))
            c.drawRightString(W - 30, 18, f"Page {page_num}")
            c.setStrokeColor(GOLD); c.setLineWidth(1.5)
            c.rect(10, 10, W - 20, H - 20, fill=0, stroke=1)

        def stat_box(x, y, w, h, label, value, val_color=GOLD):
            c.setFillColor(colors.HexColor('#222222'))
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.roundRect(x, y, w, h, 5, fill=1, stroke=1)
            c.setFont('Helvetica', 7.5); c.setFillColor(colors.HexColor('#888888'))
            c.drawString(x + 10, y + h - 15, label)
            c.setFont('Helvetica-Bold', 12); c.setFillColor(val_color)
            c.drawString(x + 10, y + 9, str(value))

        # ── Table setup ───────────────────────────────────────────────────
        tbl_x = 20
        row_h = 20
        hdr_h = 22
        cols  = [
            ('Date',       55),
            ('Company',    140),
            ('Grade',      50),
            ('Qty (m\u00b3)', 50),
            ('Rate',       55),
            ('Total \u20b9',   65),
            ('Advance \u20b9', 65),
            ('Balance \u20b9', 65),
        ]

        def draw_table_header(y):
            c.setFillColor(GOLD)
            c.rect(tbl_x, y - hdr_h, W - 40, hdr_h, fill=1, stroke=0)
            cx3 = tbl_x + 5
            for label, cw in cols:
                c.setFont('Helvetica-Bold', 8); c.setFillColor(BLACK)
                c.drawString(cx3, y - hdr_h + 7, label)
                cx3 += cw
            return y - hdr_h

        def draw_row(y, r, idx):
            bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
            c.setFillColor(bg)
            c.rect(tbl_x, y - row_h, W - 40, row_h, fill=1, stroke=0)
            c.setStrokeColor(MID_GRAY); c.setLineWidth(0.3)
            c.rect(tbl_x, y - row_h, W - 40, row_h, fill=0, stroke=1)

            m3      = float(r.get('cubic_meters', 0))
            rate    = float(r.get('rate_per_cubic', 0))
            total   = float(r.get('total_amount', 0))
            advance = float(r.get('advance_amount', 0))
            balance = float(r.get('balance_amount', 0))

            vals = [
                str(r.get('date', '')),
                str(r.get('company_name', '\u2014'))[:28],
                str(r.get('grade', '\u2014'))[:10],
                f"{m3:.2f}",
                f"\u20b9{rate:,.0f}",
                f"\u20b9{total:,.2f}",
                f"\u20b9{advance:,.2f}",
                f"\u20b9{balance:,.2f}",
            ]
            
            cx3 = tbl_x + 5
            for j, (val, (_, cw)) in enumerate(zip(vals, cols)):
                if j == 5:
                    c.setFont('Helvetica-Bold', 7.5); c.setFillColor(BLACK)
                elif j == 6:
                    c.setFont('Helvetica-Bold', 7.5); c.setFillColor(GREEN)
                elif j == 7:
                    c.setFont('Helvetica-Bold', 7.5); c.setFillColor(RED)
                else:
                    c.setFont('Helvetica', 7.5); c.setFillColor(BLACK)
                c.drawString(cx3, y - row_h + 6, val)
                cx3 += cw
            return y - row_h

        def draw_totals_row(y, t_m3, t_amt, t_adv, t_bal):
            c.setFillColor(colors.HexColor('#1a1a1a'))
            c.rect(tbl_x, y - hdr_h, W - 40, hdr_h, fill=1, stroke=0)
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.rect(tbl_x, y - hdr_h, W - 40, hdr_h, fill=0, stroke=1)
            c.setFont('Helvetica-Bold', 8); c.setFillColor(GOLD)
            c.drawString(tbl_x + 5, y - hdr_h + 7, "TOTALS")
            
            cx3 = tbl_x + 5
            for idx2, (_, cw) in enumerate(cols):
                if idx2 == 3:
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(GOLD)
                    c.drawString(cx3, y - hdr_h + 7, f"{t_m3:.2f}")
                elif idx2 == 5:
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(GOLD)
                    c.drawString(cx3, y - hdr_h + 7, f"\u20b9{t_amt:,.2f}")
                elif idx2 == 6:
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(GREEN)
                    c.drawString(cx3, y - hdr_h + 7, f"\u20b9{t_adv:,.2f}")
                elif idx2 == 7:
                    c.setFont('Helvetica-Bold', 8); c.setFillColor(RED)
                    c.drawString(cx3, y - hdr_h + 7, f"\u20b9{t_bal:,.2f}")
                cx3 += cw
            return y - hdr_h

        # ════════════════════════════════════════════════════════════════════
        # PAGE 1 — SUMMARY
        # ════════════════════════════════════════════════════════════════════
        page_num = 1
        y = draw_header("COMPANY BILLING REPORT")

        grand_m3  = sum(float(r.get('cubic_meters', 0)) for r in all_bills)
        grand_amt = sum(float(r.get('total_amount', 0)) for r in all_bills)
        grand_adv = sum(float(r.get('advance_amount', 0)) for r in all_bills)
        grand_bal = sum(float(r.get('balance_amount', 0)) for r in all_bills)
        count     = len(all_bills)

        period_lbl = f"{start_date} to {end_date}" if start_date and end_date else "All Time"

        c.setFont('Helvetica', 8.5); c.setFillColor(colors.HexColor('#888888'))
        c.drawString(30, y,       f"Company Keyword: \"{company_kw}\"   |   Period: {period_lbl}")
        c.drawString(30, y - 14,  f"Generated: {dt2.now().strftime('%d %b %Y  %H:%M')}   |   {count} record(s) matched")
        y -= 38

        # Keyword badge
        kw_box_w = c.stringWidth(f'  \"{company_kw}\"  ', 'Helvetica-Bold', 11) + 20
        c.setFillColor(colors.HexColor('#D4AF3333'))
        c.setStrokeColor(GOLD); c.setLineWidth(1)
        c.roundRect(30, y - 24, kw_box_w, 22, 4, fill=1, stroke=1)
        c.setFont('Helvetica-Bold', 11); c.setFillColor(GOLD)
        c.drawString(40, y - 16, f'"{company_kw}"')
        y -= 38

        # Stat boxes
        bw = (W - 60) / 4 - 8
        bh = 56
        stat_box(30,            y - bh, bw, bh, "TOTAL M\u00b3",     f"{grand_m3:,.2f}",      GOLD)
        stat_box(30 + bw + 8,   y - bh, bw, bh, "TOTAL AMOUNT",    f"\u20b9 {grand_amt:,.2f}", GOLD)
        stat_box(30 + 2*(bw+8), y - bh, bw, bh, "TOTAL ADVANCE",   f"\u20b9 {grand_adv:,.2f}", GREEN)
        stat_box(30 + 3*(bw+8), y - bh, bw, bh, "TOTAL BALANCE",   f"\u20b9 {grand_bal:,.2f}", RED)
        y -= bh + 20

        # Grade breakdown
        from collections import Counter
        grade_counts = Counter(r.get('grade', 'Other') for r in all_bills)
        if len(grade_counts) > 0:
            c.setFont('Helvetica-Bold', 10); c.setFillColor(GOLD)
            c.drawString(30, y, "Matched Records by Grade")
            c.setStrokeColor(GOLD); c.setLineWidth(0.8)
            c.line(30, y - 4, 300, y - 4)
            y -= 18
            for grade, cnt in grade_counts.most_common():
                gm3  = sum(float(r.get('cubic_meters', 0)) for r in all_bills if r.get('grade') == grade)
                gamt = sum(float(r.get('total_amount', 0)) for r in all_bills if r.get('grade') == grade)
                c.setFont('Helvetica', 8.5); c.setFillColor(BLACK)
                c.drawString(40, y, f"\u2022  Grade {grade}  ({cnt} record(s))")
                c.setFont('Helvetica-Bold', 8.5)
                c.drawString(180, y, f"{gm3:,.2f} m\u00b3")
                c.drawString(250, y, f"\u20b9 {gamt:,.2f}")
                y -= 15
                if y < 60:
                    draw_footer(page_num); c.showPage(); page_num += 1
                    y = draw_header("COMPANY BILLING REPORT (cont.)")

        draw_footer(page_num)

        # ════════════════════════════════════════════════════════════════════
        # PAGE 2+ — DETAIL TABLE
        # ════════════════════════════════════════════════════════════════════
        if all_bills:
            c.showPage(); page_num += 1
            y = draw_header(f'BILLING RECORDS — "{company_kw}"')

            bw2 = (W - 60) / 4 - 8
            stat_box(30,              y - bh, bw2, bh, "TOTAL M\u00b3",    f"{grand_m3:,.2f}",      GOLD)
            stat_box(30 + bw2 + 8,   y - bh, bw2, bh, "TOTAL AMOUNT",  f"\u20b9 {grand_amt:,.2f}", GOLD)
            stat_box(30 + 2*(bw2+8), y - bh, bw2, bh, "TOTAL ADVANCE", f"\u20b9 {grand_adv:,.2f}", GREEN)
            stat_box(30 + 3*(bw2+8), y - bh, bw2, bh, "TOTAL BALANCE", f"\u20b9 {grand_bal:,.2f}", RED)
            y -= bh + 14

            c.setFont('Helvetica', 8); c.setFillColor(colors.HexColor('#888888'))
            c.drawString(30, y, f"Company: \"{company_kw}\"   |   {count} record(s)   |   Sorted by date")
            y -= 20

            y = draw_table_header(y)

            for idx3, r in enumerate(all_bills):
                if y < 80:
                    draw_footer(page_num); c.showPage(); page_num += 1
                    y = draw_header(f'BILLING RECORDS — "{company_kw}" (cont.)')
                    y = draw_table_header(y)
                y = draw_row(y, r, idx3)

            if y < 80:
                draw_footer(page_num); c.showPage(); page_num += 1
                y = draw_header(f'BILLING RECORDS — "{company_kw}" (cont.)')
            y = draw_totals_row(y, grand_m3, grand_amt, grand_adv, grand_bal)

            draw_footer(page_num)

        c.save()
        buffer.seek(0)

        safe_kw = re.sub(r'[^a-zA-Z0-9_-]', '_', company_kw)[:30]
        fname = f"company_billing_{safe_kw}_{dt2.now().strftime('%Y%m%d')}.pdf"
        resp = make_response(buffer.getvalue())
        resp.headers['Content-Disposition'] = f'inline; filename={fname}'
        resp.headers['Content-Type'] = 'application/pdf'
        return resp

    except Exception as e:
        logger.error(f"Company billing report PDF error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ── All Companies Summary PDF ──────────────────────────────────────────────
@app.route('/api/reports/all-companies-summary/pdf', methods=['GET'])
def all_companies_summary_pdf():
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from datetime import datetime as dt2
    from collections import defaultdict
    import os, re

    start_date = request.args.get('start', '').strip()
    end_date   = request.args.get('end',   '').strip()

    try:
        # Fetch all billing records
        if SUPABASE_CLIENT:
            try:
                resp = SUPABASE_CLIENT.table("manual_billing").select("*").limit(10000).execute()
                all_bills = resp.data if resp.data else []
            except Exception as e:
                logger.error(f"Supabase billing fetch error: {e}")
                all_bills = MOCK_DATABASE.get("manual_billing", [])
        else:
            all_bills = MOCK_DATABASE.get("manual_billing", [])

        # Date filter
        if start_date:
            all_bills = [b for b in all_bills if str(b.get('date', '')) >= start_date]
        if end_date:
            all_bills = [b for b in all_bills if str(b.get('date', '')) <= end_date]

        # Group by normalised company name
        groups = defaultdict(lambda: {'count': 0, 'm3': 0.0, 'amount': 0.0, 'advance': 0.0, 'balance': 0.0})
        for b in all_bills:
            key = (b.get('company_name') or '—').strip().upper()
            groups[key]['count']   += 1
            groups[key]['m3']      += float(b.get('cubic_meters',   0) or 0)
            groups[key]['amount']  += float(b.get('total_amount',   0) or 0)
            groups[key]['advance'] += float(b.get('advance_amount', 0) or 0)
            groups[key]['balance'] += float(b.get('balance_amount', 0) or 0)

        sorted_companies = sorted(groups.items(), key=lambda x: x[0])

        # Totals
        grand_m3      = sum(v['m3']      for _, v in sorted_companies)
        grand_amount  = sum(v['amount']  for _, v in sorted_companies)
        grand_advance = sum(v['advance'] for _, v in sorted_companies)
        grand_balance = sum(v['balance'] for _, v in sorted_companies)

        # ── PDF Setup ──────────────────────────────────────────────────────
        buffer = BytesIO()
        W, H   = landscape(A4)  # landscape for wide table
        c      = pdf_canvas.Canvas(buffer, pagesize=landscape(A4))

        GOLD      = colors.HexColor('#D4AF37')
        BLACK     = colors.HexColor('#1a1a1a')
        WHITE     = colors.white
        LT_GRAY   = colors.HexColor('#f5f5f5')
        MID_GRAY  = colors.HexColor('#e0e0e0')
        GREEN     = colors.HexColor('#22c55e')
        RED       = colors.HexColor('#ef4444')
        DARK_BG   = colors.HexColor('#2a2a2a')

        logo_path = os.path.join('static', 'images', 'logo.png')
        period_str = f"{start_date or 'All'} to {end_date or 'Date'}"

        COL_X     = [30, 200, 330, 410, 500, 610, 700]  # column x positions
        COL_W     = [170, 130, 80,  90, 110, 90,  80]   # column widths

        def fmt_inr(v):
            return f"\u20b9{v:,.2f}"

        def draw_header():
            banner_h = 80
            banner_y = H - banner_h
            c.setFillColor(GOLD)
            c.rect(0, banner_y, W, banner_h, fill=1, stroke=0)

            lbw, lbh = 70, 58
            lbx = 14
            lby = banner_y + (banner_h - lbh) / 2
            c.setFillColor(BLACK)
            c.rect(lbx, lby, lbw, lbh, fill=1, stroke=0)
            if os.path.exists(logo_path):
                try:
                    c.drawImage(logo_path, lbx+3, lby+3, width=lbw-6, height=lbh-6,
                                preserveAspectRatio=True, mask='auto')
                except Exception:
                    pass

            cx2 = lbx + lbw + 12
            cy2 = banner_y + banner_h / 2
            c.setFont('Helvetica-Bold', 13); c.setFillColor(BLACK)
            c.drawString(cx2, cy2 + 12, "R SUNDARAM & CO")
            c.setFont('Helvetica', 9)
            c.drawString(cx2, cy2 - 2, "Overall Company Billing Summary Report")
            c.setFont('Helvetica', 8)
            c.drawString(cx2, cy2 - 14, f"Period: {period_str}   |   Generated: {dt2.now().strftime('%d-%b-%Y %I:%M %p')}")

            # right side stats
            rx = W - 260
            c.setFont('Helvetica-Bold', 10); c.setFillColor(BLACK)
            c.drawString(rx, cy2 + 14, f"Companies: {len(sorted_companies)}")
            c.drawString(rx, cy2 + 0,  f"Total Deliveries: {len(all_bills)}")
            c.drawString(rx, cy2 - 14, f"Total M\u00b3: {grand_m3:.2f}")

        def draw_table_header(y):
            headers  = ['Company Name', 'Deliveries', 'Total M\u00b3', 'Total Amount', 'Advance (₹)', 'Balance (₹)']
            col_xs   = COL_X[1:]
            col_ws   = COL_W[1:]
            row_h    = 20
            c.setFillColor(GOLD)
            c.rect(COL_X[0], y - row_h, W - COL_X[0] - 14, row_h, fill=1, stroke=0)
            c.setFont('Helvetica-Bold', 8); c.setFillColor(BLACK)
            # # column
            c.drawString(COL_X[0] + 4, y - row_h + 6, '#')
            for i, (hdr, cx, cw) in enumerate(zip(headers, col_xs, col_ws)):
                c.drawString(cx, y - row_h + 6, hdr)
            return y - row_h

        def draw_footer(page_num, total_pages=None):
            c.setFont('Helvetica', 7); c.setFillColor(colors.HexColor('#888888'))
            footer_txt = f"R Sundaram & Co  |  All Companies Summary  |  Page {page_num}"
            c.drawString(30, 18, footer_txt)
            c.drawRightString(W - 30, 18, dt2.now().strftime('%d-%b-%Y'))

        # ── Draw Pages ─────────────────────────────────────────────────────
        PAGE_TOP    = H - 95
        PAGE_BOTTOM = 40
        ROW_H       = 17
        page_num    = 1

        draw_header()
        y = draw_table_header(PAGE_TOP)

        for idx, (company, vals) in enumerate(sorted_companies):
            if y - ROW_H < PAGE_BOTTOM:
                draw_footer(page_num)
                c.showPage()
                page_num += 1
                draw_header()
                y = draw_table_header(PAGE_TOP)

            bg = LT_GRAY if idx % 2 == 0 else WHITE
            c.setFillColor(bg)
            c.rect(COL_X[0], y - ROW_H, W - COL_X[0] - 14, ROW_H, fill=1, stroke=0)

            c.setFont('Helvetica', 8); c.setFillColor(BLACK)
            c.drawString(COL_X[0] + 4, y - ROW_H + 5, str(idx + 1))
            c.setFont('Helvetica-Bold', 8)
            # truncate long names
            disp_name = company[:32] + '…' if len(company) > 32 else company
            c.drawString(COL_X[1], y - ROW_H + 5, disp_name)
            c.setFont('Helvetica', 8)
            c.drawString(COL_X[2], y - ROW_H + 5, str(vals['count']))
            c.drawString(COL_X[3], y - ROW_H + 5, f"{vals['m3']:.2f}")
            c.setFillColor(BLACK)
            c.drawString(COL_X[4], y - ROW_H + 5, fmt_inr(vals['amount']))
            c.setFillColor(GREEN)
            c.drawString(COL_X[5], y - ROW_H + 5, fmt_inr(vals['advance']))
            c.setFillColor(RED if vals['balance'] > 0 else GREEN)
            c.drawString(COL_X[6], y - ROW_H + 5, fmt_inr(vals['balance']))

            y -= ROW_H

        # Grand totals row
        if y - ROW_H * 2 < PAGE_BOTTOM:
            draw_footer(page_num)
            c.showPage()
            page_num += 1
            draw_header()
            y = draw_table_header(PAGE_TOP)

        y -= 4
        c.setFillColor(GOLD)
        c.rect(COL_X[0], y - ROW_H, W - COL_X[0] - 14, ROW_H, fill=1, stroke=0)
        c.setFont('Helvetica-Bold', 9); c.setFillColor(BLACK)
        c.drawString(COL_X[0] + 4, y - ROW_H + 5, 'GRAND TOTALS')
        c.drawString(COL_X[2], y - ROW_H + 5, str(len(all_bills)))
        c.drawString(COL_X[3], y - ROW_H + 5, f"{grand_m3:.2f}")
        c.drawString(COL_X[4], y - ROW_H + 5, fmt_inr(grand_amount))
        c.drawString(COL_X[5], y - ROW_H + 5, fmt_inr(grand_advance))
        c.drawString(COL_X[6], y - ROW_H + 5, fmt_inr(grand_balance))

        draw_footer(page_num)
        c.save()
        buffer.seek(0)

        fname = f"all_companies_summary_{dt2.now().strftime('%Y%m%d')}.pdf"
        rsp = make_response(buffer.getvalue())
        rsp.headers['Content-Disposition'] = f'inline; filename={fname}'
        rsp.headers['Content-Type'] = 'application/pdf'
        return rsp

    except Exception as e:
        logger.error(f"All companies summary PDF error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ── Client Billing Summary helpers ─────────────────────────────────────────
def _normalize_client_name(value):
    import re
    return re.sub(r'[^a-z0-9]', '', str(value or '').lower())


def _resolve_client_name(keyword, records):
    """Resolve a search keyword to one company_name (exact / contains / strict fuzzy)."""
    import difflib
    wanted = _normalize_client_name(keyword)
    if not wanted:
        return '', 0.0
    names = {str(b.get('company_name', '')).strip() for b in records if b.get('company_name')}

    # 1) Exact normalized match
    for name in names:
        if _normalize_client_name(name) == wanted:
            return name, 1.0

    # 2) Substring / contains match (prefer longest company name)
    contains = []
    for name in names:
        normalized = _normalize_client_name(name)
        if wanted in normalized or normalized in wanted:
            contains.append((name, normalized))
    if contains:
        contains.sort(key=lambda x: len(x[1]), reverse=True)
        return contains[0][0], 0.95

    # 3) Strict fuzzy (avoid accidental merges like ABC → SRI CONSTRUCTIONS)
    best_name, best_score = '', 0.0
    for name in names:
        normalized = _normalize_client_name(name)
        score = difflib.SequenceMatcher(None, wanted, normalized).ratio()
        if score > best_score:
            best_name, best_score = name, score
    min_score = 0.88 if len(wanted) >= 6 else 0.92
    if best_score >= min_score:
        return best_name, best_score
    return '', 0.0


def _collect_billed_transaction_ids():
    """IDs already settled via client_billing_summaries (works even before billing_status column exists)."""
    import json as _json
    billed = set()
    for row in _fetch_client_billing_summaries():
        raw = row.get('transaction_ids') or '[]'
        try:
            ids = _json.loads(raw) if isinstance(raw, str) else (raw or [])
        except Exception:
            ids = []
        for i in ids:
            if i:
                billed.add(i)
    return billed


def _is_unbilled_record(record, billed_ids=None):
    status = str(record.get('billing_status') or 'unbilled').strip().lower()
    if status == 'billed':
        return False
    if billed_ids is not None and record.get('id') in billed_ids:
        return False
    return True


def _is_advance_carry_record(record):
    return str(record.get('grade') or '').strip().upper() in ('ADV-CF', 'ADVANCE CF', 'ADVANCE-CF')


def _is_invoice_record(record):
    return str(record.get('grade') or '').strip().upper() == 'CBS-INV'


def _summary_from_invoice_row(row):
    """Rebuild a client_billing_summaries-shaped dict from a durable CBS-INV manual_billing row."""
    import json as _json
    meta = {}
    raw_addr = row.get('address') or ''
    try:
        if isinstance(raw_addr, dict):
            meta = raw_addr
        elif str(raw_addr).strip().startswith('{'):
            meta = _json.loads(raw_addr)
    except Exception:
        meta = {}
    txn_ids = meta.get('transaction_ids') or []
    return {
        "id": row.get('id'),
        "invoice_number": meta.get('invoice_number') or row.get('vehicle_number') or row.get('id'),
        "client_name": row.get('company_name'),
        "bill_date": row.get('date') or meta.get('bill_date'),
        "total_amount": float(row.get('total_amount') or meta.get('total_amount') or 0),
        "advance_deducted": float(row.get('advance_amount') or meta.get('advance_deducted') or 0),
        "balance_amount": float(row.get('balance_amount') or meta.get('balance_amount') or 0),
        "advance_carried_forward": float(meta.get('advance_carried_forward') or 0),
        "transaction_ids": _json.dumps(txn_ids) if not isinstance(txn_ids, str) else txn_ids,
        "pdf_path": meta.get('pdf_path'),
        "pdf_filename": meta.get('pdf_filename'),
        "notes": meta.get('notes') or row.get('driver_name'),
        "created_at": row.get('created_at'),
        "created_by": row.get('created_by'),
    }


def _fetch_all_manual_billing():
    if SUPABASE_CLIENT:
        try:
            resp = SUPABASE_CLIENT.table("manual_billing").select("*").limit(10000).execute()
            return resp.data if resp.data else []
        except Exception as e:
            logger.error(f"Supabase billing fetch error: {e}")
            return MOCK_DATABASE.get("manual_billing", [])
    return MOCK_DATABASE.get("manual_billing", [])


def _fetch_all_expenses():
    if SUPABASE_CLIENT:
        try:
            resp = SUPABASE_CLIENT.table("expenses").select("*").limit(10000).execute()
            return resp.data if resp.data else []
        except Exception as e:
            logger.error(f"Supabase expenses fetch error: {e}")
            return MOCK_DATABASE.get("expenses", [])
    return MOCK_DATABASE.get("expenses", [])


def _normalize_supplier_name(value):
    import re
    return re.sub(r'[^a-z0-9]', '', str(value or '').lower())


def _resolve_supplier_name(keyword, records):
    import difflib
    wanted = _normalize_supplier_name(keyword)
    if not wanted:
        return '', 0.0
    names = {str(r.get('company_name', '')).strip() for r in records if r.get('company_name')}

    for name in names:
        if _normalize_supplier_name(name) == wanted:
            return name, 1.0

    contains = []
    for name in names:
        normalized = _normalize_supplier_name(name)
        if wanted in normalized or normalized in wanted:
            contains.append((name, normalized))
    if contains:
        contains.sort(key=lambda x: len(x[1]), reverse=True)
        return contains[0][0], 0.95

    fuzzy_name = ''
    fuzzy_score = 1.0
    for name in names:
        normalized = _normalize_supplier_name(name)
        if not normalized:
            continue
        ratio = difflib.SequenceMatcher(a=wanted, b=normalized).ratio()
        if ratio > fuzzy_score:
            fuzzy_score = ratio
            fuzzy_name = name
    if fuzzy_score >= 0.75:
        return fuzzy_name, fuzzy_score

    return '', 0.0


def _fetch_supplier_payment_summaries():
    rows = []
    if SUPABASE_CLIENT:
        try:
            resp = SUPABASE_CLIENT.table("supplier_payment_summaries").select("*").limit(5000).execute()
            if resp.data:
                rows.extend(resp.data)
        except Exception as e:
            logger.error(f"Supabase supplier payment summaries fetch error: {e}")
    rows.extend(MOCK_DATABASE.get("supplier_payment_summaries", []))

    seen = set()
    unique = []
    for r in rows:
        rid = r.get('id')
        if not rid or rid in seen:
            continue
        seen.add(rid)
        unique.append(r)
    return unique


def _update_expense_record(expense_id, updates):
    if SUPABASE_CLIENT:
        try:
            SUPABASE_CLIENT.table("expenses").update(updates).eq("id", expense_id).execute()
            return True
        except Exception as e:
            logger.error(f"Supabase expense update error for {expense_id}: {e}")
            return False
    for record in MOCK_DATABASE.get("expenses", []):
        if record.get('id') == expense_id:
            record.update(updates)
            save_db()
            return True
    return False


def _insert_supplier_payment_summary(record):
    import json as _json
    payload = dict(record)
    if SUPABASE_CLIENT:
        try:
            resp = SUPABASE_CLIENT.table("supplier_payment_summaries").insert(payload).execute()
            if resp.data:
                return resp.data[0]
        except Exception as e:
            logger.error(f"Supabase supplier payment summary insert error: {e}")
            MOCK_DATABASE.setdefault("supplier_payment_summaries", []).append(payload)
            save_db()
            return payload
    MOCK_DATABASE.setdefault("supplier_payment_summaries", []).append(payload)
    save_db()
    return payload


def _compute_supplier_payment_summary(supplier_kw, start_date='', end_date='', include_paid=False):
    all_expenses = _fetch_all_expenses()
    matched_name, score = _resolve_supplier_name(supplier_kw, all_expenses)
    if not matched_name:
        return {
            "success": True,
            "supplier_name": supplier_kw,
            "matched": False,
            "transactions": [],
            "total_amount": 0,
            "advance_amount": 0,
            "balance_amount": 0,
            "advance_carried_forward": 0,
            "balance_label": "Balance Amount",
            "message": f'No supplier found matching "{supplier_kw}"'
        }

    supplier_records = [
        r for r in all_expenses
        if _normalize_supplier_name(r.get('company_name')) == _normalize_supplier_name(matched_name)
        and str(r.get('category', '')).strip().lower() == 'raw material'
    ]
    if start_date:
        supplier_records = [r for r in supplier_records if str(r.get('date', '')) >= start_date]
    if end_date:
        supplier_records = [r for r in supplier_records if str(r.get('date', '')) <= end_date]

    if not include_paid:
        supplier_records = [r for r in supplier_records if str(r.get('supplier_payment_status', 'unpaid')).strip().lower() != 'paid']

    supplier_records.sort(key=lambda r: str(r.get('date', '')))

    total_amount = sum(float(r.get('total_amount') or 0) for r in supplier_records)
    advance_amount = sum(float(r.get('advance_amount') or 0) for r in supplier_records)
    if advance_amount > total_amount:
        advance_deducted = total_amount
        balance_amount = 0.0
        advance_carried_forward = round(advance_amount - total_amount, 2)
        balance_label = "Advance Carried Forward"
        display_balance = advance_carried_forward
    else:
        advance_deducted = advance_amount
        balance_amount = round(total_amount - advance_amount, 2)
        advance_carried_forward = 0.0
        balance_label = "Balance Amount"
        display_balance = balance_amount

    transactions = []
    for idx, record in enumerate(supplier_records, start=1):
        transactions.append({
            "id": record.get('id'),
            "reference": record.get('id'),
            "material_type": record.get('material_type') or '—',
            "quantity": float(record.get('quantity') or 0),
            "price": float(record.get('total_amount') or 0),
            "delivery_date": record.get('date') or '',
            "advance": float(record.get('advance_amount') or 0),
            "status": record.get('supplier_payment_status') or 'unpaid'
        })

    return {
        "success": True,
        "matched": True,
        "supplier_name": matched_name,
        "match_score": round(score, 3),
        "transactions": transactions,
        "transaction_ids": [t['id'] for t in transactions],
        "total_amount": round(total_amount, 2),
        "advance_amount": round(advance_amount, 2),
        "advance_deducted": round(advance_deducted, 2),
        "balance_amount": round(balance_amount, 2),
        "advance_carried_forward": round(advance_carried_forward, 2),
        "display_balance": round(display_balance, 2),
        "balance_label": balance_label,
        "record_count": len(transactions),
        "period": {
            "start": start_date or None,
            "end": end_date or None
        }
    }


def _build_supplier_payment_pdf(summary, statement_number, statement_date):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from datetime import datetime as dt2
    import re

    buffer = BytesIO()
    W, H = A4
    c = pdf_canvas.Canvas(buffer, pagesize=A4)

    GOLD = colors.HexColor('#D4AF37')
    DARK = colors.HexColor('#1a1a1a')
    GRAY = colors.HexColor('#555555')
    GREEN = colors.HexColor('#16a34a')
    AMBER = colors.HexColor('#d97706')
    RED = colors.HexColor('#dc2626')

    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'images', 'logo.png')
    supplier_name = summary.get('supplier_name', '')
    transactions = summary.get('transactions', [])

    def draw_header():
        c.setFillColor(GOLD)
        c.rect(0, H - 90, W, 90, fill=1, stroke=0)
        c.setFillColor(DARK)
        if os.path.exists(logo_path):
            try:
                c.drawImage(logo_path, 30, H - 80, width=70, height=70, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        c.setFont('Helvetica-Bold', 18)
        c.drawString(110, H - 50, 'SUPPLIER PAYMENT STATEMENT')
        c.setFont('Helvetica', 9)
        c.drawString(110, H - 65, 'R SUNDARAM & CO | Quarry Rd, Tiruneermalai, Chennai')
        c.drawString(110, H - 78, 'Phone: 9940270304 | Email: rsundaram&co@gmail.com')

    def draw_footer(page_num):
        c.setFont('Helvetica', 7)
        c.setFillColor(GRAY)
        c.drawString(30, 20, f'Generated: {dt2.now().strftime("%d %b %Y %H:%M")}')
        c.drawRightString(W - 30, 20, f'Page {page_num}')

    draw_header()
    c.setFont('Helvetica-Bold', 10)
    c.setFillColor(DARK)
    c.drawString(30, H - 115, f'Supplier: {supplier_name}')
    c.drawString(30, H - 130, f'Statement No: {statement_number}')
    c.drawString(30, H - 145, f'Statement Date: {statement_date}')
    period_text = 'Period: '
    if summary.get('period', {}).get('start') and summary.get('period', {}).get('end'):
        period_text += f"{summary['period']['start']} to {summary['period']['end']}"
    else:
        period_text += 'All time'
    c.setFont('Helvetica', 9)
    c.setFillColor(GRAY)
    c.drawString(30, H - 160, period_text)

    table_top = H - 190
    row_height = 20
    c.setFont('Helvetica-Bold', 9)
    c.setFillColor(GOLD)
    cols = [30, 90, 240, 360, 430, 520]
    headers = ['#', 'Ref / ID', 'Material', 'Qty', 'Price (₹)', 'Delivery Date']
    for idx, header in enumerate(headers):
        c.drawString(cols[idx], table_top, header)
    c.setStrokeColor(GOLD)
    c.line(30, table_top - 4, W - 30, table_top - 4)

    c.setFont('Helvetica', 9)
    y = table_top - 22
    for i, tx in enumerate(transactions, start=1):
        if y < 90:
            draw_footer(1)
            c.showPage()
            draw_header()
            y = H - 100
        c.drawString(cols[0], y, str(i))
        c.drawString(cols[1], y, str(tx.get('reference', '—')))
        c.drawString(cols[2], y, str(tx.get('material_type', '—')))
        c.drawRightString(cols[3] + 40, y, f"{tx.get('quantity', 0):,.2f}")
        c.drawRightString(cols[4] + 40, y, f"{tx.get('price', 0):,.2f}")
        c.drawString(cols[5], y, str(tx.get('delivery_date', '—')))
        y -= row_height

    if y < 140:
        c.showPage()
        draw_header()
        y = H - 110

    c.setLineWidth(1)
    c.setStrokeColor(GOLD)
    c.line(30, y - 8, W - 30, y - 8)
    y -= 18
    c.setFont('Helvetica-Bold', 10)
    c.drawString(30, y, f"Total Amount: ₹ {summary.get('total_amount', 0):,.2f}")
    y -= 16
    c.setFont('Helvetica', 10)
    c.setFillColor(GREEN)
    c.drawString(30, y, f"Advance Deducted: ₹ {summary.get('advance_amount', 0):,.2f}")
    y -= 14
    if summary.get('advance_carried_forward', 0) > 0:
        c.setFillColor(AMBER)
        c.drawString(30, y, f"Advance Carried Forward: ₹ {summary.get('advance_carried_forward', 0):,.2f}")
    else:
        c.setFillColor(RED)
        c.drawString(30, y, f"Balance Payable: ₹ {summary.get('balance_amount', 0):,.2f}")

    draw_footer(1)
    c.save()
    buffer.seek(0)

    safe = re.sub(r'[^a-zA-Z0-9_-]', '_', supplier_name)[:24]
    fname = f"supplier_statement_{safe}_{statement_number}.pdf"
    return buffer.getvalue(), fname


def _insert_client_billing_summary(record):
    rows = []
    if SUPABASE_CLIENT:
        try:
            resp = SUPABASE_CLIENT.table("client_billing_summaries").select("*").limit(5000).execute()
            if resp.data:
                rows.extend(resp.data)
        except Exception as e:
            logger.error(f"Supabase client_billing_summaries fetch error: {e}")
    rows.extend(MOCK_DATABASE.get("client_billing_summaries", []))

    # Durable fallback invoices stored as manual_billing grade=CBS-INV
    for b in _fetch_all_manual_billing():
        if _is_invoice_record(b):
            rows.append(_summary_from_invoice_row(b))

    # Dedupe by id (dedicated table wins over fallback)
    seen = set()
    unique = []
    for r in rows:
        rid = r.get('id')
        if not rid or rid in seen:
            continue
        seen.add(rid)
        unique.append(r)
    return unique


def _update_manual_billing_record(billing_id, updates):
    """Update a manual_billing row. Falls back if settlement columns are not migrated yet."""
    if SUPABASE_CLIENT:
        try:
            SUPABASE_CLIENT.table("manual_billing").update(updates).eq("id", billing_id).execute()
            return True
        except Exception as e:
            logger.error(f"Supabase billing update error for {billing_id}: {e}")
            # Retry without settlement-only columns (pre-migration databases)
            safe = {k: v for k, v in updates.items()
                    if k not in ('billing_status', 'billing_summary_id')}
            if safe:
                try:
                    SUPABASE_CLIENT.table("manual_billing").update(safe).eq("id", billing_id).execute()
                    logger.warning(
                        "Updated without billing_status columns — run migrate_client_billing.sql when possible"
                    )
                    return True
                except Exception as e2:
                    logger.error(f"Supabase safe billing update error for {billing_id}: {e2}")
            return False
    for record in MOCK_DATABASE.get("manual_billing", []):
        if record.get('id') == billing_id:
            record.update(updates)
            save_db()
            return True
    return False


def _insert_manual_billing_record(record):
    payload = dict(record)
    if SUPABASE_CLIENT:
        try:
            resp = SUPABASE_CLIENT.table("manual_billing").insert(payload).execute()
            if resp.data:
                return resp.data[0]
            return payload
        except Exception as e:
            logger.error(f"Supabase billing insert error: {e}")
            # Retry without settlement-only columns
            safe = {k: v for k, v in payload.items()
                    if k not in ('billing_status', 'billing_summary_id')}
            try:
                resp = SUPABASE_CLIENT.table("manual_billing").insert(safe).execute()
                if resp.data:
                    return resp.data[0]
                return safe
            except Exception as e2:
                logger.error(f"Supabase safe billing insert error: {e2}")
                MOCK_DATABASE.setdefault("manual_billing", []).append(payload)
                save_db()
                return payload
    MOCK_DATABASE.setdefault("manual_billing", []).append(payload)
    save_db()
    return payload


def _insert_client_billing_summary(record):
    """Insert into client_billing_summaries when available, and always mirror as CBS-INV row."""
    import json as _json
    if SUPABASE_CLIENT:
        try:
            resp = SUPABASE_CLIENT.table("client_billing_summaries").insert(record).execute()
            if resp.data:
                record = resp.data[0]
        except Exception as e:
            logger.error(f"Supabase client_billing_summaries insert error: {e}")
            MOCK_DATABASE.setdefault("client_billing_summaries", []).append(record)
            save_db()
    else:
        MOCK_DATABASE.setdefault("client_billing_summaries", []).append(record)
        save_db()

    # Durable fallback that works without the new table (uses existing manual_billing)
    try:
        txn_ids = record.get('transaction_ids') or '[]'
        if isinstance(txn_ids, str):
            txn_list = _json.loads(txn_ids)
        else:
            txn_list = list(txn_ids)
    except Exception:
        txn_list = []

    meta = {
        "invoice_number": record.get('invoice_number'),
        "bill_date": record.get('bill_date'),
        "transaction_ids": txn_list,
        "advance_deducted": record.get('advance_deducted'),
        "balance_amount": record.get('balance_amount'),
        "advance_carried_forward": record.get('advance_carried_forward'),
        "pdf_path": record.get('pdf_path'),
        "pdf_filename": record.get('pdf_filename'),
        "notes": record.get('notes'),
    }
    inv_row = {
        "id": record.get('id'),
        "company_name": record.get('client_name'),
        "phone": "",
        "address": _json.dumps(meta, ensure_ascii=False),
        "date": record.get('bill_date') or datetime.now().strftime('%Y-%m-%d'),
        "time": datetime.now().strftime('%I:%M %p'),
        "vehicle_number": record.get('invoice_number') or record.get('id'),
        "driver_name": "Client Bill",
        "unloading_time": "N/A",
        "grade": "CBS-INV",
        "cubic_meters": 0,
        "loading_time": "N/A",
        "rate_per_cubic": 0,
        "total_amount": record.get('total_amount') or 0,
        "advance_amount": record.get('advance_deducted') or 0,
        "balance_amount": record.get('balance_amount') or 0,
        "created_at": record.get('created_at') or datetime.now().isoformat(),
        "created_by": record.get('created_by') or 'system',
    }
    # Avoid duplicating if this id already exists as CBS-INV
    existing_ids = {b.get('id') for b in _fetch_all_manual_billing()}
    if inv_row['id'] not in existing_ids:
        _insert_manual_billing_record(inv_row)
    return record

def _compute_client_billing_summary(client_kw, start_date='', end_date='', include_billed=False):
    """Return summary dict for a client: loads, totals, advance, balance / carry-forward."""
    all_bills = _fetch_all_manual_billing()
    matched_name, score = _resolve_client_name(client_kw, all_bills)
    if not matched_name:
        return {
            "success": True,
            "client_name": client_kw,
            "matched": False,
            "transactions": [],
            "total_amount": 0,
            "advance_amount": 0,
            "balance_amount": 0,
            "advance_carried_forward": 0,
            "balance_label": "Balance Amount",
            "message": f'No client found matching "{client_kw}"'
        }

    billed_ids = _collect_billed_transaction_ids()
    client_bills = [
        b for b in all_bills
        if _normalize_client_name(b.get('company_name')) == _normalize_client_name(matched_name)
        and not _is_invoice_record(b)
    ]
    if start_date:
        client_bills = [b for b in client_bills if str(b.get('date', '')) >= start_date]
    if end_date:
        client_bills = [b for b in client_bills if str(b.get('date', '')) <= end_date]

    if not include_billed:
        client_bills = [b for b in client_bills if _is_unbilled_record(b, billed_ids)]

    client_bills.sort(key=lambda x: str(x.get('date', '')))

    # Advances sum across ALL unbilled rows (including carry-forward credits)
    advance_amount = sum(float(b.get('advance_amount') or 0) for b in client_bills)

    # Load lines exclude pure advance-carry rows
    load_rows = [b for b in client_bills if not _is_advance_carry_record(b)]
    total_amount = sum(float(b.get('total_amount') or 0) for b in load_rows)

    if advance_amount > total_amount:
        advance_deducted = total_amount
        balance_amount = 0.0
        advance_carried_forward = round(advance_amount - total_amount, 2)
        balance_label = "Advance Carried Forward"
        display_balance = advance_carried_forward
    else:
        advance_deducted = advance_amount
        balance_amount = round(total_amount - advance_amount, 2)
        advance_carried_forward = 0.0
        balance_label = "Balance Amount"
        display_balance = balance_amount

    transactions = []
    for idx, b in enumerate(load_rows, start=1):
        transactions.append({
            "id": b.get('id'),
            "load_no": idx,
            "m_number": b.get('grade') or '—',
            "delivery_date": b.get('date') or '',
            "price": float(b.get('total_amount') or 0),
            "cubic_meters": float(b.get('cubic_meters') or 0),
            "rate_per_cubic": float(b.get('rate_per_cubic') or 0),
            "advance_amount": float(b.get('advance_amount') or 0),
            "billing_status": b.get('billing_status') or 'unbilled',
            "vehicle_number": b.get('vehicle_number') or '',
        })

    phone = next((b.get('phone') for b in client_bills if b.get('phone')), '')
    address = next((b.get('address') for b in client_bills if b.get('address')), '')

    return {
        "success": True,
        "matched": True,
        "client_name": matched_name,
        "match_score": round(score, 3),
        "phone": phone,
        "address": address,
        "transactions": transactions,
        "transaction_ids": [t['id'] for t in transactions],
        "advance_credit_ids": [b.get('id') for b in client_bills if _is_advance_carry_record(b)],
        "total_amount": round(total_amount, 2),
        "advance_amount": round(advance_amount, 2),
        "advance_deducted": round(advance_deducted, 2),
        "balance_amount": round(balance_amount, 2),
        "advance_carried_forward": round(advance_carried_forward, 2),
        "display_balance": round(display_balance, 2),
        "balance_label": balance_label,
        "load_count": len(transactions),
        "period": {
            "start": start_date or None,
            "end": end_date or None
        }
    }


def _build_client_billing_pdf(summary, invoice_number, bill_date):
    """Build a client billing invoice PDF (ReportLab). Returns (bytes, filename)."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from datetime import datetime as dt2
    import re

    buffer = BytesIO()
    W, H = A4
    c = pdf_canvas.Canvas(buffer, pagesize=A4)

    GOLD = colors.HexColor('#D4AF37')
    DARK_GOLD = colors.HexColor('#B8962E')
    BLACK = colors.HexColor('#1a1a1a')
    WHITE = colors.white
    LIGHT_GRAY = colors.HexColor('#f5f5f5')
    MID_GRAY = colors.HexColor('#e0e0e0')
    GREEN = colors.HexColor('#16a34a')
    RED = colors.HexColor('#dc2626')
    AMBER = colors.HexColor('#d97706')

    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'images', 'logo.png')
    client_name = summary.get('client_name', '')
    transactions = summary.get('transactions', [])

    def draw_header(title="CLIENT BILLING INVOICE"):
        banner_h = 90
        banner_y = H - banner_h
        c.setFillColor(GOLD)
        c.rect(0, banner_y, W, banner_h, fill=1, stroke=0)
        lbw, lbh = 80, 66
        lbx, lby = 16, banner_y + (banner_h - lbh) / 2
        c.setFillColor(BLACK)
        c.rect(lbx, lby, lbw, lbh, fill=1, stroke=0)
        if os.path.exists(logo_path):
            try:
                c.drawImage(logo_path, lbx + 4, lby + 4, width=lbw - 8, height=lbh - 8,
                            preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        cx2 = lbx + lbw + 14
        cy2 = banner_y + banner_h / 2
        c.setFont('Helvetica-Bold', 14)
        c.setFillColor(BLACK)
        c.drawString(cx2, cy2 + 14, "R SUNDARAM & CO")
        c.setFont('Helvetica', 8)
        c.drawString(cx2, cy2 + 2, "STRONG ROADS  BUILD TO LAST")
        c.drawString(cx2, cy2 - 10, "READYMIX CONCRETE")
        px2 = W - 160
        py2 = banner_y + 66
        c.setFont('Helvetica', 8)
        c.drawString(px2, py2, "Proprietor")
        c.setFont('Helvetica-Bold', 11)
        c.drawString(px2, py2 - 13, "S MAGHESH")
        c.setFont('Helvetica', 7.5)
        c.drawString(px2, py2 - 25, "\u25a0 9940270304")
        c.drawString(px2, py2 - 35, "\u25a0 rsundaram&co@gmail.com")
        c.drawString(px2, py2 - 45, "\u25a0 Quarry Rd, Tiruneermalai, Chennai")
        c.setStrokeColor(DARK_GOLD)
        c.setLineWidth(2)
        c.line(0, banner_y - 3, W, banner_y - 3)
        title_y = banner_y - 38
        c.setFont('Helvetica-Bold', 15)
        c.setFillColor(GOLD)
        tw = c.stringWidth(title, 'Helvetica-Bold', 15)
        c.drawString((W - tw) / 2, title_y, title)
        c.setStrokeColor(GOLD)
        c.setLineWidth(1.5)
        c.line((W - tw) / 2, title_y - 4, (W - tw) / 2 + tw, title_y - 4)
        return title_y - 28

    def draw_footer(page_num):
        c.setFillColor(BLACK)
        c.rect(0, 0, W, 30, fill=1, stroke=0)
        c.setFont('Helvetica-Bold', 7.5)
        c.setFillColor(GOLD)
        c.drawString(30, 18, "R SUNDARAM & CO  \u2022  READYMIX CONCRETE  \u2022  Strong Roads  \u2022  Build To Last")
        c.setFont('Helvetica', 7)
        c.setFillColor(colors.HexColor('#888888'))
        c.drawRightString(W - 30, 18, f"Page {page_num}")
        c.setStrokeColor(GOLD)
        c.setLineWidth(1.5)
        c.rect(10, 10, W - 20, H - 20, fill=0, stroke=1)

    page_num = 1
    y = draw_header()

    # Meta block
    c.setFont('Helvetica', 9)
    c.setFillColor(BLACK)
    c.drawString(30, y, f"Invoice No: {invoice_number}")
    c.drawRightString(W - 30, y, f"Bill Date: {bill_date}")
    y -= 16
    c.setFont('Helvetica-Bold', 11)
    c.setFillColor(GOLD)
    c.drawString(30, y, f"Client: {client_name}")
    y -= 14
    if summary.get('phone') or summary.get('address'):
        c.setFont('Helvetica', 8)
        c.setFillColor(colors.HexColor('#555555'))
        meta_bits = []
        if summary.get('phone'):
            meta_bits.append(str(summary['phone']))
        if summary.get('address'):
            meta_bits.append(str(summary['address'])[:60])
        c.drawString(30, y, "  |  ".join(meta_bits))
        y -= 18
    else:
        y -= 8

    # Summary cards
    box_w = (W - 70) / 3
    box_h = 48
    cards = [
        ("TOTAL AMOUNT", f"\u20b9 {summary['total_amount']:,.2f}", GOLD),
        ("ADVANCE DEDUCTED", f"\u20b9 {summary['advance_deducted']:,.2f}", GREEN),
        (summary['balance_label'].upper(), f"\u20b9 {summary['display_balance']:,.2f}",
         AMBER if summary['advance_carried_forward'] > 0 else RED),
    ]
    for i, (label, value, color) in enumerate(cards):
        x = 30 + i * (box_w + 5)
        c.setFillColor(colors.HexColor('#222222'))
        c.setStrokeColor(GOLD)
        c.setLineWidth(0.8)
        c.roundRect(x, y - box_h, box_w, box_h, 5, fill=1, stroke=1)
        c.setFont('Helvetica', 7)
        c.setFillColor(colors.HexColor('#888888'))
        c.drawString(x + 10, y - 14, label)
        c.setFont('Helvetica-Bold', 11)
        c.setFillColor(color)
        c.drawString(x + 10, y - box_h + 12, value)
    y -= box_h + 22

    # Table
    cols = [
        ('#', 28),
        ('M#', 55),
        ('Delivery Date', 90),
        ('Qty (m\u00b3)', 70),
        ('Rate', 80),
        ('Price \u20b9', 100),
    ]
    tbl_x = 30
    row_h, hdr_h = 18, 20

    def draw_table_header(yy):
        c.setFillColor(GOLD)
        c.rect(tbl_x, yy - hdr_h, W - 60, hdr_h, fill=1, stroke=0)
        cx = tbl_x + 6
        for label, cw in cols:
            c.setFont('Helvetica-Bold', 8)
            c.setFillColor(BLACK)
            c.drawString(cx, yy - hdr_h + 6, label)
            cx += cw
        return yy - hdr_h

    y = draw_table_header(y)
    for idx, t in enumerate(transactions):
        if y < 70:
            draw_footer(page_num)
            c.showPage()
            page_num += 1
            y = draw_header("CLIENT BILLING INVOICE (cont.)")
            y = draw_table_header(y)
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        c.setFillColor(bg)
        c.rect(tbl_x, y - row_h, W - 60, row_h, fill=1, stroke=0)
        c.setStrokeColor(MID_GRAY)
        c.setLineWidth(0.3)
        c.rect(tbl_x, y - row_h, W - 60, row_h, fill=0, stroke=1)
        vals = [
            str(t.get('load_no', idx + 1)),
            str(t.get('m_number', '—')),
            str(t.get('delivery_date', '')),
            f"{float(t.get('cubic_meters') or 0):.2f}",
            f"\u20b9{float(t.get('rate_per_cubic') or 0):,.0f}",
            f"\u20b9{float(t.get('price') or 0):,.2f}",
        ]
        cx = tbl_x + 6
        for j, (val, (_, cw)) in enumerate(zip(vals, cols)):
            c.setFont('Helvetica-Bold' if j == 5 else 'Helvetica', 8)
            c.setFillColor(BLACK)
            c.drawString(cx, y - row_h + 5, val)
            cx += cw
        y -= row_h

    # Totals row
    if y < 90:
        draw_footer(page_num)
        c.showPage()
        page_num += 1
        y = draw_header("CLIENT BILLING INVOICE (cont.)")

    y -= 8
    c.setFillColor(colors.HexColor('#1a1a1a'))
    c.rect(tbl_x, y - 70, W - 60, 70, fill=1, stroke=0)
    c.setStrokeColor(GOLD)
    c.setLineWidth(1)
    c.rect(tbl_x, y - 70, W - 60, 70, fill=0, stroke=1)
    c.setFont('Helvetica-Bold', 9)
    c.setFillColor(GOLD)
    c.drawString(tbl_x + 12, y - 18, f"Total Amount: \u20b9 {summary['total_amount']:,.2f}")
    c.setFillColor(GREEN)
    c.drawString(tbl_x + 12, y - 34, f"Advance Deducted: \u20b9 {summary['advance_deducted']:,.2f}")
    if summary['advance_carried_forward'] > 0:
        c.setFillColor(AMBER)
        c.drawString(tbl_x + 12, y - 50, f"Advance Carried Forward: \u20b9 {summary['advance_carried_forward']:,.2f}")
    else:
        c.setFillColor(RED)
        c.drawString(tbl_x + 12, y - 50, f"Balance Amount: \u20b9 {summary['balance_amount']:,.2f}")
    c.setFont('Helvetica', 7)
    c.setFillColor(colors.HexColor('#888888'))
    c.drawString(tbl_x + 12, y - 64, f"Generated: {dt2.now().strftime('%d %b %Y %H:%M')}  |  {len(transactions)} load(s)")

    draw_footer(page_num)
    c.save()
    buffer.seek(0)

    safe = re.sub(r'[^a-zA-Z0-9_-]', '_', client_name)[:24]
    fname = f"client_bill_{safe}_{invoice_number}.pdf"
    return buffer.getvalue(), fname


@app.route('/api/client-billing-summary', methods=['GET'])
def client_billing_summary():
    """Search unbilled loads for a client and return totals + advance deduction preview."""
    if 'username' not in session:
        return jsonify({"error": "Authentication required"}), 401
    if session.get('role') not in ['admin', 'manager']:
        return jsonify({"error": "Access denied"}), 403

    client = request.args.get('client', '').strip()
    start_date = request.args.get('start', '').strip()
    end_date = request.args.get('end', '').strip()
    include_billed = request.args.get('include_billed', '').lower() in ('1', 'true', 'yes')

    if not client:
        return jsonify({"error": "client name is required"}), 400

    try:
        summary = _compute_client_billing_summary(client, start_date, end_date, include_billed)
        return jsonify(summary)
    except Exception as e:
        logger.error(f"Client billing summary error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/client-billing-summary/history', methods=['GET'])
def client_billing_summary_history():
    """List previously generated client billing invoices."""
    if 'username' not in session:
        return jsonify({"error": "Authentication required"}), 401
    if session.get('role') not in ['admin', 'manager']:
        return jsonify({"error": "Access denied"}), 403

    client = request.args.get('client', '').strip()
    try:
        rows = _fetch_client_billing_summaries()
        if client:
            wanted = _normalize_client_name(client)
            rows = [r for r in rows if wanted in _normalize_client_name(r.get('client_name'))
                    or _normalize_client_name(r.get('client_name')) == wanted]
        rows.sort(key=lambda r: str(r.get('bill_date') or r.get('created_at') or ''), reverse=True)
        return jsonify({"success": True, "data": rows})
    except Exception as e:
        logger.error(f"Client billing history error: {e}")
        return jsonify({"error": str(e)}), 500


def _apply_advance_override(summary, advance_amount):
    """Recalculate deduction / balance / carry-forward from a manual advance amount."""
    total = float(summary.get('total_amount') or 0)
    advance = max(0.0, float(advance_amount or 0))
    summary['advance_amount'] = round(advance, 2)
    if advance > total:
        summary['advance_deducted'] = round(total, 2)
        summary['balance_amount'] = 0.0
        summary['advance_carried_forward'] = round(advance - total, 2)
        summary['display_balance'] = summary['advance_carried_forward']
        summary['balance_label'] = "Advance Carried Forward"
    else:
        summary['advance_deducted'] = round(advance, 2)
        summary['balance_amount'] = round(total - advance, 2)
        summary['advance_carried_forward'] = 0.0
        summary['display_balance'] = summary['balance_amount']
        summary['balance_label'] = "Balance Amount"
    return summary


@app.route('/api/client-billing-summary/confirm', methods=['POST'])
def client_billing_summary_confirm():
    """
    Confirm settlement for a client's unbilled loads:
    - Mark loads as billed (prevents duplicate billing)
    - Consume advances; create ADV-CF row if advance exceeds total
    - Generate invoice PDF and store path + metadata
    """
    if 'username' not in session:
        return jsonify({"error": "Authentication required"}), 401
    if session.get('role') not in ['admin', 'manager']:
        return jsonify({"error": "Access denied"}), 403

    data = request.json or {}
    if not data.get('confirm'):
        return jsonify({"error": "Confirmation required. Set confirm=true to settle."}), 400

    client = (data.get('client') or data.get('client_name') or '').strip()
    start_date = (data.get('start') or '').strip()
    end_date = (data.get('end') or '').strip()
    requested_ids = data.get('transaction_ids')  # optional subset
    advance_override = data.get('advance_amount', None)

    if not client:
        return jsonify({"error": "client name is required"}), 400

    try:
        summary = _compute_client_billing_summary(client, start_date, end_date, include_billed=False)
        if not summary.get('matched'):
            return jsonify({"error": summary.get('message', 'Client not found')}), 404

        txn_ids = list(summary.get('transaction_ids') or [])
        credit_ids = list(summary.get('advance_credit_ids') or [])

        if requested_ids is not None:
            requested_set = set(requested_ids)
            # Only allow settling currently unbilled ids from this summary
            txn_ids = [i for i in txn_ids if i in requested_set]
            # Recompute totals for the subset
            all_bills = _fetch_all_manual_billing()
            billed_ids = _collect_billed_transaction_ids()
            id_map = {b.get('id'): b for b in all_bills}
            subset = [id_map[i] for i in txn_ids if i in id_map and _is_unbilled_record(id_map[i], billed_ids)]
            if not subset and not credit_ids:
                return jsonify({"error": "No unbilled transactions selected"}), 400
            summary['transactions'] = []
            for idx, b in enumerate(subset, start=1):
                summary['transactions'].append({
                    "id": b.get('id'),
                    "load_no": idx,
                    "m_number": b.get('grade') or '—',
                    "delivery_date": b.get('date') or '',
                    "price": float(b.get('total_amount') or 0),
                    "cubic_meters": float(b.get('cubic_meters') or 0),
                    "rate_per_cubic": float(b.get('rate_per_cubic') or 0),
                })
            summary['transaction_ids'] = [t['id'] for t in summary['transactions']]
            summary['total_amount'] = round(sum(t['price'] for t in summary['transactions']), 2)
            txn_ids = summary['transaction_ids']

        # Manual advance from UI overrides sum of per-load advances
        try:
            if advance_override is not None:
                _apply_advance_override(summary, advance_override)
            else:
                _apply_advance_override(summary, summary.get('advance_amount', 0))
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid advance_amount"}), 400

        if not txn_ids and summary.get('advance_amount', 0) <= 0:
            return jsonify({"error": "No unbilled loads to settle for this client"}), 400

        # Re-verify none are already billed (race / duplicate guard)
        all_bills = _fetch_all_manual_billing()
        billed_ids = _collect_billed_transaction_ids()
        id_map = {b.get('id'): b for b in all_bills}
        already_billed = [i for i in txn_ids if i in id_map and not _is_unbilled_record(id_map[i], billed_ids)]
        if already_billed:
            return jsonify({
                "error": "Some loads were already billed. Refresh and try again.",
                "already_billed_ids": already_billed
            }), 409

        bill_date = datetime.now().strftime('%Y-%m-%d')
        invoice_number = f"CBS-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
        summary_id = f"cbs-{uuid.uuid4().hex[:10]}"

        pdf_bytes, pdf_filename = _build_client_billing_pdf(summary, invoice_number, bill_date)

        # Persist PDF to disk when possible
        pdf_path = None
        try:
            os.makedirs(BILLING_PDF_DIR, exist_ok=True)
            abs_path = os.path.join(BILLING_PDF_DIR, pdf_filename)
            with open(abs_path, 'wb') as f:
                f.write(pdf_bytes)
            pdf_path = abs_path
        except Exception as e:
            logger.warning(f"Could not save billing PDF to disk: {e}")
            pdf_path = f"memory:{pdf_filename}"

        import json as _json
        summary_record = {
            "id": summary_id,
            "invoice_number": invoice_number,
            "client_name": summary['client_name'],
            "bill_date": bill_date,
            "total_amount": summary['total_amount'],
            "advance_deducted": summary['advance_deducted'],
            "balance_amount": summary['balance_amount'],
            "advance_carried_forward": summary['advance_carried_forward'],
            "transaction_ids": _json.dumps(txn_ids + credit_ids),
            "pdf_path": pdf_path,
            "pdf_filename": pdf_filename,
            "notes": summary.get('balance_label'),
            "created_at": datetime.now().isoformat(),
            "created_by": session.get('username', 'system'),
        }
        _insert_client_billing_summary(summary_record)

        # Mark load rows as billed and zero their advances (consumed in this cycle)
        for tid in txn_ids:
            _update_manual_billing_record(tid, {
                "billing_status": "billed",
                "billing_summary_id": summary_id,
                "advance_amount": 0,
                "balance_amount": 0,
            })

        # Consume prior ADV-CF credit rows so advances aren't deducted twice
        for cid in credit_ids:
            _update_manual_billing_record(cid, {
                "billing_status": "billed",
                "billing_summary_id": summary_id,
                "advance_amount": 0,
                "balance_amount": 0,
            })

        # Carry forward leftover advance as a new unbilled credit row
        if summary['advance_carried_forward'] > 0:
            cf_record = {
                "id": f"bill-{uuid.uuid4().hex[:8]}",
                "company_name": summary['client_name'],
                "phone": summary.get('phone') or '',
                "address": summary.get('address') or '',
                "date": bill_date,
                "time": datetime.now().strftime('%I:%M %p'),
                "vehicle_number": "N/A",
                "driver_name": "System",
                "unloading_time": "N/A",
                "grade": "ADV-CF",
                "cubic_meters": 0,
                "loading_time": "N/A",
                "rate_per_cubic": 0,
                "total_amount": 0,
                "advance_amount": summary['advance_carried_forward'],
                "balance_amount": 0,
                "billing_status": "unbilled",
                "billing_summary_id": None,
                "created_at": datetime.now().isoformat(),
                "created_by": session.get('username', 'system'),
            }
            _insert_manual_billing_record(cf_record)

        # Keep PDF bytes in a short-lived memory cache keyed by summary id for immediate download
        if not hasattr(app, '_billing_pdf_cache'):
            app._billing_pdf_cache = {}
        app._billing_pdf_cache[summary_id] = pdf_bytes

        return jsonify({
            "success": True,
            "message": "Bill confirmed and generated successfully",
            "invoice": summary_record,
            "summary": {
                "client_name": summary['client_name'],
                "total_amount": summary['total_amount'],
                "advance_deducted": summary['advance_deducted'],
                "balance_amount": summary['balance_amount'],
                "advance_carried_forward": summary['advance_carried_forward'],
                "balance_label": summary['balance_label'],
                "loads_settled": len(txn_ids),
            },
            "pdf_url": f"/api/client-billing-summary/{summary_id}/pdf",
        })
    except Exception as e:
        logger.error(f"Client billing confirm error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/client-billing-summary/<summary_id>/pdf', methods=['GET'])
def client_billing_summary_pdf(summary_id):
    """Download a previously generated (or cached) client billing PDF."""
    if 'username' not in session:
        return jsonify({"error": "Authentication required"}), 401
    if session.get('role') not in ['admin', 'manager']:
        return jsonify({"error": "Access denied"}), 403

    try:
        # Memory cache (just-generated)
        cache = getattr(app, '_billing_pdf_cache', {})
        if summary_id in cache:
            rows = _fetch_client_billing_summaries()
            rec = next((r for r in rows if r.get('id') == summary_id), None)
            fname = (rec or {}).get('pdf_filename') or f"{summary_id}.pdf"
            resp = make_response(cache[summary_id])
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Content-Disposition'] = f'inline; filename={fname}'
            return resp

        rows = _fetch_client_billing_summaries()
        rec = next((r for r in rows if r.get('id') == summary_id), None)
        if not rec:
            return jsonify({"error": "Billing summary not found"}), 404

        pdf_path = rec.get('pdf_path') or ''
        fname = rec.get('pdf_filename') or f"{summary_id}.pdf"

        if pdf_path and not pdf_path.startswith('memory:') and os.path.exists(pdf_path):
            with open(pdf_path, 'rb') as f:
                pdf_bytes = f.read()
            resp = make_response(pdf_bytes)
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Content-Disposition'] = f'inline; filename={fname}'
            return resp

        # Fallback: regenerate from stored metadata + still-billed loads
        import json as _json
        try:
            txn_ids = _json.loads(rec.get('transaction_ids') or '[]')
        except Exception:
            txn_ids = []
        all_bills = _fetch_all_manual_billing()
        loads = [b for b in all_bills if b.get('id') in txn_ids and not _is_advance_carry_record(b)]
        loads.sort(key=lambda x: str(x.get('date', '')))
        transactions = []
        for idx, b in enumerate(loads, start=1):
            transactions.append({
                "id": b.get('id'),
                "load_no": idx,
                "m_number": b.get('grade') or '—',
                "delivery_date": b.get('date') or '',
                "price": float(b.get('total_amount') or 0),
                "cubic_meters": float(b.get('cubic_meters') or 0),
                "rate_per_cubic": float(b.get('rate_per_cubic') or 0),
            })
        regen = {
            "client_name": rec.get('client_name'),
            "phone": '',
            "address": '',
            "transactions": transactions,
            "total_amount": float(rec.get('total_amount') or 0),
            "advance_deducted": float(rec.get('advance_deducted') or 0),
            "balance_amount": float(rec.get('balance_amount') or 0),
            "advance_carried_forward": float(rec.get('advance_carried_forward') or 0),
            "display_balance": float(rec.get('advance_carried_forward') or 0) or float(rec.get('balance_amount') or 0),
            "balance_label": "Advance Carried Forward" if float(rec.get('advance_carried_forward') or 0) > 0 else "Balance Amount",
        }
        pdf_bytes, fname = _build_client_billing_pdf(
            regen, rec.get('invoice_number') or summary_id, rec.get('bill_date') or ''
        )
        resp = make_response(pdf_bytes)
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = f'inline; filename={fname}'
        return resp
    except Exception as e:
        logger.error(f"Client billing PDF error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
